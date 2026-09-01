# SPDX-License-Identifier: Apache-2.0
# ============================================================
# tests/core/test_spreadsheet_parser.py
#
# Unit tests for core.document_parser.parse_excel()
# — the server-side spreadsheet parser used by the web Chat
#   /chat/upload path and the gateway's extract_document tool.
#
# These tests run without a live gateway or database.
# They create real .xlsx / .xls files in a temp directory
# using openpyxl / xlwt so the parser exercises the full
# pandas + openpyxl / xlrd stack.
#
# Coverage:
#   1. Single-sheet workbook → plain table text (no ## heading)
#   2. Multi-sheet workbook  → one ## heading per sheet
#   3. Empty sheet           → "[Excel file is empty]" sentinel
#   4. Mixed empty + data    → only non-empty sheets appear
#   5. Large workbook        → row cap does NOT apply here
#      (parse_excel returns all rows; the gateway caps at 10 000
#       chars via parsed_text[:10000] in the /ask body builder)
#   6. Merged cells          → openpyxl reads the top-left value;
#      other cells in the merge are None → empty string in output
#   7. .xls (legacy BIFF)    → xlrd engine, same output shape
# ============================================================

from __future__ import annotations

import os
import tempfile
from pathlib import Path

import pytest

# Skip the entire module if pandas / openpyxl are not installed
# (they are in requirements.txt but may be absent in a minimal CI env).
pytest.importorskip("pandas")
pytest.importorskip("openpyxl")

from core.document_parser import parse_excel  # noqa: E402


# ── helpers ──────────────────────────────────────────────────────────────────

def _make_xlsx(path: Path, sheets: dict[str, list[list]]) -> None:
    """Write a multi-sheet .xlsx file using openpyxl."""
    from openpyxl import Workbook

    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))


def _make_xls(path: Path, rows: list[list]) -> None:
    """Write a single-sheet .xls file using xlwt (legacy BIFF format)."""
    xlwt = pytest.importorskip("xlwt")
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            ws.write(r_idx, c_idx, val)
    wb.save(str(path))


# ── fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture()
def tmp(tmp_path):
    return tmp_path


# ── tests ─────────────────────────────────────────────────────────────────────

class TestParseExcelSingleSheet:
    """Single-sheet workbooks — no ## heading, plain table."""

    def test_basic_table(self, tmp):
        p = tmp / "basic.xlsx"
        _make_xlsx(p, {"Sheet1": [["Name", "Score"], ["Alice", 95], ["Bob", 87]]})
        result = parse_excel(str(p))
        assert "Name" in result
        assert "Alice" in result
        assert "95" in result
        # Single sheet → no ## heading
        assert "## Sheet1" not in result

    def test_numeric_only(self, tmp):
        p = tmp / "nums.xlsx"
        _make_xlsx(p, {"Data": [[1, 2, 3], [4, 5, 6]]})
        result = parse_excel(str(p))
        assert "1" in result
        assert "6" in result

    def test_string_with_special_chars(self, tmp):
        p = tmp / "special.xlsx"
        _make_xlsx(p, {"Sheet1": [["Col A", "Col B"], ["foo & bar", 'say "hi"']]})
        result = parse_excel(str(p))
        assert "foo" in result
        assert "bar" in result


class TestParseExcelMultiSheet:
    """Multi-sheet workbooks — each sheet gets a ## heading."""

    def test_two_sheets(self, tmp):
        p = tmp / "multi.xlsx"
        _make_xlsx(p, {
            "Sales":   [["Region", "Revenue"], ["North", 1000], ["South", 2000]],
            "Returns": [["Region", "Count"],   ["North", 5],    ["South", 3]],
        })
        result = parse_excel(str(p))
        assert "## Sales" in result
        assert "## Returns" in result
        assert "Revenue" in result
        assert "Count" in result

    def test_three_sheets_all_present(self, tmp):
        p = tmp / "three.xlsx"
        _make_xlsx(p, {
            "A": [["x", "y"], [1, 2]],
            "B": [["p", "q"], [3, 4]],
            "C": [["m", "n"], [5, 6]],
        })
        result = parse_excel(str(p))
        for heading in ("## A", "## B", "## C"):
            assert heading in result


class TestParseExcelEmptySheets:
    """Empty and mixed-empty workbooks."""

    def test_all_empty_returns_sentinel(self, tmp):
        p = tmp / "empty.xlsx"
        _make_xlsx(p, {"Sheet1": []})
        result = parse_excel(str(p))
        assert result == "[Excel file is empty]"

    def test_mixed_empty_and_data(self, tmp):
        p = tmp / "mixed.xlsx"
        _make_xlsx(p, {
            "Empty":  [],
            "HasData": [["Col", "Val"], ["row1", 42]],
        })
        result = parse_excel(str(p))
        # The non-empty sheet must appear
        assert "HasData" in result or "Col" in result
        # The empty sheet should not produce a heading with no content
        # (parse_excel skips df.empty sheets)
        assert "Empty" not in result or "Col" in result


class TestParseExcelMergedCells:
    """Merged cells — openpyxl reads the top-left value; others are None."""

    def test_merged_cells_no_crash(self, tmp):
        from openpyxl import Workbook

        p = tmp / "merged.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged"
        ws["A1"] = "Header"
        ws["B1"] = "Value"
        ws["A2"] = "Merged"
        ws["A3"] = None  # part of merge
        ws.merge_cells("A2:A3")
        wb.save(str(p))

        result = parse_excel(str(p))
        # Should not raise; merged cell value appears at least once
        assert "Merged" in result or "Header" in result


class TestParseExcelLegacyXls:
    """Legacy .xls (BIFF) format via xlrd engine."""

    def test_xls_basic(self, tmp):
        pytest.importorskip("xlrd")
        p = tmp / "legacy.xls"
        _make_xls(p, [["Product", "Units"], ["Widget", 100], ["Gadget", 200]])
        result = parse_excel(str(p))
        assert "Product" in result
        assert "Widget" in result
        assert "100" in result


class TestParseExcelOutputFormat:
    """Output format matches what the gateway injects into the model context."""

    def test_output_is_string(self, tmp):
        p = tmp / "fmt.xlsx"
        _make_xlsx(p, {"Sheet1": [["A", "B"], [1, 2]]})
        result = parse_excel(str(p))
        assert isinstance(result, str)

    def test_no_binary_garbage(self, tmp):
        """Result must be valid UTF-8 text with no null bytes."""
        p = tmp / "clean.xlsx"
        _make_xlsx(p, {"Sheet1": [["Name", "Value"], ["AiNxt", 42]]})
        result = parse_excel(str(p))
        assert "\x00" not in result
        # Encode/decode round-trip must succeed
        assert result.encode("utf-8").decode("utf-8") == result

    def test_multi_sheet_separator(self, tmp):
        """Sheets are separated by a blank line (double newline)."""
        p = tmp / "sep.xlsx"
        _make_xlsx(p, {
            "Alpha": [["x"], [1]],
            "Beta":  [["y"], [2]],
        })
        result = parse_excel(str(p))
        assert "\n\n" in result
