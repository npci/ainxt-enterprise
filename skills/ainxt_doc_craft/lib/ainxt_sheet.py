# SPDX-License-Identifier: MIT
"""ainxt_sheet — brand-aware XLSX composition, preinstalled in the doc sandbox.

Exists so a model only has to express CONTENT. Number formats, the brand palette,
frozen headers, right-aligned numerics, brand chart colours and the mandatory
LibreOffice recalculation check are all enforced here.

    from ainxt_sheet import Book
    b = Book(title="UPI Settlement", classification="Confidential")
    d = b.sheet("Data", ["Bank", "Cycles", "Value (₹ Cr)", "Status"],
                widths=[30, 14, 18, 16], right_cols=[1, 2],
                formats={1: Book.COUNT, 2: Book.CRORE})
    d.rows([["Bank A", 1204, 8431, "Settled"], ["Bank B", 987, 6220, "Settled"]])
    d.total_row(["Total", "SUM", "SUM", ""])
    b.summary([("Total cycles", d.sum_formula(1), Book.COUNT),
               ("Total value (₹ Cr)", d.sum_formula(2), Book.CRORE)])
    b.chart(d, value_col=2, title="Value by bank")
    b.save()

Two rules this module enforces that a model reliably gets wrong on its own:
  * numeric cells stay numeric — presentation comes from number_format, never
    from pre-formatted strings, so the workbook actually sums;
  * summary formulas are bound to the real data rows, so an open-ended range
    cannot swallow a totals row and silently double-count.
"""
from __future__ import annotations

import glob
import os
import subprocess

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Brand tokens (INTERNAL_BRAND.md) ─────────────────────────────────────────
# Document identity. Empty by default so an adopter's workbooks carry their own
# brand or none, never this project's. Set DOC_BRAND_NAME when building the doc
# sandbox image (--build-arg): the container is run with no host environment
# passed in, so a runtime-only variable would never reach this code.
BRAND = os.getenv("DOC_BRAND_NAME", "")


def _brand_label(classification: str) -> str:
    return f"{BRAND} \u2014 {classification}" if BRAND else str(classification)


NAVY = "1F3864"
NAVY_TINT = "E8EDF6"
GREEN = "00A551"        # fills and chart series only
GREEN_TEXT = "00703C"   # 6.2:1 — the only green legal for 10pt cell text
AMBER = "C77700"
RED = "B3261E"
ROW_FILL = "F2F5FA"
RULE = "C9D2E3"
INK = "222222"
INK_MUTED = "5A6472"
PAPER = "FFFFFF"
FONT = "Arial"

_HAIRLINE = Border(bottom=Side(style="thin", color=RULE))
_HEAD_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
_ALT_FILL = PatternFill(start_color=ROW_FILL, end_color=ROW_FILL, fill_type="solid")
_TINT_FILL = PatternFill(start_color=NAVY_TINT, end_color=NAVY_TINT, fill_type="solid")


class Sheet:
    """One data sheet: header row 1, frozen, autofiltered, alternating body rows."""

    def __init__(self, ws, headers, widths=None, right_cols=(), formats=None):
        self.ws = ws
        self.headers = list(headers)
        self.right_cols = set(right_cols or ())
        self.formats = dict(formats or {})
        self.first_row = 2
        self.last_row = 1          # no body rows yet
        self._total_row = None

        for i, text in enumerate(self.headers, start=1):
            c = ws.cell(row=1, column=i, value=text)
            c.font = Font(name=FONT, size=10, bold=True, color=PAPER)
            c.fill = _HEAD_FILL
            c.alignment = Alignment(
                horizontal="right" if (i - 1) in self.right_cols else "left",
                vertical="center", wrap_text=True)
        for i, w in enumerate(widths or [], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False

    def rows(self, data):
        """Append body rows. Numbers stay numbers — never pre-format to strings."""
        for row in data:
            r = self.last_row + 1
            for i, v in enumerate(row, start=1):
                c = self.ws.cell(row=r, column=i, value=v)
                c.font = Font(name=FONT, size=10, color=INK)
                c.alignment = Alignment(
                    horizontal="right" if (i - 1) in self.right_cols else "left",
                    vertical="center")
                c.border = _HAIRLINE
                if r % 2 == 1:
                    c.fill = _ALT_FILL
                if (i - 1) in self.formats:
                    c.number_format = self.formats[i - 1]
            self.last_row = r
        self._finish()
        return self

    def total_row(self, spec):
        """Bold totals row. Use the string 'SUM' for a column to be summed."""
        r = self.last_row + 1
        for i, v in enumerate(spec, start=1):
            col = get_column_letter(i)
            val = (f"=SUM({col}{self.first_row}:{col}{self.last_row})"
                   if v == "SUM" else v)
            c = self.ws.cell(row=r, column=i, value=val)
            c.font = Font(name=FONT, size=10, bold=True, color=INK)
            c.alignment = Alignment(
                horizontal="right" if (i - 1) in self.right_cols else "left",
                vertical="center")
            c.border = _HAIRLINE
            if (i - 1) in self.formats:
                c.number_format = self.formats[i - 1]
        self._total_row = r
        self._finish(include_total=True)
        return self

    def _finish(self, include_total=False):
        last = self._total_row if include_total and self._total_row else self.last_row
        self.ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(self.headers))}{last}")

    def _range(self, col_idx):
        """Range covering ONLY the data rows — never the totals row."""
        col = get_column_letter(col_idx + 1)
        return f"{self.ws.title}!{col}{self.first_row}:{col}{self.last_row}"

    def sum_formula(self, col_idx):
        return f"=SUM({self._range(col_idx)})"

    def share_formula(self, value_col, match_col, match_value):
        """Guarded share-of-total, e.g. settled cycles ÷ all cycles."""
        v, m = self._range(value_col), self._range(match_col)
        return f'=IFERROR(SUMIFS({v},{m},"{match_value}")/SUM({v}),"")'


class Book:
    COUNT = "#,##0"
    CRORE = "#,##0;(#,##0)"
    MONEY = "₹#,##0"
    DEC = "#,##0.00"
    PCT = "0.0%"           # store 0.987, not 98.7
    DATE = "dd mmm yyyy"

    def __init__(self, title="Workbook", classification="Confidential",
                 out="/work/output.xlsx"):
        self.out = out
        self.title = title
        self.classification = classification
        self.wb = Workbook()
        self._summary = self.wb.active
        self._summary.title = "Summary"
        self._summary.sheet_view.showGridLines = False
        self._summary.column_dimensions["A"].width = 34
        self._summary.column_dimensions["B"].width = 20
        self._label("A1", title, size=14, bold=True, colour=NAVY)
        self._label("A2", _brand_label(classification), size=9)
        self._row = 4

    def _label(self, ref, text, size=10, bold=False, colour=INK_MUTED):
        c = self._summary[ref]
        c.value = text
        c.font = Font(name=FONT, size=size, bold=bold, color=colour)

    def sheet(self, name, headers, widths=None, right_cols=(), formats=None):
        ws = self.wb.create_sheet(name)
        return Sheet(ws, headers, widths, right_cols, formats)

    def summary(self, entries):
        """entries: [(label, formula_or_value, number_format)]"""
        self._label(f"A{self._row}", "Metric", bold=True, colour=NAVY)
        self._label(f"B{self._row}", "Value", bold=True, colour=NAVY)
        for ref in (f"A{self._row}", f"B{self._row}"):
            self._summary[ref].fill = _TINT_FILL
        self._row += 1
        for label, value, fmt in entries:
            self._label(f"A{self._row}", label, colour=INK)
            c = self._summary[f"B{self._row}"]
            c.value = value
            c.number_format = fmt
            # dark green = link to another sheet (financial-modelling convention)
            c.font = Font(name=FONT, size=10,
                          color=GREEN_TEXT if str(value).startswith("=") else INK)
            c.alignment = Alignment(horizontal="right")
            self._row += 1
        self._row += 1
        self._label(f"A{self._row}",
                    "Legend: navy = input · black = formula · dark green = link", size=9)
        self._row += 2
        return self

    def chart(self, sheet, value_col, title="", kind="bar", cat_col=0):
        """Brand-coloured native chart, anchored BELOW the summary block so it is
        not sliced by the print-page boundary on export."""
        ch = BarChart() if kind == "bar" else LineChart()
        if kind == "bar":
            ch.type = "col"
        ch.title = title or None
        ch.height, ch.width = 7.5, 16
        ch.add_data(Reference(sheet.ws, min_col=value_col + 1, min_row=1,
                              max_row=sheet.last_row), titles_from_data=True)
        ch.set_categories(Reference(sheet.ws, min_col=cat_col + 1,
                                    min_row=sheet.first_row, max_row=sheet.last_row))
        ch.y_axis.numFmt = self.CRORE
        # openpyxl defaults every series to stock Office blue.
        for s, colour in zip(ch.series, (NAVY, GREEN, INK_MUTED, AMBER)):
            s.graphicalProperties.solidFill = colour
            s.graphicalProperties.line.solidFill = colour
        if len(ch.series) == 1:
            ch.legend = None
        self._summary.add_chart(ch, f"A{self._row}")
        self._row += 18
        return self

    def save(self, recalc=True):
        self.wb.save(self.out)
        print(f"ainxt_sheet: wrote {self.out}")
        if recalc:
            print(self.recalc_check())
        return self

    def recalc_check(self):
        """openpyxl stores formulas without evaluating them, so a broken formula
        saves cleanly. LibreOffice recalculates on load; we read the cached values.
        Never reports 'none' unless a recalculated workbook actually existed."""
        errors = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
        produced, reason = [], None
        try:
            subprocess.run(
                ["soffice", "--headless", "-env:UserInstallation=file:///tmp/lo_recalc",
                 "--convert-to", "xlsx", "--outdir", "/work/recalc", self.out],
                check=False, timeout=180,
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            produced = glob.glob("/work/recalc/*.xlsx")
            if not produced:
                reason = "recalculation produced no workbook"
        except FileNotFoundError:
            # check=False does not cover a missing binary; Popen raises first.
            reason = "soffice not found"
        except (OSError, subprocess.TimeoutExpired) as exc:
            reason = f"recalculation failed ({exc})"

        if not produced:
            return f"FORMULA CHECK: DID NOT RUN — {reason}; formulas UNVERIFIED"
        found = []
        for path in produced:
            chk = load_workbook(path, data_only=True)
            for ws in chk.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value in errors:
                            found.append(f"{ws.title}!{cell.coordinate} = {cell.value}")
        return f"FORMULA ERRORS: {found if found else 'none'}"
