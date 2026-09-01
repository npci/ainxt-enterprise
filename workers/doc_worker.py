# SPDX-License-Identifier: Apache-2.0
# ============================================================
# DOC WORKER — rq job that generates a document file and
# stores the result in Redis + Postgres audit record.
#
# Flow:
#   1. Compliance gate on content_md
#   2. Call _skill_generate() — uses the ainxt_doc_craft SKILL.md + SKELETON +
#      AiNxt_BRAND.md injected as LLM system context; the LLM writes code that uses
#      the doc composition wrappers, executed in the doc sandbox via
#      sandbox.doc_executor.build() (docx/pptx/xlsx/pdf; pdf is authored as docx
#      and exported). CSV uses the stdlib csv path.
#   3. Write binary to /tmp/ainxt_docs/{file_id}.{ext}
#   4. Save audit record to generated_documents (Postgres)
#   5. Publish {status, file_id, filename} to Redis key doc:result:{job_id}
#
# NOTE: tools.doc_generator.generate() (old implementation) is DISABLED.
#       All generation now routes through _skill_generate() which delegates
#       to the ainxt_doc_craft skillset in skills/ainxt_doc_craft/ (in-house).
#       The previous silent-fallback to tools.doc_generator on skill failure
#       has been removed — _skill_generate() now hard-fails the job instead
#       of silently producing an off-brand document. This guarantees every
#       generated file follows the ainxt_doc_craft brand template.
# ============================================================

import base64
import json
import os
import re
import sys
import tempfile
import time
import uuid as _uuid_mod

from core.config import RDB_STREAM, DOC_STORAGE_DIR, user_doc_dir
from core.kv import get_kv
from core.logger import logger

# PERF: hoisted from the 3 function-local imports this module used to have
# (_safe_log, the question-mode compliance gate, the MD-session compliance
# gate). RQ forks a fresh work-horse process per job, so a function-local
# `from agents.compliance_engine import ...` paid for a genuine cold module
# import (config file read + engine init, logged as "ComplianceEngine: loaded
# config from ...") on EVERY job, not just once per process. A module-level
# import here pays that cost once when the worker process starts (or once per
# fork-parent, if the parent already imported doc_worker), not per job.
from agents.compliance_engine import compliance_engine

# DB=6 — document result delivery. Backend selected via REDIS_CLIENT_CONFIG_DB6.
_R = get_kv(RDB_STREAM, decode_responses=True)


_PCI_FALLBACK_RE = re.compile(
    r"\b(?:\d[ -]*?){13,19}\b"                       # card / long digit runs (PAN)
    r"|\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b"  # emails
    r"|\b[A-Z]{5}\d{4}[A-Z]\b"                       # PAN (Indian tax id)
)


def _safe_log(value, limit: int = 120) -> str:
    """Return a redacted, length-capped repr of user/LLM-derived text for logs."""
    try:
        text = "" if value is None else str(value)
        if not text:
            return "''"
        clipped = text[:limit]
        try:
            _r = compliance_engine.validate_input(clipped)
            redacted = _r.get("redacted_text") or clipped
        except Exception:
            redacted = clipped
        redacted = _PCI_FALLBACK_RE.sub("[REDACTED]", redacted)
        return repr(redacted)
    except Exception:
        return "'[unloggable]'"

# Persistent storage (see core.config.DOC_STORAGE_DIR). NOT /tmp — files must
# survive container restart so refresh-then-download keeps working.
DOC_DIR    = DOC_STORAGE_DIR
os.makedirs(DOC_DIR, exist_ok=True)
RESULT_TTL = 86400  # 24 h — Redis result TTL only; binary lives forever in DOC_DIR/Postgres

# ── Image provider config ─────────────────────────────────────────────────────
# "auto"    → try gemini → dalle → disabled
# "gemini"  → Gemini Imagen 3
# "dalle"   → DALL-E 3 via OpenAI
# "disabled"→ skip image generation (pure geometric design)
_PPT_IMG_PROVIDER = os.getenv("PPT_IMAGE_PROVIDER", "auto")

# ── PPT image-coverage config ─────────────────────────────────────────────────
# When false (default), only the cover (slide_type == "title") gets an
# AI-generated image; other slides fall back to geometric design. This keeps
# deck generation cheap and fast.
# Set PPT_IMAGES_ALL_SLIDES=true to restore image generation on every
# image-eligible slide (title + content + closing).
_PPT_IMAGES_ALL_SLIDES = os.getenv(
    "PPT_IMAGES_ALL_SLIDES", "false"
).strip().lower() in {"1", "true", "yes", "on"}

# ── PPTX engine selector ─────────────────────────────────────────────────────
# "sandbox" → route PPTX through sandbox.doc_executor.build()
#             (Dockerised pptxgenjs + page-image previews + sandbox-side imagen)
# "native"  → keep current path: _skill_generate() → _run_in_doc_sandbox()
# "auto"    → sandbox if docker_available() and image_present(); else native
_DOC_PPTX_ENGINE = os.getenv("DOC_PPTX_ENGINE", "auto").lower().strip()

_PPTX_FMT_ALIASES = {"pptx", "ppt", "powerpoint", "presentation", "slides"}


# ── Summary + preview helper (shared across all publish sites) ────────────────
def _attach_summary_preview(
    payload: dict,
    *,
    title: str,
    sections: list,
    question: str,
    chat_id: str | None,
    job_id: str,
) -> tuple[int, float]:
    """
    Build summary + preview and merge into the payload's `summary`, `preview`,
    and `meta.summary_*` keys. Returns (summary_tokens, summary_cost) so the
    caller can update budget accounting. Never raises — failure leaves the
    payload untouched.
    """
    try:
        from agents.doc_generator_agent import build_summary_and_preview
        summary, preview, smeta = build_summary_and_preview(
            title=title or "Document",
            sections=sections or [],
            prompt=question or "",
            chat_id=chat_id,
        )
        payload["summary"] = summary
        payload["preview"] = preview
        meta = payload.setdefault("meta", {})
        meta["summary_tokens"] = int(smeta.get("tokens") or 0)
        meta["summary_cost"]   = float(smeta.get("cost_usd") or 0.0)
        meta["summary_source"] = smeta.get("source")
        return (
            int(smeta.get("tokens") or 0),
            float(smeta.get("cost_usd") or 0.0),
        )
    except Exception as exc:
        logger.warning(
            f"[docgen] worker summary/preview attach failed | job={job_id} error={exc}"
        )
        return (0, 0.0)


# ── Atomic file write ────────────────────────────────────────────────────────
def _atomic_write_bytes(final_path: str, data: bytes) -> None:
    """Write ``data`` to ``final_path`` atomically.

    Sequence:
      1. Write the bytes to ``{final_path}.partial``.
      2. ``fh.flush()`` + ``os.fsync(fd)`` so the OS commits the write to
         the underlying device before any reader sees the final filename.
      3. ``os.replace(.partial, final_path)`` — atomic rename on POSIX
         and a same-volume Windows rename. Readers either see the old file
         (or no file) before, and the complete file after; never a partial
         file under the final name.

    Fixes the prod symptom where the first download attempt after generation
    returned a truncated 1 KB blob and a refresh returned the full file.
    Likely cause was an NFS / shared-volume visibility race between the
    writer instance and the reader instance: the file existed at its final
    name with size 0 (or partial) by the time the FastAPI ``FileResponse``
    started streaming on the other instance.
    """
    partial_path = f"{final_path}.partial"
    with open(partial_path, "wb") as fh:
        fh.write(data)
        fh.flush()
        try:
            os.fsync(fh.fileno())
        except OSError:
            # fsync is unsupported on some filesystems (e.g. /tmp on tmpfs
            # in containers). The rename is still atomic; we just lose the
            # explicit flush guarantee. Don't fail the job over this.
            pass
    os.replace(partial_path, final_path)


def _atomic_write_text(final_path: str, text: str, encoding: str = "utf-8") -> None:
    """UTF-8 text variant of :func:`_atomic_write_bytes`.

    Encodes the string up front so the write itself is a single bytes I/O,
    matching the binary path's atomicity guarantees. Used for ``.md`` outputs.
    """
    _atomic_write_bytes(final_path, text.encode(encoding))


# ── AiNxt doc-craft skillset paths ────────────────────────────────────────────
# Root of the ainxt_doc_craft folder (relative to this repo). The craft guidance
# here is authored in-house and drives the agent to build documents with the
# composition wrappers (ainxt_sheet.Book / ainxt-doc / ainxt-deck) that are
# preinstalled in the doc sandbox image (see docker/doc-sandbox/Dockerfile).
_SKILLS_ROOT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "skills", "ainxt_doc_craft",
)

# Brand file path — configurable via DOC_BRAND_FILE env var.
# Default: brand/BRAND.md (generic OSS brand — no org-specific colours or fonts).
# Internal/enterprise: set DOC_BRAND_FILE=brand/INTERNAL_BRAND.md in your .env
# Other orgs: create brand/<YourOrg>_BRAND.md and set DOC_BRAND_FILE accordingly.
# Path is relative to _SKILLS_ROOT (skills/ainxt_doc_craft/).
_DOC_BRAND_FILE = os.getenv("DOC_BRAND_FILE", "brand/BRAND.md")

# Format → sub-folder name inside _SKILLS_ROOT.
# NOTE: ainxt_doc_craft has NO pdf/ folder — PDF is authored exactly like a Word
# document (via ainxt-doc) and exported to PDF, so it reuses the docx skill.
_SKILL_FOLDER = {
    "docx": "docx", "doc": "docx", "word": "docx",
    "pptx": "pptx", "ppt": "pptx", "powerpoint": "pptx",
    "presentation": "pptx", "slides": "pptx",
    "pdf":  "docx",   # PDF reuses the docx (ainxt-doc) skill, then exports to PDF
    "xlsx": "xlsx", "xls": "xlsx", "excel": "xlsx",
    "csv":  "xlsx",   # CSV generation uses xlsx skill, outputs csv
}

# MIME types (replaces tools.doc_generator.MIME_TYPES)
MIME_TYPES = {
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "pdf":  "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv":  "text/csv",
    "txt":  "text/plain",
    "md":   "text/markdown",
}

# PPTX themes (replaces tools.doc_generator.PPTX_THEMES)
PPTX_THEMES = {
    "dark_executive": {"id": "dark_executive",  "label": "Dark Executive",  "description": "Navy & gold — boardroom authority"},
    "light_modern":   {"id": "light_modern",    "label": "Light Modern",    "description": "Clean white — contemporary clarity"},
    "vibrant_tech":   {"id": "vibrant_tech",    "label": "Vibrant Tech",    "description": "Bold gradients — innovation energy"},
}

# Per-format defensive reminders appended to the SKILL.md context in the
# code-generation prompt. These exist ONLY to surface wrapper-library pitfalls
# that the SKILL.md / SKELETON do not already cover, or that the LLM tends to
# ignore in practice. Each rule must NOT contradict its SKILL.md.
#
# The generated code uses the ainxt composition wrappers preinstalled in the doc
# sandbox (ainxt_sheet.Book / ainxt-doc / ainxt-deck), NOT raw openpyxl/docx-js/
# pptxgenjs. Those wrappers already enforce the brand palette, fonts, headers,
# number formats and (for xlsx) the LibreOffice recalculation check, and each
# writes to its default /work/output.<ext> on .save() — which is exactly where
# the sandbox reads the result from, so OUTPUT_PATH is not needed.
_FMT_TO_RULES = {
    "xlsx": (
        "Additional reminders for this run (do not override SKILL.md — extend it):\n"
        "1. Use the wrapper only: `from ainxt_sheet import Book`. Do NOT hand-roll openpyxl "
        "styling, fills, fonts or borders — the wrapper applies the brand. Do NOT `import "
        "pandas` or `faker` (neither is guaranteed in the sandbox).\n"
        "2. Build the workbook as in the SKELETON: `b = Book(title=..., classification='Confidential')`, "
        "then `s = b.sheet(name, headers, widths=[...], right_cols=[...], formats={col: Book.FMT})`, "
        "then `s.rows([...])`. Numbers must stay Python int/float — never pre-formatted strings — so "
        "cells actually sum; presentation comes from the `formats` map (Book.COUNT / CRORE / MONEY / "
        "DEC / PCT / DATE).\n"
        "3. Derived cells should be FORMULAS, not Python-computed constants: use `s.total_row([...])` "
        "with the string 'SUM' for summed columns, and `b.summary([(label, s.sum_formula(col), "
        "Book.FMT), ...])` for the Summary sheet. `sum_formula`/`share_formula` are bound to the real "
        "data rows so a totals row cannot be double-counted.\n"
        "4. Optional native chart: `b.chart(s, value_col=<0-based>, title=...)`. No pasted images.\n"
        "5. Finish with `b.save()` — it writes /work/output.xlsx and runs the recalculation check "
        "for you. Do NOT pass a path and do NOT call openpyxl's `wb.save` yourself.\n"
        "6. Sheet names must be ≤31 chars and must not contain any of: \\ / ? * [ ] :\n"
    ),
    # Bulk TABULAR dataset written to .xlsx (e.g. "500 employee records, tabular
    # only, no descriptions"). This is the openpyxl analogue of the CSV
    # test-data path: a single flat sheet of programmatically generated rows, NOT
    # the narrative multi-sheet "sections + small table" workbook. Used only when
    # the structuring step produced a `csv_schema` section (columns + row_count).
    "xlsx_testdata": (
        "TABULAR DATASET MODE (this run produces a FLAT data sheet, NOT a "
        "narrative report). Follow these rules strictly:\n"
        "1. Use the wrapper: `from ainxt_sheet import Book`. Create the book, then ONE data sheet, "
        "then append all rows, then save:\n"
        "       b = Book(title='<title>', classification='Confidential')\n"
        "       s = b.sheet('Data', HEADERS, widths=[...])   # HEADERS = exact columns, in order\n"
        "       s.rows(all_rows)                              # all_rows = list of row lists\n"
        "       b.save(recalc=False)                          # no formulas → skip recalc pass\n"
        "   Do NOT hand-roll openpyxl, and do NOT `import pandas` or `faker` (neither is guaranteed "
        "in the sandbox; the import will raise). Numbers must stay numeric (int/float), not strings.\n"
        "2. ROW COUNT: generate EXACTLY the requested number of data rows — even large counts like "
        "500, 10000, or 50000. Do NOT cap, sample, round down, or emit a placeholder subset. Build "
        "`all_rows` with a loop `for i in range(ROW_COUNT):` and pass it once to `s.rows(all_rows)`.\n"
        "3. COLUMNS: emit EVERY requested column for every row, in the given order. Do NOT drop "
        "columns, do NOT add extra columns, do NOT collapse them into a label + metrics shape.\n"
        "4. NO narrative content: this is header + data only. Do NOT add a summary, total_row, "
        "chart, or extra sheet for this mode — just the one Data sheet.\n"
        "5. Realistic values by column intent — define small curated tuples inline and pick with "
        "`random.choice(...)`, e.g.:\n"
        "       FIRST_NAMES = ('Aarav','Vihaan','Aditya','Ananya','Diya','Priya',"
        "'Rahul','Neha','Arjun','Kavya','Rohan','Riya','Ishaan','Saanvi')\n"
        "       LAST_NAMES  = ('Sharma','Verma','Iyer','Patel','Reddy','Khan',"
        "'Gupta','Singh','Nair','Mehta','Rao','Joshi','Das','Bose','Shah')\n"
        "       CITIES      = ('Mumbai','Bengaluru','Delhi','Chennai','Hyderabad',"
        "'Pune','Kolkata','Ahmedabad','Jaipur','Lucknow')\n"
        "   IDs: zero-padded sequences or uuid4. Dates: `datetime` objects or ISO `YYYY-MM-DD` "
        "strings. Salary/amount: plain numbers (no currency symbol) so the cell stays numeric.\n"
    ),
    "docx": (
        "Additional reminders for this run (do not override SKILL.md / SKELETON — extend them):\n"
        "1. Use the wrapper only: `const doc = require('ainxt-doc');` then "
        "`const d = doc.create({ title: '<title>', subtitle: '<optional>', date: '<optional>', "
        "classification: 'Confidential' });`. Do NOT require('docx') directly or hand-build "
        "Document/Paragraph/TextRun — the wrapper applies the brand, footer and page setup.\n"
        "2. Express CONTENT with the chainable methods: `d.h1(text)`, `d.h2(text)`, `d.h3(text)`, "
        "`d.p(text)`, `d.bullet(text)`, `d.step(text)` (numbered), `d.caption(text)`, "
        "`d.table(header, rows, { pct:[...], rightCols:[...] })`, `d.pageBreak()`. Pass plain "
        "strings — the wrapper wraps them in the correct docx objects. Table rows are arrays of "
        "plain strings; the wrapper handles cell construction.\n"
        "3. CODE BLOCKS. If a section has a non-empty `code` field, render each line as its own "
        "`d.p(line)` (or a monospace-styled paragraph) AFTER the section's prose/bullets, "
        "preserving indentation — do NOT reflow, summarize, or escape it.\n"
        "4. Finish with `d.save()` — it returns a Promise that writes /work/output.docx. Do NOT "
        "pass a path. The worker exports the .docx to PDF afterwards when the requested format is "
        "PDF, so author it exactly as a Word document either way.\n"
    ),
    "pptx": (
        "Additional reminders for this run (do not override SKILL.md / SKELETON — extend them):\n"
        "1. Use the wrapper only: `const deck = require('ainxt-deck');` then "
        "`const d = deck.create({ classification: 'Confidential', title: '<title>' });`. Do NOT "
        "require('pptxgenjs') directly — the wrapper applies layout, brand colours, logo, footer "
        "and accessibility rules.\n"
        "2. Build the deck from the slide-pattern methods (each returns `this`, so chain or call "
        "in sequence): `d.cover(title, subtitle, date)`, `d.contents(heading, [labels])`, "
        "`d.evidence(heading, [bullets], visual)`, `d.split(heading, [{title,bullets}, ...])`, "
        "`d.metric(heading, [{figure,label,status:'good'|'warn'|'bad'}], note)`, "
        "`d.statement(sentence, attribution)`, `d.table(heading, header, rows, {rightCols:[...]})`, "
        "`d.close(line, [nextSteps])`, `d.notes(text)`. Pass plain strings/numbers; the wrapper "
        "positions everything. Do NOT compute x/y/w/h yourself.\n"
        "3. Vary the pattern between consecutive slides (two identical layouts in a row is the "
        "top 'AI deck' tell). A typical deck is cover → contents → a mix of evidence/split/metric/"
        "table/statement → close.\n"
        "4. Finish with `d.save()` — it returns a Promise that writes /work/output.pptx. Do NOT "
        "pass a path and do NOT assume it finished synchronously.\n"
    ),
    "csv": (
        "CSV-specific rules (this is a FLAT data file, not an Excel workbook):\n"
        "1. Write the file using the stdlib `csv` module — it is ALWAYS available "
        "and avoids pandas/numpy import overhead:\n"
        "       import csv\n"
        "       with open(OUTPUT_PATH, 'w', newline='', encoding='utf-8') as fh:\n"
        "           w = csv.writer(fh)\n"
        "           w.writerow(HEADERS)\n"
        "           w.writerows(rows)\n"
        "   Do NOT use openpyxl (that produces .xlsx). Do NOT use pandas — many "
        "execution environments lack it and a flat CSV needs no DataFrame.\n"
        "2. Single header row (column names) + N data rows. No empty leading rows, "
        "no merged cells, no formulas, no multiple sheets — those are XLSX-only.\n"
        "3. Use ONLY stdlib for value generation — `random`, `random.choices`, "
        "`string`, `datetime`/`timedelta`, `uuid`. Do NOT `import faker` (it is "
        "not installed in the sandbox; the import will raise ModuleNotFoundError).\n"
        "4. For realistic-looking test data, define small curated tuples inline, e.g.:\n"
        "       FIRST_NAMES = ('Aarav','Vihaan','Aditya','Ananya','Diya','Priya',"
        "'Rahul','Neha','Arjun','Kavya','Rohan','Riya','Ishaan','Saanvi','Aryan',"
        "'Myra','Reyansh','Anika','Kabir','Zara')\n"
        "       LAST_NAMES  = ('Sharma','Verma','Iyer','Patel','Reddy','Khan','Gupta',"
        "'Singh','Nair','Mehta','Rao','Joshi','Das','Bose','Shah')\n"
        "       CITIES      = ('Mumbai','Bengaluru','Delhi','Chennai','Hyderabad',"
        "'Pune','Kolkata','Ahmedabad','Jaipur','Lucknow')\n"
        "       EMAIL_DOM   = ('gmail.com','outlook.com','yahoo.com','example.com')\n"
        "   Combine via `random.choice(...)`. Emails should look like "
        "`f\"{first.lower()}.{last.lower()}{random.randint(1,999)}@{domain}\"`.\n"
        "5. Row count: if the user gave an explicit number, honour it EXACTLY — "
        "even large counts like 10000 or 50000 (do NOT cap it at 2000; loop and "
        "stream rows rather than building giant in-memory lists where possible). "
        "Otherwise pick a sensible count between 50 and 2000 based on the topic "
        "(test/synthetic/sample data → 200–1000; demo/preview → 50–200).\n"
        "6. Domain-appropriate columns: payments → amount/currency/txn_id; "
        "users → user_id/name/email/signup_date; orders → order_id/sku/qty/price; "
        "transactions → txn_id/timestamp/amount/status. Match the user's request.\n"
        "7. Dates: use ISO format `YYYY-MM-DD`. Timestamps: `YYYY-MM-DD HH:MM:SS`. "
        "Currencies: plain numeric (e.g. `1245.60`) — no symbols or commas inside the "
        "value (CSV consumers split on `,`).\n"
        "8. If the prompt contains a USER-UPLOADED TEMPLATE block, mirror those "
        "exact column names and types — do not invent extra columns, do not drop "
        "any. Continue any obvious sequences (e.g. order_id) from the last value.\n"
    ),
    # PDF is authored exactly like a Word document with the ainxt-doc wrapper and
    # then exported to PDF by the worker — far better looking than reportlab. So
    # the PDF rules are the SAME as the docx rules; keep them in sync.
    "pdf": (
        "Additional reminders for this run (do not override SKILL.md / SKELETON — extend them):\n"
        "1. Author this PDF as a Word document using the wrapper: "
        "`const doc = require('ainxt-doc');` then "
        "`const d = doc.create({ title: '<title>', subtitle: '<optional>', date: '<optional>', "
        "classification: 'Confidential' });`. Do NOT use reportlab, and do NOT require('docx') "
        "directly — the worker exports the .docx the wrapper writes into a polished PDF.\n"
        "2. Express CONTENT with the chainable methods: `d.h1/h2/h3(text)`, `d.p(text)`, "
        "`d.bullet(text)`, `d.step(text)`, `d.caption(text)`, "
        "`d.table(header, rows, { pct:[...], rightCols:[...] })`, `d.pageBreak()`. Pass plain "
        "strings; the wrapper builds the docx objects.\n"
        "3. CODE BLOCKS. If a section has a non-empty `code` field, render each line as its own "
        "paragraph AFTER the section's prose, preserving indentation — do NOT reflow or escape it.\n"
        "4. Finish with `d.save()` — it returns a Promise that writes /work/output.docx (the "
        "worker converts it to PDF). Do NOT pass a path.\n"
    ),
}


# ── Document-sandbox Docker image ────────────────────────────────────────────
# All generated docx/pptx/xlsx/pdf code runs inside this container, which
# ships with docx-js, pptxgenjs, python-docx, openpyxl, LibreOffice, etc.
# pre-installed globally. See docker/doc-sandbox/Dockerfile.
#
# Override the image tag via DOC_SANDBOX_IMAGE if you tag it differently.
_DOC_SANDBOX_IMAGE   = os.getenv("DOC_SANDBOX_IMAGE", "ainxt-doc-sandbox:latest")
_DOC_SANDBOX_TIMEOUT = int(os.getenv("DOC_SANDBOX_TIMEOUT", "120"))
_DOC_SANDBOX_MEMORY  = os.getenv("DOC_SANDBOX_MEMORY", "1g")
_DOC_SANDBOX_CPUS    = os.getenv("DOC_SANDBOX_CPUS", "1.0")

# ── Content-preservation char limits (env-tunable) ───────────────────────────
# Used by the verbatim chat-preservation path so long, code-heavy replies survive
# intact. The default doc model is "complex" (Claude Sonnet, large context window),
# so these caps leave generous headroom while still guarding against pathological
# payloads blowing the LLM context window or the sandbox timeout.
_CHAT_PRESERVE_MAX_CHARS = int(os.getenv("DOC_CHAT_PRESERVE_MAX_CHARS", "40000"))
_SECTIONS_JSON_MAX       = int(os.getenv("DOC_SECTIONS_JSON_MAX", "60000"))

# Self-healing code generation: total number of sandbox execution attempts.
# Attempt 1 is the initial LLM-written code; each subsequent attempt feeds the
# exact stderr + broken source back to the LLM to repair it. This is what makes
# a transient "SyntaxError: Unexpected identifier" recoverable instead of fatal.
_CODE_MAX_ATTEMPTS = max(1, int(os.getenv("DOC_CODE_MAX_ATTEMPTS", "8")))

# HARD RULE (per product requirement): document generation happens ONLY through
# the ainxt_doc_craft sandbox path. There is NO fallback of any kind (no
# native in-process render, no legacy tools.doc_generator). Instead the sandbox
# is given a GENEROUS total budget — up to ~30 minutes across self-repair
# attempts — so a transient/slow build has every chance to succeed via the
# skill. If the sandbox still cannot produce a skill-generated file within the
# budget, the job fails with a retry error (never an inferior document).
#
# Total wall-clock ceiling for the whole skill-generation loop, in seconds.
# The per-attempt sandbox timeout is derived from this so all attempts together
# fit inside the ceiling.
_DOC_TOTAL_BUDGET_SEC = max(60, int(os.getenv("DOC_TOTAL_BUDGET_SEC", "1800")))  # 30 min default


# Defensive pptxgenjs shim injected before LLM-generated PPTX code (Fix #36).
# The LLM sometimes passes a bare string where pptxgenjs expects a text object or
# an {x,y,w,h} options object — e.g. `slide.addText("Manikanda Sakthi")` with the
# author name as the SECOND arg, which pptxgenjs then tries to set `.options` on,
# throwing `TypeError: Cannot create property 'options' on string`. This monkey-
# patches addText/addTable/addImage to coerce loose inputs into valid shapes and
# fill in safe default coordinates, so a minor LLM slip degrades to a rendered
# slide instead of a hard crash. Harmless for correct code.
_PPTX_SHIM_JS = r"""
;(function(){
  try {
    var _req = require;
    var P = _req('pptxgenjs');
    var proto = (P && (P.prototype)) ? P.prototype : null;
    // pptxgenjs slides are created per-call; patch at the Slide level lazily.
    var origAddSlide = proto && proto.addSlide;
    function _defaultOpts(o){
      o = (o && typeof o === 'object' && !Array.isArray(o)) ? o : {};
      if (typeof o.x !== 'number') o.x = 0.5;
      if (typeof o.y !== 'number') o.y = 0.5;
      if (typeof o.w !== 'number') o.w = 9.0;
      if (typeof o.h !== 'number') o.h = 1.0;
      return o;
    }
    function _patchSlide(slide){
      if (!slide || slide.__ainxtPatched) return slide;
      slide.__ainxtPatched = true;
      var _addText = slide.addText ? slide.addText.bind(slide) : null;
      if (_addText) {
        slide.addText = function(text, opts){
          // text may be a string, an array of runs, or a single run object.
          if (typeof text === 'string') text = [{ text: text }];
          else if (text && typeof text === 'object' && !Array.isArray(text)) {
            if (typeof text.text !== 'undefined') text = [text];
          }
          // If someone passed the options as the "text" arg and a string as opts.
          if (typeof opts === 'string') opts = { };
          return _addText(text, _defaultOpts(opts));
        };
      }
      var _addTable = slide.addTable ? slide.addTable.bind(slide) : null;
      if (_addTable) {
        slide.addTable = function(rows, opts){
          if (!Array.isArray(rows)) rows = [[ String(rows == null ? '' : rows) ]];
          rows = rows.map(function(r){ return Array.isArray(r) ? r : [r]; });
          return _addTable(rows, _defaultOpts(opts));
        };
      }
      // addImage / addShape / addChart: coerce a missing/loose opts object and
      // SWALLOW a bad single element (skip it) rather than crash the whole deck —
      // a slightly malformed image/shape/chart should degrade gracefully. (G3)
      ["addImage", "addShape", "addChart"].forEach(function(fn){
        if (typeof slide[fn] !== "function") return;
        var orig = slide[fn].bind(slide);
        slide[fn] = function(){
          try {
            var a = Array.prototype.slice.call(arguments);
            // Ensure the trailing opts arg is an object with default geometry.
            if (a.length && typeof a[a.length - 1] === "object") {
              a[a.length - 1] = _defaultOpts(a[a.length - 1]);
            } else {
              a.push(_defaultOpts({}));
            }
            return orig.apply(null, a);
          } catch (e) {
            try { console.error("[ainxt-shim] skipped bad " + fn + ": " + (e && e.message)); } catch (_e) {}
            return slide; // skip the element, keep building
          }
        };
      });
      return slide;
    }
    if (proto && origAddSlide) {
      proto.addSlide = function(){
        var s = origAddSlide.apply(this, arguments);
        return _patchSlide(s);
      };
    }
  } catch (e) {
    // If pptxgenjs isn't the target lib (docx/xlsx runs), silently ignore.
  }
})();
"""


def _run_in_doc_sandbox(
    *,
    lang: str,
    code: str,
    out_filename: str,
    workdir_host: str,
    job_id: str = "",
    timeout: int | None = None,
) -> tuple[int, bytes, str]:
    """
    Execute generated skill code (JavaScript or Python) inside the
    ainxt-doc-sandbox Docker container. The container has docx-js,
    pptxgenjs, python-docx, openpyxl, LibreOffice, etc. pre-installed —
    so no `npm install` / `pip install` happens at request time.

    Args:
        lang          : "javascript" or "python"
        code          : full source to execute (already includes OUTPUT_PATH)
        out_filename  : basename of the file the script will write to /work
        workdir_host  : host directory mounted at /work inside the container;
                        the code file is written here, the output is read here
        job_id        : for log correlation
        timeout       : seconds; defaults to DOC_SANDBOX_TIMEOUT

    Returns:
        (returncode, output_bytes_or_empty, stderr_text)
        output_bytes is non-empty only on success.
    """
    import shutil as _sh
    import subprocess as _sp

    docker_bin = _sh.which("docker")
    if not docker_bin:
        return (127, b"", "docker binary not found on PATH")

    script_name   = "gen.js" if lang == "javascript" else "gen.py"
    script_host   = os.path.join(workdir_host, script_name)
    out_host      = os.path.join(workdir_host, out_filename)
    runner_in_ctr = "node" if lang == "javascript" else "python3"
    script_in_ctr = f"/work/{script_name}"

    # Rewrite OUTPUT_PATH host paths → container path. The caller passed
    # an absolute host path; inside the container the same file lives at
    # /work/<basename>. We patch the prefix line so the script writes to
    # the bind-mounted location.
    container_out = f"/work/{out_filename}"
    if lang == "javascript":
        code = (
            f"const OUTPUT_PATH = {json.dumps(container_out)};\n"
            f"{_PPTX_SHIM_JS}\n\n{code}"
        )
    else:
        code = f"OUTPUT_PATH = {container_out!r}\n\n{code}"

    with open(script_host, "w", encoding="utf-8") as fh:
        fh.write(code)

    # The sandbox image runs as USER sandbox (uid 10001). The host tmpdir is
    # created by the worker user (e.g. ainxtappuser, different uid) with mode
    # 0700, so the container user cannot traverse it or read the script —
    # Node reports a misleading "Cannot find module" for permission-denied.
    # Make the dir+script world-readable so the bind mount is usable across
    # UIDs without touching the host's broader filesystem.
    try:
        os.chmod(workdir_host, 0o777)
        os.chmod(script_host, 0o644)
    except Exception as exc:
        logger.warning(f"[docgen] worker chmod sandbox workdir failed (continuing): {exc}")

    cmd = [
        docker_bin, "run", "--rm",
        "--network", "none",
        "--memory", _DOC_SANDBOX_MEMORY,
        "--cpus",   _DOC_SANDBOX_CPUS,
        "--read-only",
        "--tmpfs", "/tmp:rw,size=128m",
        "-v", f"{workdir_host}:/work:rw",
        "-w", "/work",
        _DOC_SANDBOX_IMAGE,
        runner_in_ctr, script_in_ctr,
    ]

    logger.info(
        f"[docgen] worker sandbox exec | job={job_id} lang={lang} "
        f"image={_DOC_SANDBOX_IMAGE} workdir={workdir_host}"
    )
    try:
        proc = _sp.run(
            cmd, capture_output=True, text=True,
            timeout=timeout or _DOC_SANDBOX_TIMEOUT,
        )
    except _sp.TimeoutExpired:
        return (124, b"", f"sandbox exec timed out after {timeout or _DOC_SANDBOX_TIMEOUT}s")
    except Exception as exc:
        return (1, b"", f"sandbox exec error: {exc}")

    if proc.returncode == 0 and os.path.exists(out_host):
        with open(out_host, "rb") as fh:
            return (0, fh.read(), proc.stderr or "")

    # Failure path: capture diagnostic context that's almost always what we
    # need (mount/permission/UID issues vs. real code errors).
    try:
        st = os.stat(workdir_host)
        diag = (
            f"\n[diag] workdir_host={workdir_host} "
            f"mode={oct(st.st_mode)} uid={st.st_uid} "
            f"listing={sorted(os.listdir(workdir_host))[:20]} "
            f"script_exists={os.path.exists(script_host)}"
        )
    except Exception as _e:
        diag = f"\n[diag] could not stat workdir: {_e}"
    return (proc.returncode, b"", (proc.stderr or "") + diag)


def _load_skill_context(fmt: str) -> str:
    """
    Load skill context files for the given format and return them concatenated
    as a single system-context string to inject into the LLM prompt.

    Files loaded per format:
      All formats : SKILL.md + brand file (DOC_BRAND_FILE)
      pptx        : SKILL.md + pptxgenjs.md + brand file (DOC_BRAND_FILE)
                    (pptxgenjs.md is the full API reference for creating from scratch)

    Returns empty string if skill files are not found (graceful degradation).
    """
    folder = _SKILL_FOLDER.get(fmt.lower().strip())
    if not folder:
        logger.warning(f"[docgen] worker no skill folder for fmt={fmt!r} — skill context unavailable")
        return ""

    skill_md_path  = os.path.join(_SKILLS_ROOT, folder, "SKILL.md")
    brand_md_path  = os.path.join(_SKILLS_ROOT, _DOC_BRAND_FILE)

    # Per-format context files. The SKELETON is a working fill-in template using
    # the ainxt composition wrapper for that format; loading it teaches the model
    # the wrapper's exact API by example (far more reliable than prose alone).
    # The optional API-reference doc (DOCXJS.md / PPTXGENJS.md) is the escape
    # hatch for anything the wrapper doesn't cover.
    files_to_load = [("SKILL", skill_md_path), ("BRAND", brand_md_path)]
    if folder == "docx":
        files_to_load.append(("SKELETON", os.path.join(_SKILLS_ROOT, "docx", "SKELETON.js")))
    elif folder == "pptx":
        files_to_load.append(("SKELETON", os.path.join(_SKILLS_ROOT, "pptx", "SKELETON.js")))
    elif folder == "xlsx":
        files_to_load.append(("SKELETON", os.path.join(_SKILLS_ROOT, "xlsx", "SKELETON.py")))

    parts = []
    for label, path in files_to_load:
        try:
            with open(path, "r", encoding="utf-8") as fh:
                parts.append(f"<!-- {label} CONTEXT: {os.path.basename(path)} -->\n{fh.read()}")
        except FileNotFoundError:
            logger.warning(f"[docgen] worker skill file not found: {path}")
        except Exception as exc:
            logger.warning(f"[docgen] worker could not read {path}: {exc}")

    return "\n\n".join(parts)


def _merge_llm_cost(base: dict, extra: dict | None, *,
                    job_id: str = "", phase: str = "") -> dict:
    """Fold an ADDITIONAL usage dict (tokens/in_tok/out_tok/cost_usd) into `base`,
    summing the numeric fields in place. Used to combine the doc-gen phases
    (outline structuring + code generation + repair retries) into ONE usage
    total that a single increment_usage() call deducts — so the expensive
    code-writer cost is no longer dropped, and it is only ever counted once.

    Numeric fields are additive; `model` is filled from `extra` only if `base`
    has none yet (the structuring model is the primary label). Fail-open:
    a missing/empty `extra` leaves `base` unchanged. Returns `base`."""
    if not extra:
        return base
    try:
        for _k in ("tokens", "in_tok", "out_tok"):
            base[_k] = int(base.get(_k) or 0) + int(extra.get(_k) or 0)
        base["cost_usd"] = float(base.get("cost_usd") or 0.0) + float(extra.get("cost_usd") or 0.0)
        if not base.get("model") and extra.get("model"):
            base["model"] = extra.get("model")
        if job_id:
            logger.info(
                f"[docgen] worker merged {phase or 'llm'} cost | job={job_id} "
                f"+tokens={int(extra.get('tokens') or 0)} "
                f"+cost_usd={float(extra.get('cost_usd') or 0.0):.6f} "
                f"→ total_cost_usd={float(base.get('cost_usd') or 0.0):.6f}"
            )
    except Exception as _merr:  # noqa: BLE001
        logger.warning(f"[docgen] worker _merge_llm_cost failed ({phase}): {_merr}")
    return base


def _skill_generate(
    job_id: str,
    fmt: str,
    question: str,
    title: str = "",
    sections: list | None = None,
    domain: str | None = None,
    theme: str = "dark_executive",
    override_prompt: str | None = None,
    llm_caller=None,
    parsed_attachment: str = "",
    source_filename: str = "",
    cost_sink: dict | None = None,
    progress_step: tuple[int, int] | None = None,
) -> tuple[bytes, str, str] | None:
    """
    Generate a document using the ainxt_doc_craft skillset exactly as designed:

    Step 1 — LLM structures content (JSON sections) enriched with SKILL.md + brand file (DOC_BRAND_FILE)
    Step 2 — LLM writes generation code using the ainxt_doc_craft composition wrappers:
               DOCX → JavaScript using the docx wrapper   (per docx/SKILL.md + SKELETON.js)
               PPTX → JavaScript using the pptx wrapper   (per pptx/SKILL.md + SKELETON.js)
               XLSX → Python using the xlsx wrapper        (per xlsx/SKILL.md + SKELETON.py)
               PDF  → JavaScript (authored as docx, exported to PDF)
               CSV  → Python stdlib `csv` (flat file; not a wrapper)
    Step 3 — Execute via sandbox.doc_executor.build() (docx/pptx/xlsx/pdf): runs the
             wrapper code, performs the docx→pdf export for PDF, and renders page-image
             previews. CSV runs via the legacy _run_in_doc_sandbox() OUTPUT_PATH path.
    Step 4 — Optional post-processing scripts (validate/recalc/thumbnail) run only if a
             scripts/ dir exists in the skill folder.

    llm_caller: optional callable(prompt: str) -> str
        Injected by callers (e.g. scenario tests) to use a specific LLM endpoint.

    cost_sink: optional dict
        If provided, it is populated (in place) with the ACCUMULATED LLM usage
        of this function — tokens/in_tok/out_tok/cost_usd/model/calls — summed
        across the content-structuring pass, the code-generation pass, and every
        self-repair retry. Callers fold this into their single increment_usage()
        deduction so budget accounting reflects the true (dominant) doc cost.
        Updated on every LLM call, so a partial total survives even if the job
        later fails. Costs nothing when omitted.

    progress_step: optional (step, total_steps) tuple matching the caller's
        own _publish_progress numbering (e.g. (4, 6) on the /ask flow). The
        self-repair loop below is the single largest silent block in doc
        generation — the caller's "Generating File" label was previously
        static for the entire loop (measured ~98s in production), giving no
        indication of whether a slow build is one long execution or several
        repair rounds. When provided, each attempt/repair round republishes
        the SAME step/total with an updated `detail` string, so a client
        polling doc:progress (or subscribed via the /docs/job/{id}/stream SSE
        endpoint) sees "Attempt 2/8…", "Repairing code (round 1)…" etc.
        instead of one frozen label. None (the default) preserves the old
        silent behaviour for callers that don't pass it.

    Generation happens ONLY via the ainxt_doc_craft sandbox (with LLM
    self-repair retries). There is NO legacy tools.doc_generator fallback — on
    persistent failure this returns None and the caller surfaces a retry error.
    Returns (file_bytes, extension, mime_type) or None on failure.
    """
    import subprocess
    import shutil
    import tempfile as _tmpmod

    logger.info(f"[docgen] worker _skill_generate START | job={job_id} fmt={fmt} title={_safe_log(title)}")

    _fmt_norm = fmt.lower().strip()

    # ── Map format to skill folder and file extension ─────────────────────────
    _FMT_TO_SKILL = {
        "docx": "docx", "doc": "docx", "word": "docx",
        "pptx": "pptx", "ppt": "pptx", "powerpoint": "pptx",
        "pdf":  "pdf",
        "xlsx": "xlsx", "xls": "xlsx", "excel": "xlsx", "csv": "xlsx",
    }
    _FMT_TO_EXT = {
        "docx": "docx", "doc": "docx", "word": "docx",
        "pptx": "pptx", "ppt": "pptx", "powerpoint": "pptx",
        "pdf":  "pdf",
        "xlsx": "xlsx", "xls": "xlsx", "excel": "xlsx",
        # CSV stays CSV — the xlsx SKILL.md explicitly covers .csv/.tsv,
        # so we reuse the xlsx skill *folder* (_FMT_TO_SKILL above) for context,
        # but the generated artefact must keep the .csv extension + text/csv MIME.
        "csv":  "csv",
        "txt":  "txt",  "md":  "md",
    }
    skill_folder = _FMT_TO_SKILL.get(_fmt_norm)
    ext          = _FMT_TO_EXT.get(_fmt_norm, _fmt_norm)

    # ── 1. Load SKILL.md + AiNxt_BRAND.md ─────────────────────────────────────
    skill_context = _load_skill_context(_fmt_norm)
    if skill_context:
        logger.info(f"[docgen] worker skill context loaded | job={job_id} fmt={_fmt_norm} chars={len(skill_context)}")
    else:
        logger.warning(f"[docgen] worker skill context missing for fmt={_fmt_norm!r} | job={job_id}")

    # ── Cost accumulator ──────────────────────────────────────────────────────
    # Every LLM call in this function (content structuring + code generation +
    # up to _CODE_MAX_ATTEMPTS self-repair retries) runs on the "complex" cloud
    # model and is the DOMINANT cost of a doc job. Historically its meta was
    # discarded, so budget accounting only saw the cheap outline pass and every
    # doc showed <$0.01. Accumulate tokens/cost here and return them so the
    # caller can fold them into the single increment_usage() deduction.
    _skill_meta = {
        "tokens": 0, "in_tok": 0, "out_tok": 0, "cost_usd": 0.0,
        "model": None, "calls": 0,
    }

    # ── Helper: call LLM ──────────────────────────────────────────────────────
    def _llm(prompt: str) -> str:
        if llm_caller is not None:
            # Injected caller (tests) — no meta available; nothing to accumulate.
            return (llm_caller(prompt) or "").strip()
        from models.model_router import model_router
        result = model_router.generate(prompt, model_hint="complex", return_meta=True)
        # model_router.generate(return_meta=True) returns {"text": ..., "meta":
        # {model,in_tok,out_tok,tokens,cost_usd,latency}} — the usage lives under
        # "meta" (see model_router.generate). Accumulate across every call.
        # Fail-open: bad/missing meta contributes 0.
        _m = (result.get("meta") if isinstance(result, dict) else None) or {}
        try:
            _skill_meta["tokens"]   += int(_m.get("tokens") or 0)
            _skill_meta["in_tok"]   += int(_m.get("in_tok") or 0)
            _skill_meta["out_tok"]  += int(_m.get("out_tok") or 0)
            _skill_meta["cost_usd"] += float(_m.get("cost_usd") or 0.0)
            _skill_meta["calls"]    += 1
            if _m.get("model"):
                _skill_meta["model"] = _m.get("model")
            # Mirror the running total into the caller's sink so a partial cost
            # survives even if generation fails after this call.
            if cost_sink is not None:
                cost_sink.update(_skill_meta)
        except Exception:  # noqa: BLE001
            pass
        _text = result.get("text") if isinstance(result, dict) else result
        return (_text or "").strip()

    # ── 2. Structure content via LLM (if sections not already provided) ───────
    if not sections:
        logger.info(f"[docgen] worker _skill_generate: structuring content | job={job_id}")
        _struct_prompt = (
            f"{skill_context}\n\n"
            f"---\n\n"
            f"You are a professional document structuring assistant following the "
            f"ainxt_doc_craft specification and AiNxt brand guidelines above.\n\n"
            f"Structure this document request into well-organized sections:\n"
            f"Request: {question}\n\n"
            f"Respond with ONLY valid JSON — no markdown fences:\n"
            f'{{"title":"<title>","domain":"<payments|banking|fintech|default>",'
            f'"sections":[{{"heading":"<h>","content":"<body>","bullets":["<b>"],'
            f'"level":1,"callout":{{"label":"<l>","text":"<t>"}},"table":null}}]}}'
        )
        try:
            raw = _llm(_struct_prompt)
            raw = re.sub(r"^```[a-zA-Z]*\s*", "", raw)
            raw = re.sub(r"\s*```\s*$", "", raw.strip()).strip()
            struct   = json.loads(raw)
            sections = struct.get("sections") or []
            if not title:
                title = _sanitize_llm_title((struct.get("title") or "").strip(), question)
            if not domain:
                domain = (struct.get("domain") or "").strip().lower() or None
            logger.info(f"[docgen] worker content structured | job={job_id} sections={len(sections)}")
        except Exception as exc:
            logger.warning(f"[docgen] worker structuring failed ({exc}) — fallback | job={job_id}")
            sections = [{"heading": "Content", "content": question, "bullets": [], "level": 1}]
            if not title:
                title = _derive_title_from_question(question)

    # ── 3. Ask LLM to write generation code using the skill's wrapper libs ────
    # The LLM reads SKILL.md + SKELETON (already in skill_context) and writes
    # code that uses the ainxt composition wrappers preinstalled in the sandbox:
    #   DOCX/PDF → JavaScript, require('ainxt-doc')     (PDF = docx exported)
    #   PPTX     → JavaScript, require('ainxt-deck')
    #   XLSX     → Python,     from ainxt_sheet import Book
    # The wrappers write to their default /work/output.<ext> on .save(), which is
    # exactly where the sandbox reads the result from — no OUTPUT_PATH needed.

    _lang_map = {"docx": "javascript", "pptx": "javascript", "pdf": "javascript", "xlsx": "python"}
    _lib_map  = {
        "docx": "the preinstalled `ainxt-doc` module (const doc = require('ainxt-doc'))",
        "pptx": "the preinstalled `ainxt-deck` module (const deck = require('ainxt-deck'))",
        "xlsx": "the preinstalled `ainxt_sheet` module (from ainxt_sheet import Book)",
    }
    # CSV reuses the xlsx skill folder but must be a FLAT-CSV file, NOT an Excel
    # workbook — so it does NOT use ainxt_sheet. Stdlib `csv` + `random` + curated
    # lists produce realistic values with no external deps.
    if ext == "csv":
        _lang = "python"
        _lib  = (
            "the stdlib `csv` module (`import csv`) with `random`, `string`, "
            "`datetime`, and `uuid` for realistic value generation. Use "
            "`csv.writer(fh).writerows(rows)` — do NOT import pandas or faker "
            "(neither is guaranteed in the execution environment), and do NOT "
            "use ainxt_sheet (that produces .xlsx, not flat .csv)."
        )
    elif skill_folder == "docx":
        # docx skill covers both .docx and .pdf (pdf is a docx exported to PDF).
        _lang = "javascript"
        _lib  = _lib_map["docx"]
    else:
        _lang = _lang_map.get(skill_folder, "python")
        _lib  = _lib_map.get(skill_folder, "the appropriate library")

    # Bulk TABULAR xlsx dataset: the structuring step wrapped the request as a
    # single `csv_schema` section (columns + row_count). Render it with the
    # openpyxl flat-sheet rules and a hard row/column directive, NOT the
    # narrative xlsx rules — otherwise the code-writer emits a small styled
    # report instead of N generated rows.
    _xlsx_schema_section = None
    if ext == "xlsx":
        for _s in (sections or []):
            if isinstance(_s, dict) and _s.get("heading") == "csv_schema":
                _xlsx_schema_section = _s
                break
    _is_xlsx_testdata = _xlsx_schema_section is not None
    if _is_xlsx_testdata:
        _xlsx_cols = _xlsx_schema_section.get("csv_columns") or []
        _xlsx_rows = _xlsx_schema_section.get("csv_row_count")
        # Honour an explicit count from the prompt if the schema omitted one.
        if not _xlsx_rows:
            _xlsx_rows = _extract_row_count(question) or 200
        # Bulk flat dataset: use the ainxt_sheet wrapper's plain data sheet
        # (Book.sheet(...).rows(...)) which stays numeric and brand-styled, with
        # the stdlib for value generation. save(recalc=False) skips the (pointless
        # for a formula-free sheet) LibreOffice recalculation pass.
        _lib = (
            "the preinstalled `ainxt_sheet` module (from ainxt_sheet import Book) "
            "with the stdlib `random`, `string`, `datetime`, and `uuid` for "
            "realistic value generation — do NOT import pandas or faker (neither "
            "is guaranteed in the execution environment)."
        )

    # Strip non-serializable internal fields (e.g. _image_bytes attached by
    # _enrich_with_images) before handing sections to the LLM. The LLM only
    # needs textual content; raw image bytes are wired into the generated
    # code separately via the output-side image pipeline.
    def _sanitize_for_json(obj):
        if isinstance(obj, dict):
            return {
                k: _sanitize_for_json(v)
                for k, v in obj.items()
                if not (isinstance(k, str) and k.startswith("_"))
                and not isinstance(v, (bytes, bytearray, memoryview))
            }
        if isinstance(obj, list):
            return [_sanitize_for_json(x) for x in obj]
        if isinstance(obj, (bytes, bytearray, memoryview)):
            return None
        return obj

    _sections_for_prompt = _sanitize_for_json(sections)
    # Cap the serialized sections (env-tunable). The previous 6000-char cap silently
    # dropped later sections and code blocks before the code-writing LLM ran; raised
    # to preserve verbatim/code-heavy content while still bounding the prompt size.
    _sections_json = json.dumps(_sections_for_prompt, ensure_ascii=False, indent=2)[:_SECTIONS_JSON_MAX]
    _brand_note = (
        "Apply AiNxt brand: Navy #1F3864 headers, Green #00A551 accents, "
        "Arial font, 'AiNxt — Confidential' footer, alternating row shading #F2F5FA."
    )

    # When the user uploaded a CSV/XLSX template (or any parseable file) and
    # asked the worker to expand / append rows / mirror the schema, hand the
    # raw parsed content to the code-writer so it can match column names,
    # ordering, and data types EXACTLY. Without this, the code-writer only
    # sees the JSON-roundtripped sections and loses the precise schema.
    _template_block = ""
    if parsed_attachment:
        _template_snippet = parsed_attachment[:4000]
        _template_block = (
            f"\nUSER-UPLOADED TEMPLATE ({source_filename or 'attachment'}):\n"
            f"```\n{_template_snippet}\n```\n"
            f"REQUIREMENT: Mirror these exact column names, order, and data "
            f"types. Append new rows that follow the same patterns. Do NOT "
            f"invent extra columns. Do NOT drop columns. When the template "
            f"contains an obvious sequence (order_id, customer_id, …) "
            f"continue it from the last value.\n"
        )

    # Select the per-format code-writer rules. A bulk tabular xlsx request uses
    # the openpyxl flat-sheet dataset rules instead of the narrative xlsx rules.
    _fmt_rules = _FMT_TO_RULES.get("xlsx_testdata" if _is_xlsx_testdata else ext, "")

    # Hard directive for bulk datasets: the exact row count and full column list
    # must be honoured (this is the fix for "500 records but <500 rows / dropped
    # columns"). Placed at the top of CRITICAL RULES so it cannot be overlooked.
    _dataset_directive = ""
    if _is_xlsx_testdata:
        _col_names = [
            (c.get("name") if isinstance(c, dict) else str(c))
            for c in _xlsx_cols
        ]
        _dataset_directive = (
            f"0. THIS IS A FLAT TABULAR DATASET. Generate EXACTLY {_xlsx_rows} "
            f"data rows (plus one header row) and emit ALL {len(_col_names)} "
            f"columns in this exact order: {_col_names}. Do NOT cap the row "
            f"count, do NOT drop or rename columns, and do NOT add any narrative "
            f"text, summary, or extra sheets. Header row + data rows only.\n"
        )

    # Output-path rule differs by engine: the wrapper formats write to the
    # wrapper's FIXED default path via .save() (no filename arg); only the legacy
    # CSV path writes to the injected OUTPUT_PATH variable. Getting this wrong is a
    # silent failure — the code "succeeds" but writes a file the worker never reads.
    if ext == "csv":
        _save_rule = (
            "1. Save the output file to the variable OUTPUT_PATH (already defined before your "
            "code runs). Do NOT hardcode a filename.\n"
        )
    else:
        _save_rule = (
            f"1. Do NOT choose or pass a filename. Use {_lib} and finish with its `.save()` "
            f"(e.g. b.save() / d.save()) with NO path argument — the wrapper writes the correct "
            f"output file automatically. Never call raw openpyxl `wb.save('name.xlsx')` or "
            f"`fs.writeFileSync('name...')`, and never import a different library — use ONLY the "
            f"wrapper named above.\n"
        )

    _code_prompt = (
        f"{skill_context}\n\n"
        f"---\n\n"
        f"Write {_lang} code using {_lib} to generate a {ext.upper()} file.\n\n"
        f"Document title: {title}\n"
        f"Domain: {domain or 'payments'}\n"
        f"{_brand_note}\n"
        f"{_template_block}\n"
        f"Content sections (render ALL of them):\n{_sections_json}\n\n"
        f"CRITICAL RULES:\n"
        f"{_dataset_directive}"
        f"{_save_rule}"
        f"2. Follow ALL patterns from the SKILL.md above exactly.\n"
        f"3. Output ONLY the code block — no explanation, no markdown prose outside the code.\n"
        f"4. The code must be complete and runnable as-is.\n"
        f"5. The code must run without raising on any input. Guard against missing/empty/None "
        f"fields in the sections data above, and tolerate optional keys being absent.\n\n"
        f"{_fmt_rules}\n"
        f"```{_lang}\n// your complete code here\n```"
    )

    logger.info(f"[docgen] worker _skill_generate: requesting {_lang} code from LLM | job={job_id}")
    try:
        code_raw = _llm(_code_prompt)
    except Exception as exc:
        logger.error(f"[docgen] worker LLM code generation failed | job={job_id}: {exc}")
        code_raw = ""

    # Extract code block from LLM response
    _code_match = re.search(
        r"```(?:javascript|js|python|py)?\s*([\s\S]+?)```",
        code_raw, re.IGNORECASE
    )
    generated_code = _code_match.group(1).strip() if _code_match else ""

    # ── 4. Execute the generated code — with self-healing retry ───────────────
    # A single LLM code draft can carry a syntax/runtime bug (e.g.
    # "SyntaxError: Unexpected identifier"). Rather than fail the whole job, we
    # feed the exact sandbox stderr + the broken source back to the LLM and ask
    # it to repair the code, retrying up to _CODE_MAX_ATTEMPTS times within the
    # 30-min budget. There is NO fallback — if every sandbox attempt fails the
    # job surfaces a retry error (see step 6). Generation is sandbox-only.
    file_bytes: bytes | None = None
    _out_basename = f"output.{ext}"

    # Formats handled by sandbox.doc_executor.build(): it runs the wrapper code,
    # performs the docx→pdf export (for pdf) and renders page-image previews, and
    # returns a DocBuildResult. CSV is NOT one of build()'s formats — it stays on
    # the older _run_in_doc_sandbox() path (a flat stdlib-csv file needs no
    # conversion/preview and writes to OUTPUT_PATH).
    _EXECUTOR_FORMATS = {"docx", "pptx", "xlsx", "pdf"}
    _use_executor = ext in _EXECUTOR_FORMATS
    # Page-image previews produced by build() (JPEG bytes), surfaced on success.
    _exec_page_images: list = []

    # ── Visual-QA (feature-flagged) ───────────────────────────────────────────
    # After a build SUCCEEDS, render pages are critiqued against the ainxt brand
    # contract by a vision model (sandbox/doc_critic). If it flags visual defects
    # (low contrast, overflow, clipped text, off-brand tables, etc.) the build
    # code is repaired against those findings and rebuilt — bounded so a visual
    # nicety can never blow the job budget. Fails OPEN: any unavailability ships
    # the document as-is. Only meaningful for the executor formats that produce
    # page images (docx/pptx/xlsx/pdf); CSV has no visual form.
    _VISION_QA_MAX_REVISIONS = int(os.getenv("DOC_VISION_QA_MAX_REVISIONS", "2"))
    _VISION_QA_BUDGET_S = int(os.getenv("DOC_VISION_QA_BUDGET_S", "900"))
    _vision_qa_revisions = 0
    _vision_qa_verdict = "not_run"

    def _exec_once(code: str, timeout: int | None = None) -> tuple[int, bytes, str]:
        """Run one build attempt. Returns (rc, output_bytes, stderr) so the
        self-repair loop below stays unchanged regardless of which engine runs.
        rc==0 with non-empty bytes == success."""
        if _use_executor:
            # Single generic engine: wrapper code → deliverable (+ pdf export +
            # previews). fmt drives the recipe (see sandbox.doc_executor._FORMATS).
            from sandbox import doc_executor
            try:
                res = doc_executor.build(code, ext, images=None)
            except Exception as _be:
                return (1, b"", f"doc_executor.build raised: {_be}")
            if res.ok and res.doc_bytes:
                if res.page_images:
                    _exec_page_images[:] = res.page_images
                return (0, res.doc_bytes, res.logs or "")
            return (1, b"", res.error or res.logs or "build failed")
        # CSV (and any non-executor format): legacy direct exec into OUTPUT_PATH.
        _tmpdir = _tmpmod.mkdtemp(prefix="ainxt_skill_")
        try:
            return _run_in_doc_sandbox(
                lang=_lang, code=code, out_filename=_out_basename,
                workdir_host=_tmpdir, job_id=job_id, timeout=timeout,
            )
        finally:
            shutil.rmtree(_tmpdir, ignore_errors=True)

    def _repair_code(broken: str, stderr: str) -> str:
        """Ask the LLM to fix the code given the exact sandbox error."""
        # The wrapper-based formats write to the wrapper's default output path via
        # .save() (no OUTPUT_PATH). Only the legacy CSV path uses OUTPUT_PATH.
        _save_hint = (
            "Keep writing the output the SAME way as before (call the wrapper's "
            "`.save()` / `b.save()` with no path — it writes the correct output "
            "file automatically)."
            if _use_executor else
            "Keep saving the output to OUTPUT_PATH (already defined)."
        )
        _fix_prompt = (
            f"The following {_lang} code (using {_lib}) was meant to generate a "
            f"{ext.upper()} file, but it FAILED when executed.\n\n"
            f"=== EXECUTION ERROR (stderr) ===\n{stderr[:2000]}\n\n"
            f"=== BROKEN CODE ===\n```{_lang}\n{broken}\n```\n\n"
            f"Fix the error and return the COMPLETE corrected code. Requirements:\n"
            f"1. Return ONLY one {_lang} code block — no prose, no explanation.\n"
            f"2. {_save_hint}\n"
            f"3. Fix the specific error above AND guard against missing/None "
            f"fields so it cannot raise on any input.\n"
            f"4. Do not use any library other than {_lib}.\n"
            f"```{_lang}\n// corrected complete code here\n```"
        )
        try:
            _resp = _llm(_fix_prompt)
        except Exception as _fe:
            logger.warning(f"[docgen] worker repair LLM call failed | job={job_id}: {_fe}")
            return ""
        _m = re.search(
            r"```(?:javascript|js|python|py)?\s*([\s\S]+?)```", _resp, re.IGNORECASE
        )
        return (_m.group(1).strip() if _m else _resp.strip())

    # ── Generous 30-min budget, sandbox-only, NO fallback ─────────────────────
    # Document generation MUST succeed via the ainxt_doc_craft sandbox. We
    # give the sandbox a large total wall-clock budget (_DOC_TOTAL_BUDGET_SEC,
    # default 30 min) spread across self-repair attempts, so a slow/transient
    # build has every chance to complete via the skill. Each attempt's timeout
    # is the remaining budget (capped so a single hang can't consume it all and
    # starve the repair loop). We stop only on success or when the budget/attempt
    # ceiling is exhausted — then the job fails (never an inferior document).
    _budget_deadline = time.monotonic() + _DOC_TOTAL_BUDGET_SEC
    _last_err = ""
    _cur_code = generated_code
    _attempt = 0
    # PERF instrumentation (A6): this loop was measured as the dominant cost
    # in the logged production example (~98s of a ~3m22s job) but with no
    # visibility into WHERE that time goes — one long sandbox execution, one
    # slow repair LLM call, or many cheap rounds adding up. These counters are
    # read-only bookkeeping (no behavior change) and summarized in one log
    # line after the loop exits, so that question can be answered from logs
    # instead of guessed at.
    _loop_t0 = time.monotonic()
    _sandbox_time_total = 0.0
    _repair_time_total = 0.0
    _repair_calls = 0
    while _cur_code and _attempt < _CODE_MAX_ATTEMPTS:
        _attempt += 1
        _remaining = _budget_deadline - time.monotonic()
        if _remaining <= 5:
            logger.warning(
                f"[docgen] worker 30-min skill budget exhausted before attempt "
                f"{_attempt} | job={job_id} (sandbox-only, no fallback)"
            )
            break
        # Per-attempt timeout: use the whole remaining budget, but leave headroom
        # for at least one repair round when earlier attempts remain.
        _attempts_left = _CODE_MAX_ATTEMPTS - _attempt + 1
        _attempt_timeout = int(max(60, min(_remaining, _remaining / max(1, _attempts_left) * 2)))
        logger.info(
            f"[docgen] worker dispatching {_lang} code to doc-sandbox | "
            f"job={job_id} attempt={_attempt}/{_CODE_MAX_ATTEMPTS} "
            f"code_len={len(_cur_code)} attempt_timeout={_attempt_timeout}s "
            f"budget_remaining={int(_remaining)}s"
        )
        if progress_step is not None:
            _p_step, _p_total = progress_step
            _p_detail = (
                f"Building {fmt.upper()} (attempt {_attempt}/{_CODE_MAX_ATTEMPTS})…"
                if _attempt == 1 else
                f"Retrying build after repair (attempt {_attempt}/{_CODE_MAX_ATTEMPTS})…"
            )
            _publish_progress(job_id, _p_step, _p_total, "Generating File", _p_detail)
        _sandbox_t0 = time.monotonic()
        try:
            rc, out_bytes, sandbox_err = _exec_once(_cur_code, timeout=_attempt_timeout)
        except Exception as exc:
            rc, out_bytes, sandbox_err = 1, b"", f"host error: {exc}"
        _sandbox_elapsed = time.monotonic() - _sandbox_t0
        _sandbox_time_total += _sandbox_elapsed

        if rc == 0 and out_bytes:
            file_bytes = out_bytes
            logger.info(
                f"[docgen] worker skill code execution SUCCESS | "
                f"job={job_id} attempt={_attempt} size={len(file_bytes):,} "
                f"sandbox_elapsed={_sandbox_elapsed:.1f}s"
            )

            # ── Visual QA pass ────────────────────────────────────────────────
            # Only when we have rendered pages (executor formats), still have a
            # revision + budget left, and haven't exhausted the build attempts.
            # A "revise" verdict repairs the code against the visual findings and
            # loops (continue) to rebuild + re-critique; anything else ships.
            _qa_can_run = (
                _use_executor
                and _exec_page_images
                and _vision_qa_revisions < _VISION_QA_MAX_REVISIONS
                and _attempt < _CODE_MAX_ATTEMPTS
                and (_budget_deadline - time.monotonic()) > 60
            )
            if _qa_can_run:
                try:
                    from sandbox import doc_critic
                    _crit = doc_critic.critique(list(_exec_page_images), ext)
                except Exception as _qae:
                    logger.warning(f"[docgen] worker visual-QA raised (shipping as-is) | job={job_id}: {_qae}")
                    _crit = None

                if _crit is not None:
                    _vision_qa_verdict = _crit.verdict
                    if _crit.needs_revision:
                        logger.info(
                            f"[docgen] worker visual-QA verdict=revise | job={job_id} "
                            f"revision={_vision_qa_revisions + 1}/{_VISION_QA_MAX_REVISIONS} "
                            f"issues={len(_crit.issues)} pages_reviewed={_crit.pages_reviewed} "
                            f"model={_crit.model or 'vision'}"
                        )
                        if progress_step is not None:
                            _p_step, _p_total = progress_step
                            _publish_progress(
                                job_id, _p_step, _p_total, "Generating File",
                                f"Polishing layout (visual review {_vision_qa_revisions + 1})…",
                            )
                        try:
                            _repaired = doc_critic.strip_code_fence(
                                _llm(doc_critic.build_repair_prompt(ext, _cur_code, _crit.issues))
                            )
                        except Exception as _re:
                            logger.warning(f"[docgen] worker visual-QA repair gen failed (shipping) | job={job_id}: {_re}")
                            _repaired = ""
                        # Only loop back if the repair actually changed something;
                        # otherwise ship the good build we already have.
                        if _repaired and _repaired != _cur_code:
                            _vision_qa_revisions += 1
                            _cur_code = _repaired
                            # Keep `file_bytes` as the last-good build so a repair
                            # that fails to build later still ships this one.
                            continue
                        logger.info(f"[docgen] worker visual-QA repair produced no change — shipping | job={job_id}")
                    else:
                        logger.info(
                            f"[docgen] worker visual-QA verdict={_crit.verdict} | job={job_id} "
                            f"pages_reviewed={_crit.pages_reviewed}"
                        )
            break

        _last_err = sandbox_err or f"rc={rc}"
        logger.warning(
            f"[docgen] worker skill code execution failed | job={job_id} "
            f"attempt={_attempt}/{_CODE_MAX_ATTEMPTS} rc={rc} "
            f"sandbox_elapsed={_sandbox_elapsed:.1f}s stderr={_last_err[:400]!r}"
        )
        # Repair for the next attempt (skip after the final attempt / no budget).
        if _attempt < _CODE_MAX_ATTEMPTS and (_budget_deadline - time.monotonic()) > 30:
            logger.info(f"[docgen] worker requesting code repair from LLM | job={job_id}")
            if progress_step is not None:
                _p_step, _p_total = progress_step
                _publish_progress(
                    job_id, _p_step, _p_total, "Generating File",
                    f"Repairing build error (round {_repair_calls + 1})…",
                )
            _repair_t0 = time.monotonic()
            _cur_code = _repair_code(_cur_code, _last_err)
            _repair_elapsed = time.monotonic() - _repair_t0
            _repair_time_total += _repair_elapsed
            _repair_calls += 1
            logger.info(
                f"[docgen] worker code repair call done | job={job_id} "
                f"repair_elapsed={_repair_elapsed:.1f}s"
            )

    logger.info(
        f"[docgen] worker skill-generation loop SUMMARY | job={job_id} "
        f"attempts={_attempt}/{_CODE_MAX_ATTEMPTS} repair_calls={_repair_calls} "
        f"loop_elapsed={time.monotonic() - _loop_t0:.1f}s "
        f"sandbox_time_total={_sandbox_time_total:.1f}s "
        f"repair_time_total={_repair_time_total:.1f}s "
        f"vision_qa_verdict={_vision_qa_verdict} vision_qa_revisions={_vision_qa_revisions} "
        f"succeeded={file_bytes is not None}"
    )

    # ── 5. Run post-processing skill scripts ──────────────────────────────────
    if file_bytes and skill_folder:
        scripts_dir = os.path.join(_SKILLS_ROOT, skill_folder, "scripts")

        # validate.py — DOCX, PPTX, XLSX
        if ext in ("docx", "pptx", "xlsx"):
            _validate = os.path.join(scripts_dir, "office", "validate.py")
            if os.path.exists(_validate):
                _vtmp = _tmpmod.mkdtemp(prefix="ainxt_validate_")
                try:
                    _vfile = os.path.join(_vtmp, f"validate.{ext}")
                    with open(_vfile, "wb") as _vf:
                        _vf.write(file_bytes)
                    _vproc = subprocess.run(
                        [sys.executable, _validate, _vfile],
                        capture_output=True, text=True, timeout=30,
                        cwd=scripts_dir,
                    )
                    if _vproc.returncode == 0:
                        logger.info(f"[docgen] worker validate.py PASSED | job={job_id} ext={ext}")
                    else:
                        logger.warning(
                            f"[docgen] worker validate.py warnings | job={job_id}: "
                            f"{_vproc.stdout[:300]} {_vproc.stderr[:300]}"
                        )
                except Exception as _ve:
                    logger.warning(f"[docgen] worker validate.py skipped: {_ve}")
                finally:
                    shutil.rmtree(_vtmp, ignore_errors=True)

        # recalc.py — XLSX MANDATORY (per SKILL.md)
        if ext == "xlsx":
            _recalc = os.path.join(scripts_dir, "recalc.py")
            if os.path.exists(_recalc):
                # ── Formula presence check ────────────────────────────────────────
                # recalc.py requires LibreOffice (soffice) on the host. Only invoke
                # it when the workbook actually contains formulas — plain data tables
                # (e.g. a product list) need no recalculation and should be delivered
                # directly without touching soffice.
                import io as _io
                import openpyxl as _openpyxl
                _formula_count = 0
                try:
                    _wb_check = _openpyxl.load_workbook(
                        _io.BytesIO(file_bytes), data_only=False
                    )
                    for _sn in _wb_check.sheetnames:
                        for _row in _wb_check[_sn].iter_rows():
                            for _cell in _row:
                                if (
                                    _cell.value
                                    and isinstance(_cell.value, str)
                                    and _cell.value.startswith("=")
                                ):
                                    _formula_count += 1
                    _wb_check.close()
                except Exception as _fce:
                    logger.warning(
                        f"[docgen] worker formula-check failed (will run recalc anyway) | "
                        f"job={job_id} err={_fce}"
                    )
                    _formula_count = 1  # assume formulas present — safe fallback

                if _formula_count == 0:
                    logger.info(
                        f"[docgen] worker skipping recalc.py — no formulas detected | "
                        f"job={job_id}"
                    )
                else:
                    logger.info(
                        f"[docgen] worker running recalc.py — {_formula_count} formula(s) found | "
                        f"job={job_id}"
                    )
                    _RECALC_ATTEMPTS = 3
                    _recalc_ok = False
                    _recalc_last_err = ""
                    for _r_attempt in range(1, _RECALC_ATTEMPTS + 1):
                        _rtmp = _tmpmod.mkdtemp(prefix="ainxt_recalc_")
                        try:
                            _rfile = os.path.join(_rtmp, "recalc.xlsx")
                            with open(_rfile, "wb") as _rf:
                                _rf.write(file_bytes)
                            _rproc = subprocess.run(
                                [sys.executable, _recalc, _rfile],
                                capture_output=True, text=True, timeout=180,
                                cwd=scripts_dir,
                            )
                            if _rproc.returncode == 0 and os.path.exists(_rfile):
                                with open(_rfile, "rb") as _rf2:
                                    file_bytes = _rf2.read()
                                logger.info(
                                    f"[docgen] worker recalc.py DONE | job={job_id} "
                                    f"attempt={_r_attempt}/{_RECALC_ATTEMPTS}"
                                )
                                _recalc_ok = True
                                break
                            _recalc_last_err = (_rproc.stderr or _rproc.stdout or f"rc={_rproc.returncode}")[:300]
                            logger.warning(
                                f"[docgen] worker recalc.py FAILED | job={job_id} "
                                f"attempt={_r_attempt}/{_RECALC_ATTEMPTS} "
                                f"rc={_rproc.returncode} err={_recalc_last_err!r}"
                            )
                        except Exception as _re:
                            _recalc_last_err = str(_re)[:300]
                            logger.warning(
                                f"[docgen] worker recalc.py error | job={job_id} "
                                f"attempt={_r_attempt}/{_RECALC_ATTEMPTS}: {_re}"
                            )
                        finally:
                            shutil.rmtree(_rtmp, ignore_errors=True)
                        if _r_attempt < _RECALC_ATTEMPTS:
                            time.sleep(2 * _r_attempt)  # linear backoff

                    if not _recalc_ok:
                        # Do NOT ship a workbook Excel will flag as corrupt.
                        logger.error(
                            f"[docgen] worker recalc.py failed after {_RECALC_ATTEMPTS} "
                            f"attempts — failing job (no corrupt xlsx published) | "
                            f"job={job_id} last_err={_recalc_last_err!r}"
                        )
                        _fail(
                            job_id,
                            "Excel recalculation failed — the spreadsheet could not "
                            "be finalized. Please retry.",
                        )
                        return None

        # thumbnail.py — PPTX visual QA
        if ext == "pptx":
            _thumb = os.path.join(scripts_dir, "thumbnail.py")
            if os.path.exists(_thumb):
                _ttmp = _tmpmod.mkdtemp(prefix="ainxt_thumb_")
                try:
                    _tfile = os.path.join(_ttmp, "slides.pptx")
                    with open(_tfile, "wb") as _tf:
                        _tf.write(file_bytes)
                    _tproc = subprocess.run(
                        [sys.executable, _thumb, _tfile],
                        capture_output=True, text=True, timeout=60,
                        cwd=scripts_dir,
                    )
                    logger.info(
                        f"[docgen] worker thumbnail.py done | job={job_id} "
                        f"rc={_tproc.returncode}"
                    )
                except Exception as _te:
                    logger.warning(f"[docgen] worker thumbnail.py skipped: {_te}")
                finally:
                    shutil.rmtree(_ttmp, ignore_errors=True)

    # ── 6. Skills-only — NO legacy fallback ───────────────────────────────────
    # RULE: document generation happens ONLY through the ainxt_doc_craft sandbox
    # path (steps 3-4, with self-healing retries that let the LLM repair its own
    # code up to _CODE_MAX_ATTEMPTS times). We DO NOT fall back to the legacy
    # tools.doc_generator (that produced off-brand, non-skill output).
    # If every sandbox attempt still fails, surface a clear retry error so the
    # user re-submits — never an inferior document.
    if file_bytes is None:
        logger.error(
            f"[docgen] worker ainxt_doc_craft sandbox produced no file within the "
            f"{_DOC_TOTAL_BUDGET_SEC}s budget / {_CODE_MAX_ATTEMPTS} attempt(s) — "
            f"failing (sandbox-only, NO fallback) | "
            f"job={job_id} fmt={_fmt_norm} last_err={_last_err[:200]!r}"
        )
        _fmt_label = {
            "pptx": "PPTX", "ppt": "PPTX", "powerpoint": "PPTX",
            "presentation": "PPTX", "slides": "PPTX",
            "docx": "DOCX", "doc": "DOCX", "word": "DOCX",
            "pdf":  "PDF",
            "xlsx": "XLSX", "xls": "XLSX", "excel": "XLSX",
            "csv":  "CSV",
        }.get(_fmt_norm, _fmt_norm.upper())
        _fail(
            job_id,
            f"{_fmt_label} generation failed. Please retry — "
            f"the document skill encountered an error.",
        )
        return None

    mime = MIME_TYPES.get(ext, "application/octet-stream")
    logger.info(f"[docgen] worker _skill_generate DONE | job={job_id} ext={ext} size={len(file_bytes):,}")
    return file_bytes, ext, mime


def _image_eligible_slide_types() -> set[str]:
    """Return the set of slide_type values that should receive an AI-generated
    image, based on the PPT_IMAGES_ALL_SLIDES flag.

    - Flag false (default) → cover-only: just the title slide.
    - Flag true            → title + content + closing (legacy behaviour).
    """
    if _PPT_IMAGES_ALL_SLIDES:
        return {"title", "content", "closing"}
    return {"title"}


def _resolve_image_provider() -> str:
    p = _PPT_IMG_PROVIDER.lower().strip()
    if p != "auto":
        return p
    # When LLM_PROXY_URL is set (production), the proxy holds all API keys —
    # always enable image generation and let the proxy decide the provider.
    if os.getenv("LLM_PROXY_URL"):
        return "gemini"   # proxy will auto-fallback to dalle if gemini unavailable
    # Local dev: check direct API keys
    if os.getenv("GOOGLE_API_KEY"):
        return "gemini"
    if os.getenv("OPENAI_API_KEY"):
        return "dalle"
    return "disabled"


# ── Sandbox PPTX path (uses sandbox.doc_executor.build) ─────────────────────
def _should_use_sandbox_pptx(fmt: str) -> bool:
    """Resolve DOC_PPTX_ENGINE to a concrete choice for this job."""
    if (fmt or "").lower() not in _PPTX_FMT_ALIASES:
        return False
    if _DOC_PPTX_ENGINE == "native":
        return False
    if _DOC_PPTX_ENGINE == "sandbox":
        return True
    # auto
    try:
        from sandbox.doc_executor import docker_available, image_present
        return bool(docker_available() and image_present())
    except Exception as exc:
        logger.warning(f"[docgen] worker sandbox preflight failed; using native: {exc}")
        return False


def _sandbox_image_provider() -> tuple[str, str]:
    """
    Pick (provider, prompt-suffix) for sandbox.doc_executor image generation.
    Order: OpenAI → Gemini-2.5-Flash-Image. _resolve_image_provider() returns
    "gemini" | "dalle" | "disabled"; map to executor's "gemini"/"openai" vocab.
    """
    p = _resolve_image_provider()
    if p == "disabled":
        return ("disabled", "")
    if p == "dalle":
        return ("openai", "")
    # Prefer OpenAI when both are available; fall back to Gemini Flash Image.
    if os.getenv("OPENAI_API_KEY") or os.getenv("LLM_PROXY_URL"):
        return ("openai", "")
    return ("gemini", "")  # proxy/executor selects gemini-2.5-flash-image


# ── Doc-generation model hint resolver ───────────────────────────────────────
#
# Resolution order (per product requirement):
#   1. The user's explicit chat-model selection (e.g. "openai-deep",
#      "claude-opus48", "local:Kimi-k2.5") — user agency wins.
#   2. Only when the user is on "auto", consult DOC_MODEL_PROVIDER env var
#      so admins can pin doc generation to a specific tier without touching
#      the user's chat-model preference.
#   3. Final fallback: "complex" → Claude Sonnet, which is what the system
#      did before this knob existed.
#
# The returned string is a model_router tier hint ("complex" | "fast" |
# "openai-deep" | "claude-opus48" | "claude-opus46" | "local" | …) — see
# models/model_router.py for the canonical list.

_DEFAULT_DOC_MODEL_HINT = "complex"

# After the first draft, run a self-critique + refine pass to deepen thin
# sections and improve flow. On by default; set DOCGEN_REFINE=0 to disable.
_DOCGEN_REFINE_ENABLED = (os.getenv("DOCGEN_REFINE", "1").strip() not in ("0", "false", "no", ""))


def _resolve_doc_model_hint(user_model_hint: str | None) -> str:
    """Pick the model_router hint for doc-generation LLM calls.

    See module-level comment for the order. ``"local:..."`` strings are
    normalised to ``"local"`` because that's the tier name; the specific
    local model name is forwarded separately when callers need it.
    """
    user_choice = (user_model_hint or "").strip().lower()
    if user_choice and user_choice != "auto":
        return "local" if user_choice.startswith("local:") else user_choice
    env_choice = (os.getenv("DOC_MODEL_PROVIDER", "") or "").strip().lower()
    if env_choice:
        return env_choice
    return _DEFAULT_DOC_MODEL_HINT


def _author_pptxgenjs_code(title: str, slides: list,
                           cost_sink: dict | None = None) -> tuple[str, list]:
    """
    Ask the LLM to author a complete Node script using pptxgenjs (the package
    is globally installed inside the ainxt-doc-sandbox image).

    Returns (code, images) where `images` is the list of
    {name, prompt, aspect_ratio, provider} that sandbox.doc_executor will
    materialise into the work dir BEFORE the build (so the JS can embed
    them via slide.addImage({path: "<filename>"})).
    """
    provider, _ = _sandbox_image_provider()
    eligible = _image_eligible_slide_types()
    logger.info(
        f"[docgen] worker pptxgenjs image gate | all_slides={_PPT_IMAGES_ALL_SLIDES} "
        f"eligible={sorted(eligible)} provider={provider}"
    )
    images: list[dict] = []
    slim_slides: list[dict] = []
    for idx, s in enumerate(slides, start=1):
        s = dict(s)
        s.pop("_image_bytes", None)  # native-path artefact; sandbox uses files
        stype = (s.get("slide_type") or "content").lower()
        img_prompt = (s.get("image_prompt") or "").strip()
        if img_prompt and provider != "disabled" and stype in eligible:
            name = f"slide_{idx}.png"
            s["image_file"] = name  # LLM sees this and uses it in addImage
            images.append({
                "name": name,
                "prompt": img_prompt,
                "aspect_ratio": "16:9",
                "provider": provider,
            })
        else:
            # Drop any stale image_file so non-eligible slides render purely
            # via geometric layouts in the LLM-authored pptxgenjs script.
            s.pop("image_file", None)
        slim_slides.append(s)

    skill_ctx = _load_skill_context("pptx")
    schema_json = json.dumps(
        {"title": title, "slides": slim_slides},
        ensure_ascii=False, indent=2,
    )
    prompt = (
        f"{skill_ctx}\n\n"
        "Author a COMPLETE, self-contained Node.js script that builds the "
        "PowerPoint deck described by the JSON below using the preinstalled "
        "`ainxt-deck` composition module.\n\n"
        "HARD REQUIREMENTS:\n"
        "1. `const deck = require('ainxt-deck');` then "
        "`const d = deck.create({ classification: 'Confidential', title: '<title>' });`. "
        "Do NOT require('pptxgenjs') directly and do NOT require anything else — "
        "ainxt-deck applies the configured layout, brand colours, logo, footer and "
        "accessibility rules (this is what guarantees legible text/background "
        "contrast; hand-rolled pptxgenjs colours are the cause of low-contrast "
        "and off-brand slides).\n"
        "2. Build every slide from the schema in order using the deck pattern "
        "methods only: d.cover(title, subtitle, date), d.contents(heading, [labels]), "
        "d.evidence(heading, [bullets], visual), d.split(heading, [{title,bullets}]), "
        "d.metric(heading, [{figure,label,status}], note), d.statement(sentence, attribution), "
        "d.table(heading, header, rows, {rightCols}), d.close(line, [nextSteps]), d.notes(text). "
        "Pass plain strings/numbers; never compute x/y/w/h yourself. Vary the pattern "
        "between consecutive slides.\n"
        "3. For any slide whose JSON contains `image_file`, pass it as the evidence "
        "visual: `d.evidence(heading, bullets, { image: '<image_file>' })`. The file "
        "already exists in the working directory. Never add stock photography, "
        "clip-art, or decorative images of your own.\n"
        "4. Finish with `d.save()` — it writes /work/output.pptx automatically. Do "
        "NOT pass a filename and do NOT call pres.writeFile yourself.\n"
        "5. No network access, no extra npm packages, no shell-outs.\n"
        "6. Output ONLY the JavaScript code. No prose, no markdown fences.\n\n"
        f"DECK SCHEMA:\n{schema_json}\n"
    )

    from models.model_router import model_router
    # return_meta=True so the (dominant) PPTX code-authoring cost is captured
    # for budget accounting instead of being dropped.
    result = model_router.generate(prompt, model_hint="complex", return_meta=True)
    if isinstance(result, dict):
        raw = result.get("text") or ""
        # Usage lives under "meta" (see model_router.generate return contract).
        _m = result.get("meta") or {}
        if cost_sink is not None:
            try:
                cost_sink["tokens"]   = int(_m.get("tokens") or 0)
                cost_sink["in_tok"]   = int(_m.get("in_tok") or 0)
                cost_sink["out_tok"]  = int(_m.get("out_tok") or 0)
                cost_sink["cost_usd"] = float(_m.get("cost_usd") or 0.0)
                cost_sink["model"]    = _m.get("model")
            except Exception:  # noqa: BLE001
                pass
    elif isinstance(result, tuple):
        raw = result[0]
    else:
        raw = result
    code = (raw or "").strip()

    m = re.match(r"^```(?:javascript|js)?\s*([\s\S]+?)\s*```$", code)
    if m:
        code = m.group(1).strip()
    return code, images


def _generate_pptx_via_sandbox(
    *,
    job_id: str,
    title: str,
    slides: list,
    user_id: str,
    chat_id: str | None,
    content_md: str,
    source_doc_name: str = "",
    prev_doc_name: str = "",
    question: str = "",
    llm_meta: dict | None = None,
    artifact_id: str | None = None,
    version: int | None = None,
) -> bool | None:
    """
    Build a PPTX via sandbox.doc_executor.build() and publish the result the
    same way doc_skill_worker.build_doc_skill_job does.

    ``artifact_id``/``version`` chain this build to an existing logical doc
    (revise/convert follow-up). For a fresh one-shot PPTX both are None and the
    durable Canvas handle becomes the file_id (audit row keeps artifact_id NULL,
    which /versions resolves via id==artifact_id).

    Returns:
        True  — success, result published.
        False — a terminal error occurred AFTER the doc was built (e.g. file
                write) and _fail() has already been called; caller must stop.
        None  — the sandbox could not build the doc; caller should fall back to
                the native _skill_generate() path instead of failing.
    """
    from sandbox.doc_executor import build as _sandbox_build

    logger.info(f"[docgen] worker PPTX sandbox path START | job={job_id} slides={len(slides)}")
    # On any build failure we return None (NOT False) so the caller falls back
    # to the native _skill_generate() path — which has its own self-healing
    # retries + guaranteed local renderer — instead of hard-failing the job.
    # Capture the (dominant) PPTX code-authoring LLM cost so it is deducted and
    # surfaced — the sandbox path returns early and never reaches the caller's
    # budget deduction, so it must account for its own cost here.
    _pptx_cost: dict = {}
    try:
        code, images = _author_pptxgenjs_code(title, slides, cost_sink=_pptx_cost)
    except Exception as exc:
        logger.warning(f"[docgen] worker pptxgenjs authoring failed — will fall back to native | job={job_id}: {exc}")
        return None
    if not code.strip():
        logger.warning(f"[docgen] worker pptxgenjs authoring produced no code — will fall back to native | job={job_id}")
        return None

    try:
        result = _sandbox_build(code, "pptx", images=images)
    except Exception as exc:
        logger.warning(f"[docgen] worker sandbox build crashed — will fall back to native | job={job_id}: {exc}")
        return None

    if not result.ok:
        logger.warning(f"[docgen] worker sandbox build failed — will fall back to native | job={job_id}: {result.error}")
        return None

    # ── Visual QA (feature-flagged, fails open) ───────────────────────────────
    # The PPTX sandbox path bypasses _skill_generate, so its visual-QA loop lives
    # here too. Critique the rendered slides against the ainxt brand contract; a
    # "revise" verdict re-authors the pptxgenjs code against the visual findings
    # and rebuilds. Bounded by DOC_VISION_QA_MAX_REVISIONS; any unavailability or
    # error just ships the current build (never blocks a deliverable). This is the
    # path that catches slide-level defects like low text/background contrast.
    _qa_max_rev = int(os.getenv("DOC_VISION_QA_MAX_REVISIONS", "2"))
    _qa_rev = 0
    _qa_verdict = "not_run"
    while _qa_rev < _qa_max_rev and result.page_images:
        try:
            from sandbox import doc_critic
            _crit = doc_critic.critique(list(result.page_images), "pptx")
        except Exception as _qae:
            logger.warning(f"[docgen] worker PPTX visual-QA raised (shipping as-is) | job={job_id}: {_qae}")
            break
        _qa_verdict = _crit.verdict
        if not _crit.needs_revision:
            logger.info(
                f"[docgen] worker PPTX visual-QA verdict={_crit.verdict} | job={job_id} "
                f"pages_reviewed={_crit.pages_reviewed}"
            )
            break
        logger.info(
            f"[docgen] worker PPTX visual-QA verdict=revise | job={job_id} "
            f"revision={_qa_rev + 1}/{_qa_max_rev} issues={len(_crit.issues)} "
            f"pages_reviewed={_crit.pages_reviewed} model={_crit.model or 'vision'}"
        )
        _publish_progress(job_id, 4, 6, "Generating File",
                          f"Polishing layout (visual review {_qa_rev + 1})…")
        try:
            from models.model_router import model_router as _mr_qa
            _repaired = doc_critic.strip_code_fence(
                _mr_qa.generate(doc_critic.build_repair_prompt("pptx", code, _crit.issues),
                                model_hint="complex") or ""
            )
        except Exception as _re:
            logger.warning(f"[docgen] worker PPTX visual-QA repair gen failed (shipping) | job={job_id}: {_re}")
            break
        if not _repaired or _repaired == code:
            logger.info(f"[docgen] worker PPTX visual-QA repair produced no change — shipping | job={job_id}")
            break
        try:
            _re_result = _sandbox_build(_repaired, "pptx", images=images)
        except Exception as _rbe:
            logger.warning(f"[docgen] worker PPTX visual-QA rebuild crashed — keeping prior build | job={job_id}: {_rbe}")
            break
        if not _re_result.ok or not _re_result.doc_bytes:
            logger.warning(f"[docgen] worker PPTX visual-QA rebuild failed — keeping prior build | job={job_id}: {_re_result.error}")
            break
        # Repaired build is good — adopt it and re-critique next loop.
        code, result = _repaired, _re_result
        _qa_rev += 1
    logger.info(f"[docgen] worker PPTX visual-QA done | job={job_id} verdict={_qa_verdict} revisions={_qa_rev}")

    _user_dir = user_doc_dir(user_id, chat_id)
    file_id  = str(_uuid_mod.uuid4())
    ext      = result.ext or "pptx"
    from tools.doc_generator import smart_filename
    if prev_doc_name:
        # ── UPDATE/FOLLOW-UP: version the previous doc name (-updated, -v2, …)
        # so each PPTX revision is distinct and content-derived — matching the
        # native (non-sandbox) doc paths. NOTE: prev_doc_name must NOT be folded
        # into source_doc_name; doing so makes smart_filename concatenate the
        # generic "Presentation" label twice (→ "Presentation-Presentation").
        _base = _versioned_basename(prev_doc_name)
    else:
        _base    = smart_filename(
        title=title, question=question or title,
        source_doc_name=source_doc_name, fmt_ext=ext,
    )
    filename = f"{_base}.{ext}"
    path     = os.path.join(_user_dir, f"{file_id}.{ext}")
    try:
        _atomic_write_bytes(path, result.doc_bytes)
    except Exception as exc:
        _fail(job_id, f"File write error: {exc}")
        return False

    preview_pages = 0
    for i, img in enumerate(result.page_images, start=1):
        try:
            _atomic_write_bytes(
                os.path.join(_user_dir, f"{file_id}.page-{i}.jpg"), img
            )
            preview_pages = i
        except Exception:
            break

    try:
        _save_audit(
            file_id=file_id, job_id=job_id, user_id=user_id, chat_id=chat_id,
            fmt=ext, title=title, filename=filename, file_path=path,
            content_md=content_md,
            artifact_id=artifact_id,
            version=version,
        )
    except Exception as exc:
        logger.warning(f"[docgen] worker audit save failed (non-fatal): {exc}")

    # ── Budget deduction (single point for the sandbox path) ───────────────────
    # This path returns True and short-circuits the caller BEFORE its own
    # increment_usage(), so it must deduct here. Combine the outline-structuring
    # cost (llm_meta, from the caller) with the PPTX code-authoring cost so the
    # total reflects true spend and is counted exactly once.
    _combined = dict(llm_meta or {})
    _merge_llm_cost(_combined, _pptx_cost, job_id=job_id, phase="pptx_authoring")
    _p_tokens   = int(_combined.get("tokens") or 0)
    _p_cost_usd = float(_combined.get("cost_usd") or 0.0)
    if user_id and user_id not in ("unknown", "default", "") \
            and (_p_tokens > 0 or _p_cost_usd > 0.0):
        try:
            from store.budget_store import increment_usage
            increment_usage(user_id, tokens=_p_tokens, cost_usd=_p_cost_usd)
            logger.info(
                f"doc_worker: budget deducted (pptx sandbox) | job={job_id} "
                f"user={user_id} tokens={_p_tokens} cost_usd={_p_cost_usd:.6f}"
            )
        except Exception as _bu_err:  # noqa: BLE001
            logger.warning(f"doc_worker: pptx sandbox budget update failed | job={job_id}: {_bu_err}")

    payload_out = {
        "status":        "done",
        "file_id":       file_id,
        "user_id":       str(user_id),   # owner — enforced by doc_job_status IDOR guard
        "artifact_id":   artifact_id or file_id,
        "filename":      filename,
        "format":        ext,
        "size":          len(result.doc_bytes),
        "preview_pages": preview_pages,
        "preview_url":   f"/ainxt/v1/api/docs/preview/{file_id}" if preview_pages else "",
        "meta": {
            "model":    _combined.get("model") or "unknown",
            "tokens":   _p_tokens,
            "in_tok":   int(_combined.get("in_tok") or 0),
            "out_tok":  int(_combined.get("out_tok") or 0),
            "cost_usd": _p_cost_usd,
            "engine":   "sandbox",
        },
    }
    _attach_summary_preview(
        payload_out, title=title, sections=slides,
        question=question or title, chat_id=chat_id, job_id=job_id,
    )
    _R.setex(f"doc:result:{job_id}", RESULT_TTL, json.dumps(payload_out))
    logger.info(
        f"[docgen] worker PPTX sandbox path DONE | job={job_id} "
        f"{filename} ({len(result.doc_bytes):,} bytes, {preview_pages} preview pages)"
    )
    return True


def generate_doc_job(payload: dict) -> None:
    """
    RQ job entry point.

    payload keys:
      job_id        str  — used as Redis result key suffix
      format        str  — docx | pptx | pdf | xlsx | txt | md
      title         str  — document title
      sections      list — [{heading, content, bullets, level}]
      content_md    str  — raw markdown (audit trail stored in Postgres)
      user_id       str  — requesting user id
      chat_id       str  — originating chat id (nullable)
      use_template  bool — use custom pptx template if available (pptx only)
    """
    job_id       = payload.get("job_id", "unknown")
    user_id      = payload.get("user_id", "unknown")
    chat_id      = payload.get("chat_id")

    from core.log_job_context import job_log_context
    with job_log_context(
        job_id=job_id, user_id=user_id, chat_id=chat_id or "",
        request_id=payload.get("request_id") or "",
        correlation_id=payload.get("correlation_id") or payload.get("request_id") or "",
        job_kind=payload.get("job_kind") or "structured",
        agent_id="doc_worker.generate_doc_job",
    ):
        # Guarantee a terminal status is ALWAYS published (Fix #34). Previously an
        # unhandled exception (e.g. during PPTX authoring/sandbox) let the RQ job die
        # WITHOUT writing doc:result:{job_id}, so the UI polled "Generating PPT…"
        # forever until its 30-min timeout. Now any escape publishes an error result.
        try:
            return _generate_doc_job_impl(payload)
        except Exception as exc:  # noqa: BLE001
            logger.exception(f"[docgen] worker generate_doc_job crashed | job={job_id} error={exc}")
            try:
                _fail(job_id, f"Document generation failed: {exc}")
            except Exception as _fe:
                logger.error(f"[docgen] worker could not publish failure | job={job_id} error={_fe}")
            return None


def _md_to_sections(md: str) -> list:
    """Parse authored/edited markdown into the worker's section structure
    deterministically (no LLM call) so supplied content_md renders verbatim.
    Used by the revise/convert path where the content already exists."""
    sections, cur = [], None

    def _new(heading="", level=2):
        s = {"heading": heading, "content": "", "bullets": [], "level": min(max(level, 2), 4)}
        sections.append(s)
        return s

    for line in (md or "").splitlines():
        s = line.strip()
        if not s:
            continue
        if s.startswith("#"):
            level = len(s) - len(s.lstrip("#"))
            heading = s.lstrip("#").strip()
            if level <= 1:        # top-level heading == the title; skip
                continue
            cur = _new(heading, level)
        elif s[:2] in ("- ", "* ") or s.startswith("• "):
            cur = cur or _new()
            cur["bullets"].append(s.lstrip("-*• ").strip())
        else:
            cur = cur or _new()
            cur["content"] = (cur["content"] + " " + s).strip() if cur["content"] else s
    if not sections:
        sections = [{"heading": "", "content": (md or "").strip(), "bullets": [], "level": 2}]
    return sections


def _generate_doc_job_impl(payload: dict) -> None:
    # Drafting already happened upstream (carried in llm_meta); this job only
    # builds the file. Total latency = upstream drafting + this build.
    _job_t0      = time.time()
    job_id       = payload.get("job_id", "unknown")
    fmt          = payload.get("format", "pdf")
    title        = payload.get("title", "Document")
    sections     = payload.get("sections") or []
    content_md   = payload.get("content_md") or ""
    user_id      = payload.get("user_id", "unknown")
    chat_id      = payload.get("chat_id")
    use_template = bool(payload.get("use_template", False))
    theme        = payload.get("theme", "dark_executive")
    llm_meta     = payload.get("llm_meta") or {}
    source_doc_name = payload.get("source_doc_name") or ""
    prev_doc_name   = payload.get("prev_doc_name") or ""
    # When a caller (revise engine) supplies an artifact_id, this build is a NEW
    # VERSION of an existing logical document; otherwise it's a fresh one-shot
    # (NULL artifact_id, version 1).
    _artifact_id    = payload.get("artifact_id") or None
    _version        = payload.get("version") or None

    _user_dir = user_doc_dir(user_id, chat_id)

    # ── Compliance gate ────────────────────────────────────────
    check_text = content_md or title
    try:
        chk = compliance_engine.validate_input(check_text[:4000])
        if chk.get("blocked"):
            _fail(job_id, "Content blocked by compliance policy")
            return
    except Exception as _ce_err:
        logger.warning(f"doc_worker: compliance check failed (fail-open): {_ce_err}")

    # ── If the caller supplied edited markdown (revise/convert path), render THAT
    #    verbatim — do NOT re-author from the instruction, which would discard the
    #    user's edits. Parse the markdown into sections.
    if not sections and content_md.strip():
        sections = _md_to_sections(content_md)
        logger.info(f"doc_worker: rendering {len(sections)} sections from supplied content_md | job={job_id}")

    # ── If still no sections, use LLM to structure from title/question ──
    if not sections:
        question_text = payload.get("question") or title
        logger.info(f"doc_worker: no sections — auto-structuring from title for job {job_id}")
        from workers._doc_preview import make_preview_callbacks
        _on_title, _on_section, _preview_done = make_preview_callbacks(job_id)
        result = _llm_structure(job_id, fmt, question_text,
                                on_section=_on_section, on_title=_on_title,
                                user_model_hint=payload.get("user_model_hint"))
        if result is None:
            return  # _fail already called
        sections, struct_meta, llm_title = result
        _preview_done()
        if not llm_meta:
            llm_meta = struct_meta
        # Use a generated title if the caller didn't supply a meaningful one.
        # Sanitize the LLM title (guards against raw-prompt/verb leaks); if it's
        # not trustworthy, name the doc from its actual content via the local model.
        if not title or title == "Document":
            _heur = _derive_title_from_question(question_text)
            _clean = _sanitize_llm_title(llm_title or "", question_text) if llm_title else ""
            if _clean and _clean != _heur:
                title = _clean
            else:
                title = _title_from_content(question_text, sections=sections,
                                            content_md=content_md, heuristic=_heur)
            logger.info(f"doc_worker: resolved document title: {_safe_log(title)} | job={job_id}")

    # ── PPTX engine selection (sandbox via doc_executor vs native) ────────
    use_pptx_sandbox = _should_use_sandbox_pptx(fmt)
    logger.info(
        f"[docgen] worker pptx engine resolved | job={job_id} "
        f"requested={_DOC_PPTX_ENGINE!r} sandbox={use_pptx_sandbox} fmt={fmt}"
    )

    # ── Enrich PPTX slides with AI-generated images (native path only) ────
    if fmt in ("pptx", "ppt", "powerpoint", "presentation", "slides") and not use_pptx_sandbox:
        sections = _enrich_with_images(sections)

    # ── PPTX sandbox path: short-circuits the rest of this function ───────
    # Returns None if the sandbox couldn't build — in which case we fall
    # through to the native _skill_generate() path below (which enriches the
    # slides with images first) so the job never hard-fails.
    if use_pptx_sandbox:
        _sandbox_ok = _generate_pptx_via_sandbox(
            job_id=job_id, title=title, slides=sections,
            user_id=user_id, chat_id=chat_id, content_md=content_md,
            source_doc_name=source_doc_name, prev_doc_name=prev_doc_name,
            question=payload.get("question") or title,
            llm_meta=llm_meta,
            artifact_id=_artifact_id, version=_version,
        )
        if _sandbox_ok is not None:
            return  # True → published; False → _fail already called
        logger.warning(
            f"[docgen] worker PPTX sandbox unavailable — falling back to native "
            f"skill path | job={job_id}"
        )
        sections = _enrich_with_images(sections)

    # ── Generate file via platform skillset ───────────────────
    domain = llm_meta.get("domain") if llm_meta else None

    # ── PLAIN-TEXT CSV short-circuit ──────────────────────────────────────
    # When the structuring step classified a `.csv` request as informational
    # (csv_mode == "plain_text"), write the content verbatim (one cell per
    # line) instead of fabricating tabular test data. Mirrors the same branch
    # in _generate_doc_from_question_impl so BOTH doc entry points behave
    # consistently. content_md was supplied by the caller (chat_worker) or
    # built above from sections.
    _csv_mode = (llm_meta.get("csv_mode") if llm_meta else "") or ""
    if fmt == "csv" and _csv_mode == "plain_text":
        if not content_md:
            content_md = f"# {title}\n\n" + "\n\n".join(
                "\n".join(p for p in (
                    f"## {s.get('heading','')}".rstrip(),
                    s.get("content", "") or "",
                    "\n".join(f"- {b}" for b in (s.get("bullets") or [])),
                ) if p) for s in sections
            )
        logger.info(
            f"[docgen] worker CSV plain_text mode — writing content verbatim "
            f"(bypassing test-data code-writer) | job={job_id} "
            f"content_chars={len(content_md)}"
        )
        data = _render_plaintext_csv(content_md)
        ext, mime = "csv", MIME_TYPES.get("csv", "text/csv")
    else:
        # Capture the code-writer's (dominant) LLM cost so the deduction below
        # reflects true doc cost, not just any pre-structured outline cost.
        _skill_cost: dict = {}
        skill_result = _skill_generate(
            job_id=job_id, fmt=fmt, question=payload.get("question") or title,
            title=title, sections=sections, domain=domain, theme=theme,
            cost_sink=_skill_cost,
        )
        if skill_result is None:
            return  # _fail already called inside _skill_generate
        data, ext, mime = skill_result
        if llm_meta is None:
            llm_meta = {}
        _merge_llm_cost(llm_meta, _skill_cost, job_id=job_id, phase="skill_generate")

    # ── Write to temp dir ──────────────────────────────────────
    file_id  = str(_uuid_mod.uuid4())
    # Context-aware naming — MUST match the marker name chosen in
    # chat_worker._handle_doc_generation so the download button name is stable:
    #   • UPDATE/follow-up (prev_doc_name) → version the previous name.
    #   • NEW doc from CHAT (no uploaded source_doc_name) → question="" so the
    #     LLM content title drives the name, not the generic chat prompt.
    from tools.doc_generator import smart_filename
    if prev_doc_name:
        _base = _versioned_basename(prev_doc_name)
    else:
        _question_hint = payload.get("question") or ""
        _from_chat     = not source_doc_name
        _base = smart_filename(
            title=title,
            question="" if _from_chat else _question_hint,
            source_doc_name=source_doc_name,
            fmt_ext=ext,
        )
    filename = f"{_base}.{ext}"
    path     = os.path.join(_user_dir, f"{file_id}.{ext}")

    try:
        _atomic_write_bytes(path, data)
    except Exception as exc:
        logger.error(f"doc_worker: file write failed for job {job_id}: {exc}")
        _fail(job_id, f"File write error: {exc}")
        return

    # ── Postgres audit record ──────────────────────────────────
    _save_audit(
        file_id=file_id, job_id=job_id, user_id=user_id, chat_id=chat_id,
        fmt=ext, title=title, filename=filename, file_path=path,
        content_md=content_md, artifact_id=_artifact_id, version=_version,
    )

    # ── Page-image preview (Canvas / inline preview) ───────────
    # The native skill path produces only the deliverable file; render page
    # images here so the in-app Canvas can SHOW the document (docx/xlsx/pdf/…).
    # Best-effort: render_preview_pages never raises and returns 0 if no
    # renderer is available, in which case the UI falls back to a text preview.
    preview_pages = 0
    try:
        from services.doc_preview import render_preview_pages
        preview_pages = render_preview_pages(path, ext, _user_dir, file_id)
        logger.info(
            f"doc_worker: preview render | job={job_id} fmt={ext} pages={preview_pages}"
        )
    except Exception as _pv_err:
        logger.warning(f"doc_worker: preview render failed (non-fatal) job={job_id}: {_pv_err}")

    # ── Token accounting + budget update ───────────────────────
    # Single source of truth for budget deduction on the doc generation path.
    # llm_meta is either:
    #   a) populated by _llm_structure() above (when sections were empty), OR
    #   b) passed in the payload from chat_worker (pre-structured path).
    # Either way, deduct here — chat_worker deliberately does NOT call
    # increment_usage so there is no double-counting.
    _tokens   = int(llm_meta.get("tokens") or 0)
    _cost_usd = float(llm_meta.get("cost_usd") or 0.0)
    if llm_meta and user_id and user_id not in ("unknown", "default", "") \
            and (_tokens > 0 or _cost_usd > 0.0):
        try:
            from store.budget_store import increment_usage
            increment_usage(
                user_id,
                tokens=_tokens,
                cost_usd=_cost_usd,
            )
            logger.info(
                f"doc_worker: budget deducted | job={job_id} user={user_id} "
                f"tokens={_tokens} cost_usd={_cost_usd:.6f}"
            )
        except Exception as _bu_err:
            logger.warning(f"doc_worker: budget update failed for job {job_id}: {_bu_err}")

    in_tok  = int(llm_meta.get("in_tok") or 0)
    out_tok = int(llm_meta.get("out_tok") or 0)

    if not (in_tok or out_tok):
        total = int(llm_meta.get("tokens") or 0)
        in_tok = total
        out_tok = 0

    tokens  = in_tok + out_tok
    cost    = float(llm_meta.get("cost_usd") or 0.0)
    latency = round(float(llm_meta.get("latency") or 0.0) + (time.time() - _job_t0), 3)
    model   = llm_meta.get("model") or "unknown"

    # ── Publish result ─────────────────────────────────────────
    # artifact_id is the durable handle the Canvas uses for version history +
    # AI edits. One-shot docs have no explicit artifact_id, so the handle IS the
    # file_id (the audit row was saved with artifact_id NULL; the /versions
    # endpoint resolves NULL via id==artifact_id). Surfacing it here is what
    # makes the "Edit in Canvas" button appear.
    result = {
        "status":        "done",
        "file_id":       file_id,
        "user_id":       str(user_id),   # owner — enforced by doc_job_status IDOR guard
        "artifact_id":   _artifact_id or file_id,
        "filename":      filename,
        "format":        ext,
        "size":          len(data),
        "preview_pages": preview_pages,
        "preview_url":   f"/ainxt/v1/api/docs/preview/{file_id}" if preview_pages else "",
        "meta": {
            "model":   model,
            "tokens":  tokens,
            "in_tok":  in_tok,
            "out_tok": out_tok,
            "cost_usd": cost,
            "latency": latency,
        } if llm_meta else {},
    }
    _attach_summary_preview(
        result, title=title, sections=sections,
        question=payload.get("question") or title,
        chat_id=chat_id, job_id=job_id,
    )
    _R.setex(f"doc:result:{job_id}", RESULT_TTL, json.dumps(result))
    logger.info(f"doc_worker: job {job_id} done — {_safe_log(filename)} ({len(data)} bytes)")

    # ── BUGFIX: persist MD edit-session so follow-up "update" requests
    # regenerate in the ORIGINAL format (docx/pdf/xlsx/…) instead of
    # silently falling through to .md. Path A (_generate_doc_from_question_impl)
    # already does this at line 1830; Path B (slash-command / pre-structured
    # generation via chat_worker → generate_doc_job) was missing the same write,
    # which caused `_is_doc_edit_followup` to return False on update and the
    # edit branch in doc_worker_agent to never see `original_format=docx`.
    # Without this session, the edit either reverts to plain chat OR — if a
    # later `_handle_md_generation` call writes an md-session — gets locked
    # to `original_format="md"` forever.
    if chat_id:
        try:
            _save_md_session_for_chat(
                chat_id=chat_id,
                job_id=job_id,
                file_id=file_id,
                title=title,
                domain=domain,
                sections=sections,
                content_md=content_md,
                filename=filename,
                file_path=path,
                original_format=ext,  # the REAL output extension (docx/pdf/xlsx/…)
                question=payload.get("question") or title,
                llm_meta=llm_meta or {},
            )
        except Exception as _sess_err:
            logger.warning(
                f"doc_worker: md-session persist failed for job {job_id}: {_sess_err}"
            )


def generate_doc_from_question(payload: dict) -> None:
    """
    RQ entry point for question-driven doc generation (frontend path).
    Uses LLM to structure raw user question into sections, then generates the file.

    payload keys:
      job_id          str  — Redis result key suffix
      question        str  — raw user question/request
      format          str  — docx | pptx | pdf | xlsx | txt | md
      user_id         str  — requesting user id
      chat_id         str  — originating chat id (nullable)
      source_doc_name str  — uploaded file name for context-aware filename generation
      prev_doc_name   str  — filename of the previously generated doc in this chat (follow-up requests)
      chat_last_response str — verbatim last assistant reply (preserve all content incl. code blocks)
    """
    job_id          = payload.get("job_id", "unknown")
    question        = (payload.get("question") or "").strip()
    fmt             = payload.get("format", "pdf")
    user_id         = payload.get("user_id", "unknown")
    chat_id         = payload.get("chat_id")
    source_doc_name = payload.get("source_doc_name") or ""
    prev_doc_name   = payload.get("prev_doc_name") or ""
    attachment_ids  = payload.get("attachment_ids") or []
    chat_context    = (payload.get("chat_context") or "").strip()  # conversation history as context
    chat_last_response = (payload.get("chat_last_response") or "").strip()  # verbatim last reply
    doc_source_scope_hint = (payload.get("doc_source_scope") or "").strip() or None  # gateway CIL scope
    all_conversation = bool(payload.get("all_conversation"))  # P8: include full conversation

    from core.log_job_context import job_log_context
    with job_log_context(
        job_id=job_id, user_id=user_id, chat_id=chat_id or "",
        request_id=payload.get("request_id") or "",
        correlation_id=payload.get("correlation_id") or payload.get("request_id") or "",
        job_kind=payload.get("job_kind") or "doc",
        agent_id="doc_worker.generate_doc_from_question",
    ):
        return _generate_doc_from_question_impl(
            payload=payload, job_id=job_id, question=question, fmt=fmt,
            user_id=user_id, chat_id=chat_id, source_doc_name=source_doc_name,
            prev_doc_name=prev_doc_name, attachment_ids=attachment_ids,
            chat_context=chat_context, chat_last_response=chat_last_response,
            doc_source_scope_hint=doc_source_scope_hint,
            all_conversation=all_conversation,
        )


def _generate_doc_from_question_impl(*, payload, job_id, question, fmt, user_id,
                                     chat_id, source_doc_name, attachment_ids,
                                     chat_context, prev_doc_name="", chat_last_response="",
                                     doc_source_scope_hint=None, all_conversation=False):
    # Drafting happens inside this job, so whole-job wall-clock is the total.
    _q_job_t0 = time.time()
    # ── BUG #3 FIX: Strip slash command prefixes from question ────────────────
    # The frontend sends the raw question including the command prefix (e.g.
    # "/pdf generate a report"). Strip it so the LLM sees a clean topic.
    _CMD_PREFIX_RE = re.compile(
        r"^/(pdf|docx?|word|doc|xlsx?|excel|csv|pptx?|pptagent|ppt|md|txt|text)\s*",
        re.IGNORECASE,
    )
    question = _CMD_PREFIX_RE.sub("", question).strip()

    logger.info(f"[docgen] worker START generate_doc_from_question | job={job_id} fmt={fmt} user={user_id}")
    logger.info(f"[docgen] worker question preview: {_safe_log(question)}")

    if not question:
        logger.warning(f"[docgen] worker ABORT job={job_id} — no question provided")
        _fail(job_id, "No question provided")
        return

    _user_dir = user_doc_dir(user_id, chat_id)

    # ── Compliance gate on question ────────────────────────────
    logger.info(f"[docgen] worker STEP 1/6 — compliance gate | job={job_id}")
    _publish_progress(job_id, 1, 6, "Compliance Check", "Validating content safety…")
    try:
        chk = compliance_engine.validate_input(question[:4000])
        if chk.get("blocked"):
            logger.warning(f"[docgen] worker BLOCKED by compliance | job={job_id}")
            _fail(job_id, "Content blocked by compliance policy")
            return
        logger.info(f"[docgen] worker compliance gate PASSED | job={job_id}")
    except Exception as _ce_err:
        logger.warning(f"[docgen] worker compliance check failed (fail-open): {_ce_err}")

    # ── Fetch parsed content from uploaded attachments (MULTI-DOC) ────────────
    # Loop over EVERY uploaded file — not just the first — so "combine these 3
    # PDFs" / "extract from all my files" work. Each file is redacted (PCI/PII)
    # independently and labeled so the LLM can attribute facts to a source.
    # For a single file, _parsed_attachment is that file's content verbatim
    # (preservation/convert branches depend on this). For multiple files it is a
    # labeled, per-file budgeted concatenation.
    _parsed_attachment = ""
    _attachment_filename = source_doc_name
    _source_ext = ""   # extension of the FIRST/primary file (drives format inference)
    _attachment_count = 0
    if attachment_ids:
        logger.info(f"[docgen] worker STEP 1b/6 — fetching attachment content | job={job_id} ids={attachment_ids}")
        _publish_progress(job_id, 1, 6, "Reading Uploaded Files", "Fetching parsed file content…")
        # Per-file char budget so many docs don't blow the context window.
        _n = max(1, len(attachment_ids))
        _per_file_budget = 12000 if _n == 1 else max(2000, 24000 // _n)
        _blocks: list[str] = []
        _first_clean = ""   # raw (unlabeled) content of the primary file
        try:
            from db.database import SessionLocal as _AttSL
            from db.models import ChatAttachment as _AttModel
            _attdb = _AttSL()
            try:
                for _aid in attachment_ids:
                    _att = _attdb.query(_AttModel).filter(_AttModel.id == _aid).first()
                    if not (_att and _att.parsed_text):
                        continue
                    _att_owner = getattr(_att, "user_id", None)
                    if _att_owner and user_id and str(_att_owner) != str(user_id):
                        logger.warning(
                            f"[docgen] worker skipping attachment {_aid} — owner "
                            f"mismatch (not owned by user={user_id}) | job={job_id}"
                        )
                        continue
                    _raw_att = _att.parsed_text or ""
                    _fname = _att.file_name or source_doc_name or f"file-{_attachment_count+1}"
                    _ext = (_att.file_type or "").lower().strip(".")
                    # Redact PCI/PII before sending to the LLM proxy.
                    try:
                        _r = compliance_engine.validate_input(_raw_att[:_per_file_budget])
                        _clean = _r.get("redacted_text") or _raw_att[:_per_file_budget]
                        if _r.get("was_redacted"):
                            logger.info(
                                f"[docgen] worker attachment redacted types={_r.get('redacted_types')} "
                                f"file={_safe_log(_fname)} | job={job_id}"
                            )
                    except Exception:
                        _clean = _raw_att[:_per_file_budget]
                    if _attachment_count == 0:
                        # First file is the "primary" — drives filename + format inference.
                        _attachment_filename = _fname
                        _source_ext = _ext
                        _first_clean = _clean
                    _blocks.append(f"===== SOURCE FILE {_attachment_count+1}: {_fname} =====\n{_clean}")
                    _attachment_count += 1
            finally:
                _attdb.close()
        except Exception as _att_err:
            logger.warning(f"[docgen] worker attachment fetch failed (continuing without): {_att_err}")

        if _attachment_count == 1:
            # Single file → raw content so preservation/convert see it verbatim.
            _parsed_attachment = _first_clean
        elif _attachment_count > 1:
            _parsed_attachment = "\n\n".join(_blocks)
        logger.info(
            f"[docgen] worker attachments loaded: count={_attachment_count} "
            f"primary={_attachment_filename!r} ext={_source_ext!r} "
            f"({len(_parsed_attachment):,} chars) | job={job_id}"
        )

    # ── BUG #1 FIX: Infer format from source file extension when format is ambiguous ──
    # When the user uploads a CSV/XLSX and asks to "add rows" or "expand the file",
    # _detect_doc_format returns "pdf" (no format keyword in the text).
    # Use the uploaded file's extension to pick a sensible default format.
    _EXT_TO_FMT = {
        # CSV uploads keep CSV output (per user preference: mirror upload format).
        # Previously this upgraded csv→xlsx, which silently changed the user's
        # output format when they uploaded a .csv and asked to expand it.
        "csv":  "csv",
        "xlsx": "xlsx",
        "xls":  "xlsx",
        "docx": "docx",
        "doc":  "docx",
        "pptx": "pptx",
        "ppt":  "pptx",
        "txt":  "txt",
        "md":   "md",
        "pdf":  "pdf",
    }
    # ── Resolve the DOC PLAN — the single authority (services.doc_router) ─────
    # One call decides intent (generate/summarize/convert/extract/revise),
    # attachment reproduce-vs-generate (preserve), which prior artifact a
    # follow-up targets, and whether the request is too ambiguous to act on
    # (needs_clarification). The client's doc_intent/format are passed as HINTS
    # only — the backend is authoritative. Fail-open: defaults to "generate".
    from services.doc_router import resolve_doc_plan as _resolve_doc_plan
    _plan = _resolve_doc_plan(
        question,
        has_attachments=bool(_parsed_attachment),
        attachment_filename=_attachment_filename,
        attachment_count=_attachment_count,
        chat_id=chat_id or "", user_id=user_id,
        format_hint=fmt,
        intent_hint=(payload.get("doc_intent") or "").lower().strip() or None,
        hint_confidence=payload.get("doc_confidence"),
        has_chat_context=bool(chat_context or chat_last_response),
        source_scope_hint=doc_source_scope_hint,  # gateway CIL scope — eliminates double-classification
    )
    _intent = _plan.intent
    _is_preserve_intent = _plan.preserve and bool(_parsed_attachment)
    _from_chat_source = (_plan.source_scope == "chat")

    # ── Resolve the OUTPUT FORMAT (user intent wins over source extension) ────
    # The classifier extracts the TARGET format from the user's words (e.g.
    # "convert the attached presentation to PDF" → pdf). Honour it.
    if _plan.format:
        fmt = _plan.format
    # Fallback ONLY: no explicit target format AND not an explicit convert →
    # mirror the uploaded file's own format (e.g. "add rows to my .csv" with no
    # format word). NEVER override an explicit convert target (that was the bug
    # where a .pptx upload + "convert to PDF" produced a .pptx again).
    elif fmt == "pdf" and _intent != "convert" and _source_ext in _EXT_TO_FMT:
        _inferred_fmt = _EXT_TO_FMT[_source_ext]
        if _inferred_fmt != "pdf":
            logger.info(
                f"[docgen] worker format inferred from source ext: "
                f"pdf → {_inferred_fmt} (source={_source_ext!r}) | job={job_id}"
            )
            fmt = _inferred_fmt

    # ── Clarify hook: ask the user rather than guess ──────────────────────────
    # When the request plausibly targets a prior doc but we can't confidently
    # resolve which one (or new-vs-existing), publish a clarify result and stop.
    # The frontend renders quick-reply buttons; the choice resumes via
    # POST /docs/clarify-resume with a disambiguated intent/artifact_id.
    if _plan.needs_clarification:
        logger.info(
            f"[docgen] worker doc plan needs clarification ({_plan.reason}) | "
            f"job={job_id} options={len(_plan.clarify_options)}"
        )
        _clarify(job_id, _plan.clarify_question, _plan.clarify_options,
                 original_question=question, fmt=fmt, doc_intent=_intent,
                 attachment_ids=list(attachment_ids or []),
                 chat_id=chat_id or "",
                 user_model_hint=payload.get("user_model_hint") or "auto")
        return

    # ── REVISE / CONVERT-OF-PRIOR delegation ──────────────────────────────────
    # A revise (edit) or a convert of a PRIOR generated doc (no uploaded file)
    # must act on that doc's stored source — revise edits it, convert just
    # retargets the format — NOT re-author from the instruction. Delegate to the
    # shared reviser, which loads content_md and enqueues a versioned rebuild. We
    # forward the new job_id under THIS job's result key so the frontend poller
    # (already polling job_id) transparently follows the result.
    _delegate_revise = bool(
        _plan.target_artifact_id and not _parsed_attachment
        and _intent in ("revise", "convert")
    )
    if _delegate_revise:
        try:
            from services.doc_reviser import revise as _revise
            _tgt_fmt = _plan.format if (_intent == "convert" or _plan.format != fmt) else None
            _rv = _revise(
                artifact_id=_plan.target_artifact_id,
                instruction=(question if _intent == "revise" else "Convert to the requested format without changing the content."),
                user_id=user_id, chat_id=chat_id or None,
                target_format=_tgt_fmt,
                user_model_hint=payload.get("user_model_hint") or "auto",
            )
            if _rv.get("ok"):
                _R.setex(
                    f"doc:result:{job_id}", 3600,
                    json.dumps({"status": "redirect", "job_id": _rv["job_id"]})
                )
                logger.info(
                    f"[docgen] worker revise delegated | job={job_id} → "
                    f"{_rv['job_id']} artifact={_plan.target_artifact_id} v{_rv.get('version')}"
                )
                return
            logger.warning(
                f"[docgen] worker revise delegation failed ({_rv.get('error')}) — "
                f"falling back to generate | job={job_id}"
            )
            _intent = "generate"
        except Exception as _rex:  # noqa: BLE001
            logger.warning(f"[docgen] worker revise delegation error — generating instead: {_rex}")
            _intent = "generate"

    logger.info(
        f"[docgen] worker resolved doc intent={_intent!r} preserve={_is_preserve_intent} "
        f"target_artifact={_plan.target_artifact_id!r} conf={_plan.confidence:.2f} | job={job_id}"
    )

    # ── Call LLM to structure content into sections ────────────
    logger.info(
        f"[docgen] worker STEP 2/6 — LLM structuring | job={job_id} fmt={fmt} "
        f"preserve={_is_preserve_intent} has_attachment={bool(_parsed_attachment)}"
    )
    _publish_progress(job_id, 2, 6, "Drafting", "Streaming sections…")

    from workers._doc_preview import make_preview_callbacks
    _on_title, _on_section, _preview_done = make_preview_callbacks(job_id)

    # ── COMPARE: assemble two labeled sources (uploads + prior generated docs) ──
    # The plan resolved which prior artifact(s) fill the gap when fewer than two
    # files were uploaded. Load each prior doc's stored content_md and append it
    # as an additional labeled SOURCE so the comparison prompt sees both sides.
    _compare_sources = ""
    if _intent == "compare":
        _blocks_cmp = []
        if _parsed_attachment:
            _blocks_cmp.append(_parsed_attachment)   # already labeled per file
        try:
            from services.doc_context import load_latest_source as _lls_cmp
            _next_idx = (_attachment_count or 0) + 1
            # A clarify-resume pins the exact prior doc the user chose to compare
            # against — honour it first, then any auto-resolved ones.
            _cmp_ids = list(_plan.compare_prior_artifact_ids or [])
            _pinned_cmp = payload.get("compare_artifact_id")
            if _pinned_cmp and _pinned_cmp not in _cmp_ids:
                _cmp_ids.insert(0, _pinned_cmp)
            for _aid in _cmp_ids:
                _ref = _lls_cmp(_aid, user_id)
                if _ref and (_ref.content_md or "").strip():
                    _blocks_cmp.append(
                        f"===== SOURCE FILE {_next_idx}: {_ref.title} "
                        f"({_ref.format}, v{_ref.version}) =====\n{_ref.content_md[:12000]}"
                    )
                    _next_idx += 1
        except Exception as _cmp_err:  # noqa: BLE001
            logger.warning(f"[docgen] worker compare prior-doc load failed: {_cmp_err}")
        _compare_sources = "\n\n".join(_blocks_cmp)
        logger.info(
            f"[docgen] worker compare sources assembled | job={job_id} "
            f"uploads={_attachment_count} priors={len(_plan.compare_prior_artifact_ids or [])} "
            f"chars={len(_compare_sources)}"
        )

    # Intent-driven override prompts (summarize/extract/convert/compare) all feed
    # the same _llm_structure call — build the prompt here, then dispatch below.
    # IMPORTANT: every override_prompt MUST end with this JSON schema suffix so
    # Claude/GPT returns structured JSON instead of prose Markdown. Without it
    # _parse_llm_json receives Markdown and crashes with JSONDecodeError.
    _JSON_SCHEMA_SUFFIX = (
        "\n\nRespond with ONLY valid JSON — no markdown fences, no explanation.\n\n"
        "JSON SCHEMA (follow exactly):\n"
        "{\n"
        '  "title": "<concise 3-7 word noun phrase — NO verbs like generate/create/write/make>",\n'
        '  "domain": "<single keyword: general|payments|banking|ai|fintech|healthcare|'
        'government|education|retail|hr|legal|security|default>",\n'
        '  "sections": [\n'
        "    {\n"
        '      "heading": "<section heading>",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<full section body — paragraphs separated by \\n\\n>",\n'
        '      "bullets": ["<bullet point>"],\n'
        '      "callout": {"label": "<short label>", "text": "<callout text or empty>"},\n'
        '      "table": null\n'
        "    }\n"
        "  ]\n"
        "}\n\n"
        "RULES:\n"
        "1. title: professional noun phrase derived from the SOURCE content — NOT from the user request verb.\n"
        "2. Minimum 3 sections, each with substantive content from the source.\n"
        "3. Do NOT invent content not present in the source.\n"
        "4. Output RAW JSON ONLY. Any non-JSON text is a failure.\n"
    )
    _intent_prompt = None
    if _intent == "compare" and _compare_sources:
        # Structured diff of two documents → a comparison report.
        _intent_prompt = (
            f"You are an expert analyst. COMPARE the SOURCE DOCUMENTS below and produce "
            f"a clear, well-structured COMPARISON REPORT as a {fmt.upper()} document. "
            f"Structure it as:\n"
            f"1. Overview — what each document is, in one line.\n"
            f"2. Key Similarities — points both share.\n"
            f"3. Key Differences — a side-by-side table (attribute | Document 1 | "
            f"Document 2) covering the material differences.\n"
            f"4. Notable Additions / Removals — content present in one but not the other.\n"
            f"5. Assessment & Recommendation — which is stronger for the user's goal, or "
            f"what changed and why it matters.\n"
            f"Attribute every fact to its source document. Do NOT invent content that is "
            f"not present in the sources.\n\n"
            f"SOURCE DOCUMENTS:\n{_compare_sources[:20000]}\n\nUser request: {question}"
            + _JSON_SCHEMA_SUFFIX
        )
    elif _intent == "summarize" and (_parsed_attachment or chat_context or chat_last_response):
        # Summarize existing content → a summary document, not a fresh essay.
        _src = _parsed_attachment or chat_last_response or chat_context or ""
        _intent_prompt = (
            f"You are an expert analyst. Produce a clear, well-structured SUMMARY as a "
            f"{fmt.upper()} document of the SOURCE below. Capture the key points, decisions, "
            f"figures and conclusions faithfully — do NOT invent content or add topics that "
            f"are not in the source. Use concise headings and bullet points.\n\n"
            f"SOURCE:\n{_src[:12000]}\n\nUser request: {question}"
            + _JSON_SCHEMA_SUFFIX
        )
    elif _intent == "extract" and _parsed_attachment:
        # Extract/merge from uploaded doc(s) (already aggregated) into one doc.
        _intent_prompt = (
            f"You are an expert analyst. From the SOURCE DOCUMENT(S) below, EXTRACT and "
            f"CONSOLIDATE the information the user asks for into a single well-structured "
            f"{fmt.upper()} document. Merge related facts across sources, remove duplication, "
            f"and cite which source a figure came from when it matters. Do NOT fabricate data "
            f"that is not present in the sources.\n\n"
            f"SOURCE DOCUMENT(S):\n{_parsed_attachment[:16000]}\n\nUser request: {question}"
            + _JSON_SCHEMA_SUFFIX
        )
    elif _intent == "extract" and _from_chat_source and (chat_context or chat_last_response):
        # Extract from chat context (no attachment) — user said "extract this into a pdf"
        # after a conversation. Treat like summarize but with extraction framing.
        _src = chat_last_response or chat_context or ""
        _intent_prompt = (
            f"You are an expert analyst. From the CONVERSATION below, EXTRACT and "
            f"CONSOLIDATE the key information the user asks for into a single well-structured "
            f"{fmt.upper()} document. Capture key facts, decisions, code snippets, and insights. "
            f"Do NOT fabricate data not present in the conversation.\n\n"
            f"CONVERSATION:\n{_src[:12000]}\n\nUser request: {question}"
            + _JSON_SCHEMA_SUFFIX
        )
    elif _intent == "convert" and _parsed_attachment:
        # Faithful format change → reuse the preservation prompt.
        _intent_prompt = _build_preservation_prompt(
            parsed_content=_parsed_attachment[:10000],
            source_filename=_attachment_filename, target_format=fmt,
        )

    if _intent_prompt is not None:
        logger.info(f"[docgen] worker using {_intent.upper()} prompt | job={job_id}")
        doc_result = _llm_structure(job_id, fmt, question, override_prompt=_intent_prompt,
                                    on_section=_on_section, on_title=_on_title,
                                    user_model_hint=payload.get("user_model_hint"))
    elif _is_preserve_intent:
        # PRESERVATION path: faithfully convert uploaded file content → no hallucination
        _pres_prompt = _build_preservation_prompt(
            parsed_content=_parsed_attachment[:10000],
            source_filename=_attachment_filename,
            target_format=fmt,
        )
        logger.info(f"[docgen] worker using PRESERVATION prompt | job={job_id} file={_safe_log(_attachment_filename)}")
        _umh = payload.get("user_model_hint")
        doc_result = _llm_structure(job_id, fmt, question, override_prompt=_pres_prompt,
                                    on_section=_on_section, on_title=_on_title,
                                    user_model_hint=_umh)
    elif _parsed_attachment:
        # GENERATION WITH FILE CONTEXT: user wants new content using uploaded file as seed
        # e.g. "add 1000 rows following this pattern"
        _ctx_question = (
            f"Source file ({_attachment_filename}) content for reference:\n"
            f"{_parsed_attachment[:6000]}\n\n"
            f"User request: {question}"
        )
        logger.info(f"[docgen] worker using GENERATION prompt with attachment context | job={job_id}")
        _umh = payload.get("user_model_hint")
        # This is a GENERATE-with-file-context request (preserve=false), e.g.
        # "add 1000 rows following this pattern". For xlsx the flat-dataset
        # decision depends on WHAT was uploaded:
        #   • a DATA file (.csv/.xlsx/.xls) → the user is working with tabular
        #     data, so allow the test-data probe (bulk row generation works).
        #   • anything else (a report doc, pdf, etc.) → keep the narrative
        #     report format by default so "add a section to my report" does NOT
        #     flatten a McKinsey workbook into a raw dataset.
        # Either way _llm_structure honours an EXPLICIT flat-data request
        # (_wants_flat_dataset) inside the xlsx branch.
        _src_is_data_file = _source_ext in ("csv", "xlsx", "xls")
        doc_result = _llm_structure(job_id, fmt, _ctx_question,
                                    on_section=_on_section, on_title=_on_title,
                                    user_model_hint=_umh,
                                    xlsx_allow_testdata=_src_is_data_file)
    elif (chat_last_response and _intent != "generate"
          and (_from_chat_source or _intent in ("convert", "summarize", "extract"))):
        # VERBATIM CHAT PRESERVATION: only when the user is acting on the
        # conversation. The authority is the LLM intent classifier — either it
        # tagged the source as the chat (source_scope="chat", e.g. "summarize
        # this chat session into a document") or the intent is a chat-acting one
        # (convert/summarize/extract "this reply"). A fresh "generate a doc on
        # <topic>" (source_scope="none") must NOT reproduce the previous answer —
        # doing so bled a prior Python-code reply into a "PDF on UPI growth".
        _chat_pres_prompt = _build_chat_preservation_prompt(
            last_response=chat_last_response[:_CHAT_PRESERVE_MAX_CHARS],
            user_request=question,
            target_format=fmt,
        )
        logger.info(
            f"[docgen] worker using CHAT-PRESERVATION prompt | job={job_id} "
            f"intent={_intent} scope={_plan.source_scope!r} resp_len={len(chat_last_response)}"
        )
        _umh = payload.get("user_model_hint")
        doc_result = _llm_structure(job_id, fmt, question, override_prompt=_chat_pres_prompt,
                                    on_section=_on_section, on_title=_on_title,
                                    user_model_hint=_umh)
    elif chat_context and (_from_chat_source or _intent in ("convert", "summarize", "extract")):
        # GENERATION FROM CHAT CONTEXT: the user wants a doc built from the
        # conversation (e.g. "summarize this chat into a report", "make a doc of
        # what we discussed"). Gated by the LLM classifier's source_scope="chat"
        # so a topic-only request never pulls unrelated prior turns in. The
        # chat_context is STRICTLY this chat's history (rolling summary + recent
        # turns assembled in gateway.py) — never the knowledge base or codebase.
        # P8: all_conversation flag → comprehensive prompt covering ALL topics.
        if all_conversation:
            _chat_ctx_question = (
                f"The following is a COMPLETE conversation between a user and an AI assistant.\n"
                f"Generate a comprehensive, well-structured {fmt.upper()} document that captures "
                f"ALL key information, insights, and content from this conversation. "
                f"Include ALL topics discussed — do not omit any subject matter. "
                f"Use clear section headings for each distinct topic.\n\n"
                f"Conversation:\n{chat_context[:20000]}\n\n"
                f"User request: {question}"
            )
        else:
            _chat_ctx_question = (
                f"The following is a conversation between a user and an AI assistant.\n"
                f"Generate a well-structured {fmt.upper()} document that captures the key "
                f"information, insights, and content from this conversation.\n\n"
                f"Conversation:\n{chat_context[:8000]}\n\n"
                f"User request: {question}"
            )
        logger.info(
            f"[docgen] worker using GENERATION prompt with chat context | job={job_id} "
            f"intent={_intent} scope={_plan.source_scope!r} context_len={len(chat_context)} "
            f"all_conv={all_conversation}"
        )
        _umh = payload.get("user_model_hint")
        doc_result = _llm_structure(job_id, fmt, _chat_ctx_question,
                                    on_section=_on_section, on_title=_on_title,
                                    user_model_hint=_umh)
    else:
        # PURE GENERATION: no attachment, no chat context — generate from question alone
        _umh = payload.get("user_model_hint")
        doc_result = _llm_structure(job_id, fmt, question,
                                    on_section=_on_section, on_title=_on_title,
                                    user_model_hint=_umh)

    if doc_result is None:
        logger.error(f"[docgen] worker LLM structuring returned None — aborting | job={job_id}")
        return  # _fail already called
    sections, llm_meta, llm_title = doc_result
    _preview_done()
    logger.info(f"[docgen] worker LLM structuring DONE | job={job_id} sections={len(sections)} model={llm_meta.get('model')} tokens={llm_meta.get('tokens')}")

    # ── Check point A: bail out if user cancelled while LLM was running ──
    from core.generation_registry import is_stopped_redis
    if is_stopped_redis(job_id):
        logger.info(f"[docgen] worker job {job_id} cancelled after _llm_structure — stopping")
        return

    # ── PPTX engine selection (sandbox via doc_executor vs native) ────────
    use_pptx_sandbox = _should_use_sandbox_pptx(fmt)
    logger.info(
        f"[docgen] worker pptx engine resolved | job={job_id} "
        f"requested={_DOC_PPTX_ENGINE!r} sandbox={use_pptx_sandbox} fmt={fmt}"
    )

    # ── Enrich PPTX slides with AI-generated images (native path only) ────
    if fmt in ("pptx", "ppt", "powerpoint", "presentation", "slides") and not use_pptx_sandbox:
        logger.info(f"[docgen] worker STEP 3/6 — enriching PPTX slides with images | job={job_id}")
        _publish_progress(job_id, 3, 6, "Image Enrichment", "Generating slide images…")
        sections = _enrich_with_images(sections)
        logger.info(f"[docgen] worker image enrichment DONE | job={job_id}")
        if is_stopped_redis(job_id):
            logger.info(f"[docgen] worker job {job_id} cancelled after _enrich_with_images — stopping")
            return
    else:
        reason = ("sandbox path (images handled in-container)"
                  if use_pptx_sandbox else "not PPTX")
        logger.info(f"[docgen] worker STEP 3/6 — skipping image enrichment ({reason}) | job={job_id}")
        _publish_progress(job_id, 3, 6, "Image Enrichment",
                          "Handled in sandbox" if use_pptx_sandbox else "Skipped (not PPTX)")

    # Resolve the document title (also drives the filename).
    #   1. LLM-generated content title from the structuring pass — but ALWAYS run
    #      it through _sanitize_llm_title so a raw-prompt / request-verb leak
    #      (e.g. "Summarize this doc") never becomes the title.
    #   2. If that yields nothing trustworthy, NAME THE DOC FROM ITS CONTENT using
    #      the fast local model (_title_from_content) — this is how Claude/GPT
    #      title documents and is robust to typos and vague prompts.
    #   3. _derive_title_from_question (regex heuristic) is only the last-resort
    #      fallback inside those helpers if the model is unavailable.
    _from_chat_doc = bool(chat_last_response or chat_context)
    _clean_llm_title = _sanitize_llm_title(llm_title or "", question) if llm_title else ""
    # A sanitized title is "trustworthy" only if it wasn't downgraded to the
    # heuristic derivation of the (possibly junky) prompt.
    _heuristic = _derive_title_from_question(question)
    if _clean_llm_title and _clean_llm_title != _heuristic:
        title = _clean_llm_title
    else:
        # Author-quality title from the actual document content.
        title = _title_from_content(question, sections=sections, heuristic=_heuristic)
    logger.info(
        f"[docgen] worker document title: {title!r} "
        f"(llm_title={llm_title!r}, clean={_clean_llm_title!r}, from_chat={_from_chat_doc})"
    )

    def _section_md(s):
        # Build markdown for one section, appending any preserved code as a fenced
        # block so the audit trail (and md/txt fallback rendering) keeps code intact.
        parts = [
            f"## {s.get('heading','')}".rstrip(),
            s.get("content", s.get("key_message", "")) or "",
        ]
        bl = s.get("bullets") or []
        if bl:
            parts.append("\n".join(f"- {b}" for b in bl))
        code = (s.get("code") or "").strip()
        if code:
            lang = (s.get("language") or "").strip()
            parts.append(f"```{lang}\n{code}\n```")
        return "\n".join(p for p in parts if p)

    content_md = f"# {title}\n\n" + "\n\n".join(_section_md(s) for s in sections)

    # ── PPTX sandbox path: short-circuits the rest of this function ───────
    # Returns None if the sandbox couldn't build — fall through to the native
    # skill path (enriching slides with images first) so the job never fails.
    if use_pptx_sandbox:
        _publish_progress(job_id, 4, 6, "Generating File", "Building PPTX in sandbox…")
        _sbx_artifact = payload.get("artifact_id") or _plan.target_artifact_id or None
        _sbx_version  = None
        if _sbx_artifact and _plan.target_artifact_id == _sbx_artifact and _plan.target_version:
            _sbx_version = int(_plan.target_version) + 1
        _sandbox_ok = _generate_pptx_via_sandbox(
            job_id=job_id, title=title, slides=sections,
            user_id=user_id, chat_id=chat_id, content_md=content_md,
            source_doc_name=source_doc_name,prev_doc_name=prev_doc_name,
            question=question, llm_meta=llm_meta,
            artifact_id=_sbx_artifact, version=_sbx_version,
        )
        if _sandbox_ok is not None:
            return  # True → published; False → _fail already called
        logger.warning(
            f"[docgen] worker PPTX sandbox unavailable — falling back to native "
            f"skill path | job={job_id}"
        )
        sections = _enrich_with_images(sections)

    # ── Generate file ──────────────────────────────────────────
    logger.info(f"[docgen] worker STEP 4/6 — generating {fmt} file via platform skillset | job={job_id}")
    _publish_progress(job_id, 4, 6, "Generating File", f"Building {fmt.upper()} layout…")
    domain = llm_meta.get("domain") if llm_meta else None
    logger.info(f"[docgen] worker domain for palette selection: {domain!r} | job={job_id}")

    # ── PLAIN-TEXT CSV short-circuit ──────────────────────────────────────
    # When the structuring LLM classified a `.csv` request as informational
    # (csv_mode == "plain_text"), or a preservation/override path produced real
    # sections, do NOT hand it to the synthetic test-data code-writer. Write the
    # SAME content the live preview shows directly to the .csv (one cell per
    # line). This fixes: (a) informational CSV requests getting fake transaction
    # tables, and (b) the preview-vs-download divergence for "this chat as CSV".
    _csv_mode = (llm_meta.get("csv_mode") if llm_meta else "") or ""
    if fmt == "csv" and _csv_mode == "plain_text":
        logger.info(
            f"[docgen] worker CSV plain_text mode — writing content verbatim "
            f"(bypassing test-data code-writer) | job={job_id} "
            f"content_chars={len(content_md)}"
        )
        data = _render_plaintext_csv(content_md)
        ext, mime = "csv", MIME_TYPES.get("csv", "text/csv")
    else:
        # ── OLD tools.doc_generator.generate() DISABLED — use platform skillset ──
        # When the user uploaded a file (CSV/XLSX template) and asked to expand it,
        # hand the parsed content directly to the code-writer LLM so column names,
        # ordering and types survive the round-trip. Skipped automatically when
        # _is_preserve_intent (the structuring LLM already absorbs the content).
        _attach_for_code = (
            _parsed_attachment
            if (_parsed_attachment and not _is_preserve_intent)
            else ""
        )
        # Capture the code-writer's (dominant) LLM cost so it lands in the
        # budget deduction below — not just the cheap _llm_structure pass.
        _skill_cost: dict = {}
        skill_result = _skill_generate(
            job_id=job_id, fmt=fmt, question=question,
            title=title, sections=sections, domain=domain, theme="dark_executive",
            parsed_attachment=_attach_for_code,
            source_filename=_attachment_filename or "",
            cost_sink=_skill_cost,
            # B3: this is the single largest silent block on the /ask flow
            # (step 4/6, "Generating File" — measured ~98s in production with
            # no sub-progress at all). Publishes per-attempt/per-repair-round
            # detail under the same step/total already announced above.
            progress_step=(4, 6),
        )
        if skill_result is None:
            return  # _fail already called inside _skill_generate
        data, ext, mime = skill_result
        _merge_llm_cost(llm_meta, _skill_cost, job_id=job_id, phase="skill_generate")
    logger.info(f"[docgen] worker file generation DONE | job={job_id} ext={ext} size={len(data):,} bytes domain={domain!r}")

    # ── Write to temp dir ──────────────────────────────────────
    logger.info(f"[docgen] worker STEP 5/6 — writing file to disk | job={job_id}")
    _publish_progress(job_id, 5, 6, "Saving File", "Writing document to disk…")
    file_id  = str(_uuid_mod.uuid4())
    from tools.doc_generator import smart_filename
    _src_name = source_doc_name or prev_doc_name

    if prev_doc_name:
        # ── UPDATE/FOLLOW-UP: derive the name from the previous doc and append a
        # version suffix so each revision is distinct (-updated, -v2, -v3, …),
        # instead of overwriting with the same name.
        _base = _versioned_basename(prev_doc_name)
    else:
        # ── NEW DOC ────────────────────────────────────────────────
        # When the doc is generated FROM CHAT CONTENT (no upload), the user's
        # request ("generate this explanation into a word doc") is generic noise —
        # the meaningful signal is the LLM-generated `title` derived from the
        # response. Pass question="" so smart_filename uses the content title
        # (Priority 4) instead of the generic prompt topic (Priority 3).
        _from_chat = bool(chat_last_response or chat_context)
        _base = smart_filename(
            title=title,
            question="" if _from_chat else question,
            source_doc_name=_src_name,
            fmt_ext=ext,
        )
    filename = f"{_base}.{ext}"
    path     = os.path.join(_user_dir, f"{file_id}.{ext}")
    logger.info(f"[docgen] worker file path | job={job_id} file={file_id} path={path}")

    try:
        _atomic_write_bytes(path, data)
        logger.info(f"[docgen] worker file write DONE | job={job_id} filename={_safe_log(filename)}")
    except Exception as exc:
        logger.error(f"[docgen] worker file write FAILED | job={job_id} error={exc}", exc_info=True)
        _fail(job_id, f"File write error: {exc}")
        return

    # ── Postgres audit record ──────────────────────────────────
    logger.info(f"[docgen] worker STEP 6/6 — saving audit record | job={job_id}")
    _publish_progress(job_id, 6, 6, "Finalizing", "Recording audit trail…")
    # Version chaining: prefer an explicit client artifact_id; otherwise use the
    # one the doc plan resolved (a revise/convert follow-up targeting a prior
    # doc). When we version an existing artifact, bump version = prev + 1 so the
    # follow-up forms a proper chain (Canvas/Pages parity) instead of a new doc.
    _audit_artifact = (payload.get("artifact_id") or _plan.target_artifact_id or None)
    _audit_version  = (payload.get("version") or None)
    if _audit_artifact and not _audit_version:
        # Prefer the version the plan already loaded (no extra DB read). Only
        # query as a fallback when the plan didn't resolve it (e.g. an
        # explicit client-supplied artifact_id the plan never saw).
        if _plan.target_artifact_id == _audit_artifact and _plan.target_version:
            _audit_version = int(_plan.target_version) + 1
        else:
            try:
                from services.doc_context import load_latest_source as _lls
                _prev = _lls(_audit_artifact, user_id)
                _audit_version = (int(_prev.version) + 1) if _prev else None
            except Exception:  # noqa: BLE001
                _audit_version = None
    _save_audit(
        file_id=file_id, job_id=job_id, user_id=user_id, chat_id=chat_id,
        fmt=ext, title=title, filename=filename, file_path=path,
        content_md=content_md,
        artifact_id=_audit_artifact,
        version=_audit_version,
    )
    logger.info(f"[docgen] worker audit record saved | job={job_id}")

    # ── Save MD session for @edit_doc follow-up support ────────
    # Persists title, sections, content snapshot, and original format to
    # Redis (md:session:{chat_id}) so that generate_md_job(mode="edit")
    # can load this context and regenerate in the same format after editing.
    if chat_id:
        _save_md_session_for_chat(
            chat_id=chat_id,
            job_id=job_id,
            file_id=file_id,
            title=title,
            domain=domain,
            sections=sections,
            content_md=content_md,
            filename=filename,
            file_path=path,
            original_format=ext,
            question=question,
            llm_meta=llm_meta,
        )

    # ── Publish result ─────────────────────────────────────────
    logger.info(f"[docgen] worker publishing result to Redis | job={job_id}")
    # Computed once; reused for the result meta and the persisted ChatMessage.
    _q_job_latency = round(max(0.0, time.time() - _q_job_t0), 3)
    doc_result = {
        "status":   "done",
        "file_id":  file_id,
        "user_id":  str(user_id),   # owner — enforced by doc_job_status IDOR guard
        "artifact_id": _audit_artifact or file_id,
        "filename": filename,
        "format":   ext,
        "size":     len(data),
        "meta": {
            "model":   llm_meta.get("model"),
            "tokens":  llm_meta.get("tokens"),
            "in_tok":  llm_meta.get("in_tok"),
            "out_tok": llm_meta.get("out_tok"),
            "cost_usd": llm_meta.get("cost_usd"),
            "latency": _q_job_latency,
        },
    }
    _attach_summary_preview(
        doc_result, title=title, sections=sections,
        question=question, chat_id=chat_id, job_id=job_id,
    )
    _R.setex(f"doc:result:{job_id}", RESULT_TTL, json.dumps(doc_result))
    logger.info(f"[docgen] worker COMPLETE question-job {job_id} — {_safe_log(filename)} ({len(data):,} bytes)")

    if not chat_id:
        logger.info(f"[docgen] worker no chat_id — skipping chat message metadata update | job={job_id}")
        return

    logger.info(f"[docgen] worker updating chat message metadata | job={job_id} chat_id={chat_id}")
    try:
        from db.database import SessionLocal
        from db.models import ChatMessage

        _ast_id = payload.get("assistant_message_id")

        def _find_msg(db):
            # Prefer the explicit assistant_message_id stamped by /docs/generate
            # (stable, primary-key lookup). Fall back to a substring scan for
            # legacy callers (chat_worker flow) that don't propagate the id.
            if _ast_id:
                m = db.query(ChatMessage).filter(ChatMessage.id == _ast_id).first()
                if m:
                    return m
            return (
                db.query(ChatMessage)
                .filter(
                    ChatMessage.chat_id == chat_id,
                    ChatMessage.role == "assistant",
                    ChatMessage.content.contains(job_id),
                )
                .order_by(ChatMessage.created_at.desc())
                .first()
            )

        # The chat-history Kafka consumer is asynchronous; for very fast doc
        # jobs (~few seconds) the assistant row may not be committed yet. Retry
        # the lookup a few times before giving up on the metadata attach.
        msg = None
        with SessionLocal() as db:
            for _ in range(5):
                msg = _find_msg(db)
                if msg:
                    break
                time.sleep(0.5)
            if not msg:
                logger.warning(f"[docgen] worker chat message not found for job_id={job_id}")
                return

            msg.model_used = llm_meta.get("model")
            msg.tokens_used = llm_meta.get("tokens")
            msg.in_tok = llm_meta.get("in_tok")
            msg.out_tok = llm_meta.get("out_tok")
            msg.cost_usd = llm_meta.get("cost_usd")
            msg.latency = _q_job_latency
            logger.info(f"[docgen] worker chat message metadata updated | job={job_id}")
    except Exception as e:
        logger.warning(f"[docgen] worker chat message metadata update failed: {e}")

    try:
        from store.budget_store import increment_usage

        increment_usage(
            user_id,
            tokens=llm_meta.get("tokens",0),
            cost_usd=llm_meta.get("cost_usd",0.0)
        )
    except Exception as e:
        logger.warning(f"doc_worker: budget update failed: {e}")

    _sibling_formats = payload.get("sibling_formats") or []
    _sibling_job_ids = payload.get("sibling_job_ids") or []
    _MAX_SIBLINGS = 2
    if len(_sibling_formats) > _MAX_SIBLINGS:
        logger.warning(
            f"[docgen] worker clamping sibling_formats "
            f"{len(_sibling_formats)}→{_MAX_SIBLINGS} | job={job_id}"
        )
        _sibling_formats = _sibling_formats[:_MAX_SIBLINGS]
        _sibling_job_ids = _sibling_job_ids[:_MAX_SIBLINGS]
    if _sibling_formats and content_md.strip():
        try:
            from core.job_queue import enqueue_job as _enqueue_job, Q_DOC as _Q_DOC
            for _idx, _sfmt in enumerate(_sibling_formats):
                _sjid = (_sibling_job_ids[_idx]
                         if _idx < len(_sibling_job_ids) else str(_uuid_mod.uuid4()))
                _sib_payload = {
                    "job_id":      _sjid,
                    "format":      _sfmt,
                    "title":       title,
                    "content_md":  content_md,   # verbatim render — no re-author
                    "user_id":     user_id,
                    "chat_id":     chat_id,
                    "assistant_message_id": None,  # metadata attaches by content scan
                }
                _enqueue_job(
                    "workers.doc_worker.generate_doc_job",
                    _sib_payload, queue_name=_Q_DOC, timeout=1800, retry_count=0,
                )
                logger.info(
                    f"[docgen] worker sibling render enqueued | primary={job_id} "
                    f"sibling={_sjid} fmt={_sfmt}"
                )
        except Exception as _sib_err:  # noqa: BLE001
            logger.warning(f"[docgen] worker sibling render fan-out failed | job={job_id}: {_sib_err}")

    # ── Distinct-mode sequential fan-out ─────────────────────────────────────
    # pending_sibling_* keys are only set by the webchat /ask distinct-mode
    # path in gateway.py. All other callers never set these keys → no-op.
    # Each job enqueues only the NEXT pending sibling after it completes,
    # forming a chain that prevents queue starvation when workers are busy.
    _pending_fmts    = payload.get("pending_sibling_formats") or []
    _pending_job_ids = payload.get("pending_sibling_job_ids") or []
    _pending_intents = payload.get("pending_sibling_intents") or []

    if _pending_fmts:
        _pending_fmts    = _pending_fmts[:2]   # hard cap (max 3 docs total)
        _pending_job_ids = _pending_job_ids[:2]
        _pending_intents = _pending_intents[:2]
        try:
            from core.job_queue import enqueue_job as _enqueue_job_d, Q_DOC as _Q_DOC_d  # noqa: PLC0415
            _pjid    = _pending_job_ids[0] if _pending_job_ids else str(_uuid_mod.uuid4())
            _pfmt    = _pending_fmts[0]
            _pintent = _pending_intents[0] if _pending_intents else ""
            _pend_payload = {
                "job_id":     _pjid,
                "format":     _pfmt,
                "question":   payload.get("question", ""),
                "user_id":    user_id,
                "chat_id":    chat_id,
                "doc_intent": _pintent,
                # Same top-level classifier run as the primary job (docs[]
                # entries all share one _di.confidence — see gateway.py) so
                # the sibling's resolve_doc_plan can also trust the hint.
                "doc_confidence":     payload.get("doc_confidence"),
                "user_model_hint":    payload.get("user_model_hint", "auto"),
                "chat_context":       payload.get("chat_context", ""),
                "chat_last_response": payload.get("chat_last_response", ""),
                # Pass the remainder so the next job continues the chain
                "pending_sibling_formats":  _pending_fmts[1:],
                "pending_sibling_job_ids":  _pending_job_ids[1:],
                "pending_sibling_intents":  _pending_intents[1:],
            }
            _enqueue_job_d(
                "workers.doc_worker_agent.generate_doc_from_question",
                _pend_payload, queue_name=_Q_DOC_d, timeout=1800, retry_count=0,
            )
            logger.info(
                f"[docgen] distinct-mode sequential fan-out | "
                f"primary={job_id} next={_pjid} fmt={_pfmt} intent={_pintent!r}"
            )
        except Exception as _pend_err:  # noqa: BLE001
            logger.warning(
                f"[docgen] distinct-mode fan-out failed | job={job_id}: {_pend_err}"
            )


# ── File-conversion job ───────────────────────────────────────────────────────

def convert_doc_job(payload: dict) -> None:
    """
    RQ job: parse an uploaded file (docx/pdf/txt/…) and re-generate it in the
    requested target format using the branded doc_generator templates.

    payload keys:
      job_id          str  — Redis result key suffix
      storage_path    str  — path to the already-saved source file (from ChatAttachment)
      source_filename str  — original filename, e.g. "report.docx"
      source_ext      str  — file extension without dot, e.g. "docx"
      target_format   str  — "pdf" | "docx"
      user_id         str
      chat_id         str  (nullable)
    """
    job_id          = payload.get("job_id", "unknown")
    user_id         = payload.get("user_id", "unknown")
    chat_id         = payload.get("chat_id")

    from core.log_job_context import job_log_context
    with job_log_context(
        job_id=job_id, user_id=user_id, chat_id=chat_id or "",
        request_id=payload.get("request_id") or "",
        correlation_id=payload.get("correlation_id") or payload.get("request_id") or "",
        job_kind=payload.get("job_kind") or "convert",
        agent_id="doc_worker.convert_doc_job",
    ):
        return _convert_doc_job_impl(payload)


def _convert_doc_job_impl(payload: dict) -> None:
    job_id          = payload.get("job_id", "unknown")
    storage_path    = payload.get("storage_path", "")
    source_filename = payload.get("source_filename", "document")
    source_ext      = payload.get("source_ext", "").lower().strip(".")
    target_format   = payload.get("target_format", "pdf").lower().strip()
    user_id         = payload.get("user_id", "unknown")
    chat_id         = payload.get("chat_id")

    logger.info(
        f"[docgen] worker convert_doc_job START | job={job_id} "
        f"src={source_filename!r} ext={source_ext} → {target_format}"
    )

    _user_dir = user_doc_dir(user_id, chat_id)

    # ── 1. Load source file bytes via storage backend ─────────────────────────
    # storage_path is an opaque key returned by storage.save():
    #   "minio:<object_id>"  — stored in MinIO
    #   "local:<abs_path>"   — stored on local disk
    # Use storage.load() to retrieve bytes regardless of backend, then write
    # to a temp file so document_parser can open it by path.
    _publish_progress(job_id, 1, 4, "Parsing Document", f"Reading {source_filename}…")
    try:
        from core.storage import storage as _storage
        file_bytes = _storage.load(storage_path)
    except Exception as exc:
        logger.error(f"[docgen] worker convert_doc_job: storage.load failed: {exc}")
        _fail(job_id, f"Could not load source file: {exc}")
        return

    if not file_bytes:
        logger.error(f"[docgen] worker convert_doc_job: storage.load returned empty for {storage_path!r}")
        _fail(job_id, f"Source file not found or empty: {source_filename}")
        return

    # Write bytes to a temp file so document_parser can open it by path
    import tempfile as _tmpfile
    _tmp_suffix = f".{source_ext}" if source_ext else ""
    try:
        with _tmpfile.NamedTemporaryFile(delete=False, suffix=_tmp_suffix) as _tmp:
            _tmp.write(file_bytes)
            src_path = _tmp.name
    except Exception as exc:
        logger.error(f"[docgen] worker convert_doc_job: temp file write failed: {exc}")
        _fail(job_id, f"Could not stage source file: {exc}")
        return

    # ── 2. Parse source file → structured Markdown text ──────────────────────
    # document_parser returns clean Markdown preserving headings, tables, lists.
    # text_to_sections() then splits this into doc_generator section dicts.
    try:
        from core.document_parser import parse_file
        text = parse_file(src_path, source_ext, source_filename)
    except Exception as exc:
        logger.error(f"[docgen] worker convert_doc_job: parse failed: {exc}")
        _fail(job_id, f"Could not read source file: {exc}")
        return
    finally:
        # Always clean up the temp file
        try:
            os.unlink(src_path)
        except Exception:
            pass

    if not text or text.startswith("["):
        _fail(job_id, f"Could not extract text from {source_filename}")
        return

    # ── 3. Build title from filename ──────────────────────────────────────────
    import pathlib as _pl
    title = _pl.Path(source_filename).stem.replace("_", " ").replace("-", " ").title()

    # ── 4. Convert text → sections ────────────────────────────────────────────
    _publish_progress(job_id, 2, 4, "Structuring Content", "Splitting into sections…")
    try:
        from tools.doc_generator import text_to_sections
        sections = text_to_sections(text)
    except Exception as exc:
        logger.error(f"[docgen] worker convert_doc_job: text_to_sections failed: {exc}")
        _fail(job_id, f"Content structuring failed: {exc}")
        return

    # ── 5. Generate output file via platform skillset ────────────────────────
    # OLD: tools.doc_generator.generate() DISABLED
    _publish_progress(job_id, 3, 4, "Generating File", f"Building {target_format.upper()}…")
    # Hand the parsed source text to the code-writer as a template so the
    # converted file mirrors the source structure (especially important for
    # CSV → CSV row expansion).
    # Capture the code-writer's LLM cost so the conversion is billed for its
    # true (dominant) cost — this path previously deducted nothing at all.
    _skill_cost: dict = {}
    skill_result = _skill_generate(
        job_id=job_id, fmt=target_format, question=f"Convert {source_filename} to {target_format}",
        title=title, sections=sections,
        parsed_attachment=text or "",
        source_filename=source_filename or "",
        cost_sink=_skill_cost,
    )
    if skill_result is None:
        return  # _fail already called
    data, ext, mime = skill_result

    # ── 6. Write to temp dir ──────────────────────────────────────────────────
    _publish_progress(job_id, 4, 4, "Finalizing", "Saving converted file…")
    file_id  = str(_uuid_mod.uuid4())
    from tools.doc_generator import smart_filename
    _base    = smart_filename(
        title=title,
        source_doc_name=source_filename,
        fmt_ext=ext,
    )
    filename = f"{_base}.{ext}"
    path     = os.path.join(_user_dir, f"{file_id}.{ext}")

    try:
        _atomic_write_bytes(path, data)
    except Exception as exc:
        logger.error(f"[docgen] worker convert_doc_job: file write failed: {exc}")
        _fail(job_id, f"File write error: {exc}")
        return

    # ── 7. Audit record ───────────────────────────────────────────────────────
    _save_audit(
        file_id=file_id, job_id=job_id, user_id=user_id, chat_id=chat_id,
        fmt=ext, title=title, filename=filename, file_path=path,
        content_md=text[:5000],
    )

    # ── 8. Budget deduction (single point) ─────────────────────────────────────
    _c_tokens   = int(_skill_cost.get("tokens") or 0)
    _c_cost_usd = float(_skill_cost.get("cost_usd") or 0.0)
    if user_id and user_id not in ("unknown", "default", "") \
            and (_c_tokens > 0 or _c_cost_usd > 0.0):
        try:
            from store.budget_store import increment_usage
            increment_usage(user_id, tokens=_c_tokens, cost_usd=_c_cost_usd)
            logger.info(
                f"doc_worker: budget deducted (convert) | job={job_id} "
                f"user={user_id} tokens={_c_tokens} cost_usd={_c_cost_usd:.6f}"
            )
        except Exception as _bu_err:  # noqa: BLE001
            logger.warning(f"doc_worker: convert budget update failed | job={job_id}: {_bu_err}")

    # ── 9. Publish result ─────────────────────────────────────────────────────
    result = {
        "status":   "done",
        "file_id":  file_id,
        "user_id":  str(user_id),   # owner — enforced by doc_job_status IDOR guard
        "artifact_id": file_id,
        "filename": filename,
        "format":   ext,
        "size":     len(data),
        "meta":     {
            "model":    _skill_cost.get("model"),
            "tokens":   _c_tokens,
            "in_tok":   int(_skill_cost.get("in_tok") or 0),
            "out_tok":  int(_skill_cost.get("out_tok") or 0),
            "cost_usd": _c_cost_usd,
        },
    }
    _attach_summary_preview(
        result, title=title, sections=sections,
        question=f"Convert {source_filename} to {target_format}",
        chat_id=chat_id, job_id=job_id,
    )
    _R.setex(f"doc:result:{job_id}", RESULT_TTL, json.dumps(result))
    logger.info(f"[docgen] worker convert_doc_job DONE | job={job_id} → {_safe_log(filename)} ({len(data):,} bytes)")


# ── Helpers ───────────────────────────────────────────────────

def _versioned_basename(prev_doc_name: str) -> str:
    """
    Given the filename of a previously generated doc, return a NEW base name
    (without extension) for the updated/follow-up revision, so each revision is
    distinct instead of overwriting with the same name.

    Versioning rule (first revision → "-updated", then numbered "-v2", "-v3", …):
      "upi-payments.docx"          → "upi-payments-updated"
      "upi-payments-updated.docx"  → "upi-payments-v2"
      "upi-payments-v2.pdf"        → "upi-payments-v3"
      "upi-payments-v9"            → "upi-payments-v10"
    """
    import os as _os
    import re as _re

    # Strip directory + extension to get the bare base.
    base = _os.path.splitext(_os.path.basename((prev_doc_name or "").strip()))[0]
    if not base:
        return "generated-document-updated"

    # Already numbered: "...-v2" → bump to "...-v3".
    m = _re.search(r"^(.*)-v(\d+)$", base, _re.IGNORECASE)
    if m:
        stem, num = m.group(1), int(m.group(2))
        return f"{stem}-v{num + 1}"

    # First revision marker present: "...-updated" → start numbering at "-v2".
    m = _re.search(r"^(.*)-updated$", base, _re.IGNORECASE)
    if m:
        return f"{m.group(1)}-v2"

    # Pristine name (no version marker yet): first revision → "-updated".
    return f"{base}-updated"


def _derive_title_from_question(question: str) -> str:
    """
    Derive a clean, professional document title from the raw user question
    when the LLM does not return one.  Strips common request verbs and
    format keywords so the result reads like a proper document title.

    Examples:
      "generate a pdf report on UPI payments in India"
        → "UPI Payments in India"
      "write a word document about AI trends 2025"
        → "AI Trends 2025"
      "create a report on IPL cricket analytics"
        → "IPL Cricket Analytics"
    """
    import re as _re
    text = (question or "").strip()

    # Strip leading request verbs + format keywords
    _STRIP_PREFIX = _re.compile(
        r"^(please\s+)?"
        r"(generate|create|make|write|export|produce|draft|build|prepare|give|get|"
        r"want|need|show|provide|send|share|download|fetch|output|"
        r"summari[sz]e|summari[sz]ation|tl;?dr|rewrite|reword|shorten|expand|"
        r"convert|transform|turn|extract|pull|merge|combine|consolidate|update|revise|edit)\s+"
        r"(me\s+)?(this\s+|that\s+|these\s+|those\s+|it\s+)?(a\s+|an\s+|the\s+)?"
        r"(pdf|docx?|word|excel|xlsx?|pptx?|powerpoint|presentation|slides?|"
        r"spreadsheet|markdown|text|txt|report|document|doc|file|summary|analysis)?\s*"
        r"(report|document|doc|file|on|about|for|regarding|covering|of|into|to|from)?\s*",
        _re.IGNORECASE,
    )
    cleaned = _STRIP_PREFIX.sub("", text).strip()

    # Capitalise each word (title case), cap at 80 chars
    if cleaned:
        title = " ".join(w.capitalize() for w in cleaned.split())
        return title[:80]

    # Last resort: first 80 chars of original question, title-cased
    return " ".join(w.capitalize() for w in text.split())[:80] or "Document"


def _sanitize_llm_title(llm_title: str, question: str) -> str:
    """
    Guard against LLM returning the raw question or a request-verb phrase as the title.
    If the returned title looks like a raw prompt (starts with a request verb, or is
    longer than 100 chars), fall back to _derive_title_from_question.

    Examples of bad LLM titles that get replaced:
      "Generate a PDF report on UPI payments in India"  → "UPI Payments in India"
      "Create a detailed Word document about AI trends" → "AI Trends"
      "Write me a report on IPL cricket analytics"      → "IPL Cricket Analytics"
    """
    import re as _re
    t = (llm_title or "").strip()
    if not t:
        return _derive_title_from_question(question)

    # Reject if title is suspiciously long (raw question leak)
    if len(t) > 100:
        return _derive_title_from_question(question)

    # Reject if title starts with a request verb
    _BAD_START = _re.compile(
        r"^(please\s+)?(generate|create|make|write|export|produce|draft|build|prepare|"
        r"give|get|want|need|show|provide|send|share|download|fetch|output|"
        r"summari[sz]e|summari[sz]ation|tl;?dr|rewrite|reword|shorten|expand|"
        r"convert|transform|turn|extract|pull|merge|combine|consolidate|update|revise|edit)\b",
        _re.IGNORECASE,
    )
    if _BAD_START.match(t):
        return _derive_title_from_question(question)

    return t


# Env knob: which local model names the document. Defaults to the local tier;
# forward-ready for kimi-k2.7 / glm-5.2 (set DOC_INTENT_MODEL=local:<model-id>).
_TITLE_MODEL_HINT = (os.getenv("DOC_INTENT_MODEL", "") or "local").strip() or "local"


def _title_from_content(question: str, sections: list | None = None,
                         content_md: str = "", heuristic: str | None = None) -> str:
    """
    AUTHORITATIVE title: name the document from its ACTUAL CONTENT using the fast
    in-house local model — the way Claude/GPT title a document. Regex heuristics
    (`_derive_title_from_question`) are only a fail-open fallback because they
    cannot handle typos ("summarizr") or vague prompts ("summarize this doc").

    `heuristic` may be passed in to avoid recomputing _derive_title_from_question.
    Never raises — any failure degrades to the heuristic title so a document is
    never blocked on titling.
    """
    _fallback = heuristic if heuristic is not None else _derive_title_from_question(question)
    # Build a compact content snapshot for the namer.
    snapshot = (content_md or "").strip()
    if not snapshot and sections:
        parts = []
        for s in sections[:6]:
            if not isinstance(s, dict):
                continue
            h = (s.get("heading") or s.get("subheading") or "").strip()
            c = (s.get("content") or "").strip()
            if h:
                parts.append(h)
            if c:
                parts.append(c[:400])
        snapshot = "\n".join(parts)
    snapshot = snapshot[:2500]

    if not snapshot:
        # Nothing authored yet — best effort from the request text.
        return _fallback

    try:
        from models.model_router import model_router
        prompt = (
            "You are naming a document. Return ONLY a concise, professional title "
            "(3-9 words, Title Case, no quotes, no trailing punctuation, no file "
            "extension, do NOT restate the user's instruction verbs like "
            "'summarize'/'generate'/'convert'). Base it on the document CONTENT.\n\n"
            f"User request (context only): {(question or '').strip()[:300]}\n\n"
            f"Document content:\n{snapshot}\n\nTITLE:"
        )
        raw = (model_router.generate(prompt, model_hint=_TITLE_MODEL_HINT,
                                     return_meta=False) or "").strip()
        # Take first line, strip quotes/fences/trailing punctuation.
        raw = raw.splitlines()[0].strip() if raw else ""
        raw = raw.strip('`"\'' ).strip()
        raw = re.sub(r"^(title|document title)\s*[:\-]\s*", "", raw, flags=re.IGNORECASE).strip()
        raw = raw.rstrip(".;:, ").strip()
        # Reuse the guard: rejects request-verb starts / raw-prompt leaks / empties.
        cleaned = _sanitize_llm_title(raw, question)
        return cleaned[:120] if cleaned else _fallback
    except Exception as _terr:  # noqa: BLE001 — titling must never break generation
        logger.warning(f"[docgen] worker content-title generation failed, using heuristic: {_terr}")
        return _fallback


def _wants_flat_dataset(question: str) -> bool:
    """
    True when the user EXPLICITLY asks for a flat/plain tabular DATA sheet
    (a raw dataset), as opposed to a styled analytical report.

    Used both to decide whether an EDIT/EXPAND of an existing xlsx report may
    be turned into the programmatic test-data path, AND (as of the xlsx probe
    perf fix — see _llm_structure's xlsx_allow_testdata docstring) as the
    on-demand trigger for the CSV test-data classifier on a fresh xlsx
    request. By default we keep the narrative report format; this only flips
    it when the user clearly wants raw data. Deliberately conservative —
    matches strong, unambiguous phrases, plus an explicit row/record count
    (a near-certain dataset signal that a narrative report request never uses).

    Signals: "tabular only", "flat file/table", "raw data", "just the data",
    "plain data/spreadsheet", "data dump", "no descriptions/summaries", "as a
    dataset", "sample/synthetic/mock/dummy/test data", "N rows/records/entries".
    """
    q = (question or "").lower()
    _phrases = (
        "tabular only", "tabular format only", "only tabular",
        "flat file", "flat table", "flat data",
        "raw data", "just the data", "only the data", "data only",
        "plain data", "plain spreadsheet", "plain table",
        "data dump", "as a dataset", "as a data set",
        "no descriptions", "no description", "no summaries", "no summary",
        "no explanations", "no explanation", "no narrative", "no prose",
        "sample data", "synthetic data", "mock data", "dummy data",
        "test data", "fake data", "generate realistic sample",
        "rows of data", "records of data", "dummy records", "sample rows",
        "sample records", "dataset of", "data set of",
    )
    if any(p in q for p in _phrases):
        return True
    # An explicit "N rows/records/entries/..." count is only ever meaningful
    # for a bulk dataset — a narrative plan/report request has no reason to
    # specify one. Reuse the same parser the test-data prompt builder uses so
    # the two stay in lockstep.
    return _extract_row_count(question) is not None


def _extract_slide_count(question: str) -> int | None:
    """
    Parse the user's question for an explicit slide / page / section count.

    Handles patterns like:
      "1 slide pdf"          → 1
      "3-slide presentation" → 3
      "5 slides"             → 5
      "10 page report"       → 10
      "a 4 slide deck"       → 4
      "generate 2 slides"    → 2
    Returns None when no count is found (caller uses its own default).
    """
    import re as _re
    # Match: <number> (optional hyphen) slide(s)/page(s)/section(s)
    # OR:    slide(s)/page(s) <number>
    patterns = [
        r"\b(\d+)\s*[-–]?\s*(?:slides?|pages?|sections?)\b",
        r"\b(?:slides?|pages?|sections?)\s*[-–]?\s*(\d+)\b",
    ]
    for pat in patterns:
        m = _re.search(pat, question, _re.IGNORECASE)
        if m:
            n = int(m.group(1))
            # Sanity-clamp: 1–50
            return max(1, min(n, 50))
    return None


# Upper bound for CSV data-file row counts. Honours large explicit requests
# (e.g. "10000 records") while preventing a runaway sandbox job from a typo
# like "50000000 rows" that would OOM / hang the executor.
_CSV_MAX_ROWS = 100_000


def _extract_row_count(question: str) -> int | None:
    """
    Parse the user's question for an explicit CSV row / record count.

    Unlike _extract_slide_count (which matches slides/pages/sections and caps
    at 50), this matches data-file vocabulary and allows a much higher ceiling
    so requests like "10000 records" are honoured exactly instead of being
    silently truncated to the old 2000 schema default.

    Handles patterns like:
      "10000 records"             → 10000
      "generate 5000 rows"        → 5000
      "a record count of 10000"   → 10000
      "csv with 2500 entries"     → 2500
      "1,00,000 rows"             → 100000   (commas stripped)
    Returns None when no count is found (caller uses its own default).
    """
    import re as _re
    _data_kw = r"(?:rows?|records?|entries|entry|lines?|datapoints?|samples?)"
    patterns = [
        # "<N> rows/records/..."  (allow grouping commas inside the number)
        rf"\b([\d][\d,]*)\s*[-–]?\s*{_data_kw}\b",
        # "rows/records/... (of|count of|:)? <N>"
        rf"\b{_data_kw}\b(?:\s+(?:count|of|=|:|is|to))*\s*([\d][\d,]*)\b",
    ]
    for pat in patterns:
        m = _re.search(pat, question, _re.IGNORECASE)
        if m:
            try:
                n = int(m.group(1).replace(",", ""))
            except (TypeError, ValueError):
                continue
            if n <= 0:
                continue
            # Sanity-clamp: 1 – _CSV_MAX_ROWS
            return max(1, min(n, _CSV_MAX_ROWS))
    return None


def _render_plaintext_csv(content_md: str) -> bytes:
    """
    Render informational content as a PLAIN-TEXT .csv file — one cell per line.

    Used for the `csv_mode == "plain_text"` path: the user asked an informational
    question ("tell me about Madhya Pradesh", "explain UPI") or asked to convert a
    chat reply to CSV. There is no tabular dataset to fabricate, so we write the
    exact same content the live preview shows — each source line becomes a single
    quoted CSV field/row. This guarantees the downloaded file == the preview and
    contains NO synthetic transaction/record data.

    `csv.writer` handles RFC-4180 quoting/escaping (embedded commas, quotes, and
    newlines), so the result opens cleanly as a single column in Excel / Sheets
    while remaining faithful, readable plain text.
    """
    import csv as _csv
    import io as _io

    buf = _io.StringIO()
    writer = _csv.writer(buf, quoting=_csv.QUOTE_MINIMAL)
    # Split on newlines so headings, paragraphs, and bullets each land on their
    # own row. Keep blank lines (as empty rows) to preserve visual spacing.
    for line in (content_md or "").split("\n"):
        writer.writerow([line])
    # Excel/Sheets detect UTF-8 reliably with a BOM; harmless for other readers.
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _build_pptx_prompt(question: str) -> str:
    return (
        "You are an elite presentation designer equivalent to Presentation.ai, "
        "specializing in executive storytelling decks for senior leadership.\n\n"

        "Create a HIGHLY VISUAL, DESIGN-LED PowerPoint presentation about:\n"
        f"{question}\n\n"

        "Respond with ONLY valid JSON. NO markdown, NO explanations, NO commentary.\n\n"

        "JSON SCHEMA (follow exactly — use single braces for JSON, fill in all placeholder values):\n"
        '{"title":"<concise 3-6 word presentation title>","slides":[\n'
        ' {"slide_type":"title","heading":"<title>","key_message":"<8-12 word bold insight>",'
        '"icon":"","image_prompt":"<cinematic professional photograph, 16:9, ultra-clean, no text>",'
        '"speaker_notes":"<2-3 concise presenter notes>"},\n'

        ' {"slide_type":"agenda","heading":"Agenda","bullets":["<topic 1>","<topic 2>","<topic 3>","<topic 4>"],'
        '"speaker_notes":""},\n'

        ' {"slide_type":"content","heading":"<4-6 word heading>",'
        '"key_message":"<one strong insight sentence>",'
        '"bullets":["<insight 1>","<insight 2>","<insight 3>"],'
        '"icon":"",'
        '"image_prompt":"<modern corporate photo, dramatic lighting, shallow depth, no text>",'
        '"speaker_notes":""},\n'

        ' {"slide_type":"stats","heading":"<short heading>",'
        '"key_message":"<what this data proves>",'
        '"stats":[{"value":"<9.7B>","label":"<metric label>"},{"value":"<42%>","label":"<metric label>"}],'
        '"speaker_notes":""},\n'

        ' {"slide_type":"quote","heading":"",'
        '"quote":"<impactful leadership quote, 15-30 words>",'
        '"attribution":"— Name, Title",'
        '"speaker_notes":""},\n'

        ' {"slide_type":"two_column","heading":"<comparison or contrast>",'
        '"two_col_left":{"title":"<left title>","bullets":["<point 1>","<point 2>","<point 3>"]},'
        '"two_col_right":{"title":"<right title>","bullets":["<point 1>","<point 2>","<point 3>"]},'
        '"speaker_notes":""},\n'

        ' {"slide_type":"closing","heading":"Next Steps",'
        '"key_message":"<forward-looking closing insight>",'
        '"bullets":["<action 1>","<action 2>","<action 3>"],'
        '"icon":"",\n'
        '"image_prompt":"<confident business success imagery, clean background, no text>",'
        '"speaker_notes":""}\n'
        ']}\n\n'

        "ABSOLUTE DESIGN RULES (DO NOT VIOLATE):\n"
        "1. Design-first, minimal text, executive visual storytelling.\n"
        "2. Slides should look generated by Presentation.ai, NOT Microsoft default PPT.\n"
        "3. Prefer strong visuals + short copy over explanations.\n"
        "4. Bullets: MAX 8 words, no punctuation.\n"
        "5. Headings: MAX 6 words.\n"
        "6. key_message must be INSIGHTFUL, not a summary.\n"
        "7. Always assume image-backed layouts for title, content, and closing slides.\n"
        "8. Professional, premium, cinematic aesthetic — no cartoons, no clipart.\n"
        "9. Write for a global professional audience — use universally applicable facts, data, and examples relevant to the topic.\n"
        "10. title field: 3-6 words, professional, NO request verbs (generate/create/write/make).\n"
        "11. Output RAW JSON ONLY. Any non-JSON text is a failure.\n"
    )


# ── Shared rich section schema used by both DOCX and PDF prompts ─────────────
#
# Each section supports:
#   heading      str   — section title (required)
#   subheading   str   — optional H3 sub-label
#   level        int   — 1 = H1 major section, 2 = H2 subsection
#   content      str   — body paragraphs (\\n\\n separated)
#   bullets      list  — • bullet points (complete sentences)
#   callout      dict  — {"label": "<descriptive label>", "text": "..."}
#                        label examples: "Key Highlight", "Market Inflection",
#                        "Strategic Vision", "Risk Alert", "Data Insight",
#                        "Critical Finding", "Growth Driver", "Action Required"
#   table        dict  — {"headers": [...], "rows": [[...], ...]}
#
# Top-level:
#   title        str   — document title
#   domain       str   — industry keyword for palette selection
#                        (payments | ai | healthcare | government | sports |
#                         heritage | executive | cybersecurity | esg | ...)
#
# Design standard:
#   - Cover page: clean minimal — primary top band, large white title, accent stripe,
#     subtitle + "Prepared by" + date below band, CONFIDENTIAL badge
#   - Header (3 zones): LEFT doc title | CENTER Confidential © YYYY | RIGHT Page N
#   - No Table of Contents — content starts immediately after cover
#   - Section numbering: H1 → "1.  Heading", H2 → "1.1  Heading"
#   - Callout labels: descriptive (not fixed KEY/STAT/INSIGHT)
#   - Bullets: • (not →)


def _build_chat_preservation_prompt(last_response: str, user_request: str, target_format: str) -> str:
    """
    Convert the LAST assistant chat reply into structured JSON for rendering,
    PRESERVING all original content verbatim — fenced code blocks (in a dedicated
    `code` field) and markdown tables (in a structured `table` field) — while ALSO
    honoring any additions the user explicitly asked for in their follow-up request
    (e.g. "convert this to a word doc AND add more info about X").

    Originals are never altered, summarized, or reordered; requested additions are
    appended as new sections clearly after the preserved content.

    Used when the user generates a doc from a chat answer (no file uploaded). The
    output is fed to _llm_structure via override_prompt so no summarization happens.
    """
    return (
        "You are a faithful content-preservation assistant. Convert the assistant "
        f"reply below into structured JSON for rendering as a {target_format.upper()} "
        "document. PRESERVE the existing content EXACTLY, then ADD only what the user's "
        "follow-up request explicitly asks for.\n\n"

        "CRITICAL RULES:\n"
        "1. PRESERVE ALL existing text verbatim. Do NOT summarize, shorten, rewrite, "
        "reorder, or drop anything from the assistant reply.\n"
        "2. Every fenced code block (```...```) in the source MUST be captured in a "
        "section's `code` field EXACTLY as written — preserve newlines, indentation, and "
        "all whitespace. NEVER put code into the `content` field and NEVER reflow it.\n"
        "3. Set `language` to the fence's language tag (e.g. 'python', 'go', 'java') or "
        "an empty string if none.\n"
        "4. Every markdown table (lines using '|' with a '---' separator row) MUST be "
        "extracted into the structured `table` field as "
        '{"headers": ["Col1","Col2"], "rows": [["a","b"],["c","d"]]} — do NOT leave the '
        "table as raw pipe-text inside `content`.\n"
        "5. Keep content in reading order: prose preceding a code block goes in that "
        "section's `content`, the code in its `code` field, and a table in its `table` "
        "field. If multiple code blocks/tables are separated by prose, emit multiple sections.\n"
        "6. ADDITIONS: If the user's follow-up request asks to add/update/include more "
        "information, append NEW section(s) AFTER all the preserved sections covering that "
        "additional information. Clearly scope additions to exactly what the user asked — "
        "do NOT modify or interleave them with the preserved original content. If the "
        "request asks only to convert/produce the doc (no additions), add nothing.\n"
        "7. Preserve inline formatting and bullet lists as they appear.\n\n"

        f"USER FOLLOW-UP REQUEST (preserve the reply below, then honor any explicit "
        f"additions requested here): {user_request}\n\n"
        "ASSISTANT REPLY TO PRESERVE:\n"
        "---\n"
        f"{last_response}\n"
        "---\n\n"

        "Respond with ONLY valid JSON — no markdown fences, no explanation.\n\n"
        "JSON SCHEMA:\n"
        '{\n'
        '  "title": "<short title derived from the reply topic — no verbs like generate/create>",\n'
        '  "domain": "<single keyword: payments|banking|fintech|technology|ai|default>",\n'
        '  "sections": [\n'
        '    {\n'
        '      "heading": "<short heading, or empty string if the source had none>",\n'
        '      "subheading": "",\n'
        '      "level": 2,\n'
        '      "content": "<verbatim prose text for this section, paragraphs split by \\n\\n; '
        'EMPTY string if this section is only a code block or table>",\n'
        '      "bullets": ["<verbatim bullet if the source used a list>"],\n'
        '      "code": "<verbatim code block exactly as written, or empty string if none>",\n'
        '      "language": "<code fence language tag, or empty string>",\n'
        '      "callout": null,\n'
        '      "table": null\n'
        '    }\n'
        '  ]\n'
        '}\n\n'
        "For a section that contains a markdown table, set the table field like this "
        '(a JSON object, NOT a string): "table": {"headers": ["Col1","Col2"], '
        '"rows": [["a","b"],["c","d"]]}\n\n'
        "REMINDER: Preserve originals verbatim (code → `code`, tables → structured `table`). "
        "Append requested additions as new trailing sections. No other changes."
    )


def _build_preservation_prompt(parsed_content: str, source_filename: str, target_format: str) -> str:
    """
    Build a prompt that instructs the LLM to faithfully convert already-parsed
    file content into doc_generator sections — WITHOUT inventing or adding anything.

    Used when the user uploads a file and asks to convert/reproduce it as PDF/DOCX.
    The parsed_content is the Markdown output from document_parser.py.
    """
    return (
        "You are a document conversion assistant. Your ONLY job is to convert the "
        "parsed content of an uploaded file into a structured JSON format for rendering "
        f"as a {target_format.upper()} document.\n\n"

        "CRITICAL RULES — you MUST follow these exactly:\n"
        "1. Use ONLY the content provided below. Do NOT add, invent, or infer any content.\n"
        "2. Do NOT add sections that are not in the source (no Executive Summary, no "
        "   Strategic Recommendations, no Conclusion unless they exist in the source).\n"
        "3. Preserve ALL headings, tables, bullet points, and text exactly as they appear.\n"
        "4. The title must come from the document itself, not from the filename.\n"
        "5. Every section heading must match a heading from the source document.\n"
        "6. Tables in the source MUST appear as tables in the output JSON.\n\n"

        f"SOURCE FILE: {source_filename}\n\n"
        "PARSED CONTENT TO CONVERT:\n"
        "---\n"
        f"{parsed_content}\n"
        "---\n\n"

        "Respond with ONLY valid JSON — no markdown fences, no explanation.\n\n"

        "JSON SCHEMA:\n"
        '{\n'
        '  "title": "<exact document title from the source content>",\n'
        '  "domain": "<single keyword: payments|banking|fintech|healthcare|government|legal|default>",\n'
        '  "sections": [\n'
        '    {\n'
        '      "heading": "<exact heading from source>",\n'
        '      "subheading": "<sub-heading if present, else empty string>",\n'
        '      "level": 1,\n'
        '      "content": "<exact paragraph text from source for this section>",\n'
        '      "bullets": ["<bullet from source if any>"],\n'
        '      "callout": null,\n'
        '      "table": null\n'
        '    }\n'
        '  ]\n'
        '}\n\n'

        "For sections that contain a table in the source, set the table field like this:\n"
        '  "table": {"headers": ["Col1","Col2"], "rows": [["val","val"],["val","val"]]}\n\n'

        "REMINDER: Only use content from the source. No hallucination. No added sections."
    )


def _build_freeform_prompt(question: str, fmt: str) -> str:
    """
    Skill-aligned free-form structuring prompt — replaces _build_docx_prompt and
    _build_pdf_prompt.

    Mirrors the prompt used by scenario/test_premium_design_docs.py
    (_FREE_FORM_SYSTEM_PROMPT), which produces the correct platform skill
    output (natural topic-driven headings, no forced Executive Summary,
    no forced "Closing Insight" callout label).

    The skill itself (_skill_generate → SKILL.md + AiNxt_BRAND.md) applies all
    visual styling. This prompt's ONLY job is to convert the user's request
    into a JSON section list without injecting opinionated structure.
    """
    requested_sections = _extract_slide_count(question)
    if requested_sections:
        section_rule = (
            f"CRITICAL: The user explicitly requested exactly {requested_sections} "
            f"section(s). Produce exactly {requested_sections} section(s) in the "
            f"'sections' array — no more, no less.\n\n"
        )
    else:
        section_rule = ""

    fmt_label = "Word document" if fmt in ("docx", "word", "doc") else "PDF report"

    return (
        "You are an expert document author and information architect.\n\n"
        f"Author a high-quality {fmt_label} for the user's request as a JSON "
        "payload of sections. Follow these principles:\n"
        "- If the user SPECIFIED the structure/headings/sections, follow that "
        "structure exactly — same sections, same order.\n"
        "- If the request is OPEN-ENDED, design a clear, logical structure with "
        "natural topic-driven headings and write substantive, specific content "
        "for each section (like a subject-matter expert would).\n"
        "- Do NOT inject a generic 'Executive Summary' or 'Conclusion' / 'Closing "
        "Insight' section unless it genuinely serves the document or the user "
        "asked for it.\n"
        "- Never fabricate specific facts, statistics or names you cannot support.\n\n"
        + section_rule +
        f"User request:\n{question}\n\n"
        "OUTPUT FORMAT\n"
        "Respond with ONLY valid JSON — no markdown fences, no explanation, no preamble.\n\n"
        "JSON SCHEMA (follow exactly):\n"
        '{\n'
        '  "title": "<concise 3-7 word document title derived from the user\'s request — '
        'NO request verbs like generate/create/write/make>",\n'
        '  "domain": "<single lowercase keyword: payments|banking|ai|technology|'
        'cybersecurity|legal|education|heritage|sports|esg|government|fintech|'
        'healthcare|executive|default>",\n'
        '  "sections": [\n'
        '    {\n'
        '      "heading": "<section heading as requested by the user>",\n'
        '      "subheading": "<sub-label if the user specified one, otherwise empty string>",\n'
        '      "level": 1,\n'
        '      "content": "<full section body text, paragraphs separated by \\n\\n, '
        'written exactly as the user requested>",\n'
        '      "bullets": ["<bullet point as requested>"],\n'
        '      "callout": {"label": "<short label>", "text": "<callout text if applicable, otherwise empty>"},\n'
        '      "table": null\n'
        '    }\n'
        '  ]\n'
        '}\n\n'
        "RULES\n"
        "1. Follow the user's requested structure exactly — produce every section "
        "they listed, in the same order. Use natural topic-driven headings drawn "
        "from the request itself.\n"
        "2. Write the content the user asked for — do not invent extra sections or "
        "change the scope. Do NOT prepend an Executive Summary or append a "
        "Conclusion / Closing Insight unless explicitly requested.\n"
        "3. If the user requests a table, populate 'table' as "
        '{"headers": [...], "rows": [[...], ...]}.\n'
        "4. Output RAW JSON ONLY. Any non-JSON text causes a parse failure.\n"
        "5. All text must be grammatically complete and correctly spelled.\n"
    )


def _build_docx_prompt(question: str) -> str:
    """
    DEPRECATED — DO NOT CALL.

    The hardcoded "Executive Summary" first section and "Closing Insight"
    callout in this schema produced the off-brand legacy DOCX template.
    All DOCX generation now routes through _build_freeform_prompt instead,
    which mirrors the local test runner (test_premium_design_docs.py) and
    produces the platform skill output the user expects.

    Kept temporarily for git-history reference. Will be removed once we
    confirm no external callers depend on it.
    """
    requested_sections = _extract_slide_count(question)
    if requested_sections:
        section_rule = (
            f"CRITICAL: The user explicitly requested exactly {requested_sections} section(s). "
            f"You MUST produce exactly {requested_sections} section(s) in the 'sections' array — no more, no less.\n\n"
        )
    else:
        section_rule = ""
    return (
        "You are a professional document designer and senior analyst. "
        "You produce McKinsey/BCG-quality branded Word documents.\n\n"

        + section_rule +

        "Create a rich, multi-section Word document report on:\n"
        f"{question}\n\n"

        "Respond with ONLY valid JSON — no markdown fences, no explanation.\n\n"

        "JSON SCHEMA (follow exactly — use standard single-brace JSON syntax):\n"
        '{\n'
        '  "title": "<concise 3-7 word professional title — NO request verbs like generate/create/write/make>",\n'
        '  "domain": "<single industry keyword: payments|ai|healthcare|government|banking|'
        'fintech|cybersecurity|legal|education|heritage|sports|esg|retail|luxury|startup|'
        'media|travel|food|hr|executive|infrastructure|default>",\n'
        '  "sections": [\n'
        '    {\n'
        '      "heading": "Executive Summary",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<2-3 substantive paragraphs: context, key findings, significance>",\n'
        '      "bullets": [],\n'
        '      "callout": {"label": "Key Highlight", "text": "<single most important insight from this section>"},\n'
        '      "table": null\n'
        '    },\n'
        '    {\n'
        '      "heading": "<Core Section Heading>",\n'
        '      "subheading": "<optional H3 sub-label>",\n'
        '      "level": 1,\n'
        '      "content": "<3-4 analytical paragraphs with real data, percentages, named entities>",\n'
        '      "bullets": ["<complete insight sentence>","<complete insight sentence>","<complete insight sentence>"],\n'
        '      "callout": {"label": "Market Inflection", "text": "<key statistic or data point from this section>"},\n'
        '      "table": {"headers": ["<Col 1>","<Col 2>","<Col 3>"], "rows": [["<v>","<v>","<v>"],["<v>","<v>","<v>"],["<v>","<v>","<v>"]]}\n'
        '    },\n'
        '    {\n'
        '      "heading": "<Analysis Section>",\n'
        '      "subheading": "",\n'
        '      "level": 2,\n'
        '      "content": "<2-3 paragraphs>",\n'
        '      "bullets": ["<insight>","<insight>","<insight>"],\n'
        '      "callout": {"label": "Critical Finding", "text": "<analytical insight>"},\n'
        '      "table": null\n'
        '    },\n'
        '    {\n'
        '      "heading": "Challenges & Risk Factors",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<2-3 paragraphs identifying key risks>",\n'
        '      "bullets": ["<risk 1>","<risk 2>","<risk 3>"],\n'
        '      "callout": {"label": "Risk Alert", "text": "<primary risk or challenge>"},\n'
        '      "table": null\n'
        '    },\n'
        '    {\n'
        '      "heading": "Strategic Recommendations",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<2-3 paragraphs with actionable guidance>",\n'
        '      "bullets": ["<recommendation 1>","<recommendation 2>","<recommendation 3>","<recommendation 4>"],\n'
        '      "callout": {"label": "Strategic Vision", "text": "<primary strategic goal>"},\n'
        '      "table": null\n'
        '    },\n'
        '    {\n'
        '      "heading": "Conclusion",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<strong closing paragraph with forward-looking statement>",\n'
        '      "bullets": [],\n'
        '      "callout": {"label": "Closing Insight", "text": "<closing insight or call to action>"},\n'
        '      "table": null\n'
        '    }\n'
        '  ]\n'
        '}\n\n'

        + (
            f"DOCUMENT QUALITY RULES:\n"
            f"1. title: 3-7 words, professional noun phrase, NO verbs like generate/create/write/make/report/document.\n"
            f"   GOOD: 'UPI Payments in India 2025'  BAD: 'Generate a Report on UPI'\n"
            f"2. {'Produce EXACTLY ' + str(requested_sections) + ' section(s) — this is a hard requirement.' if requested_sections else 'Minimum 5-7 sections — no thin or placeholder sections.'}\n"
            "3. Each section's 'content' must be 2-4 full analytical paragraphs with real data.\n"
            "4. All data must be realistic and specific: real percentages, years, named entities.\n"
            "5. Bullets must be complete, insightful sentences — not fragments.\n"
            "6. Every section must have a callout box with a DESCRIPTIVE label (e.g. 'Key Highlight',\n"
            "   'Market Inflection', 'Critical Finding', 'Risk Alert', 'Strategic Vision',\n"
            "   'Growth Driver', 'Data Insight', 'Action Required', 'Closing Insight').\n"
            "   Do NOT use generic labels like KEY, STAT, INSIGHT, GOAL, NOTE.\n"
            "7. At least 2 sections must include a data table with 3+ columns and 3+ rows.\n"
            "8. domain must be a single lowercase keyword from the allowed list.\n"
            "9. Output tone: Executive briefing — authoritative, data-driven, factual.\n"
            "10. Output RAW JSON ONLY. Any non-JSON text is a failure.\n"
        )
    )


def _build_pdf_prompt(question: str) -> str:
    """
    DEPRECATED — DO NOT CALL.

    The hardcoded "Executive Summary" first section and "Closing Insight"
    callout in this schema produced the off-brand legacy PDF template.
    All PDF generation now routes through _build_freeform_prompt instead.

    Kept temporarily for git-history reference. Will be removed once we
    confirm no external callers depend on it.
    """
    requested_sections = _extract_slide_count(question)
    if requested_sections:
        section_rule = (
            f"CRITICAL: The user explicitly requested exactly {requested_sections} section(s). "
            f"You MUST produce exactly {requested_sections} section(s) in the 'sections' array — no more, no less.\n\n"
        )
    else:
        section_rule = ""
    return (
        "You are a professional document designer and senior analyst. "
        "You produce McKinsey/BCG-quality branded PDF executive reports.\n\n"

        + section_rule +

        "Create a rich, multi-section PDF report on:\n"
        f"{question}\n\n"

        "Respond with ONLY valid JSON — no markdown fences, no explanation.\n\n"

        "JSON SCHEMA (follow exactly — use standard single-brace JSON syntax):\n"
        '{\n'
        '  "title": "<concise 3-7 word formal report title — NO request verbs like generate/create/write/make>",\n'
        '  "domain": "<single industry keyword: payments|ai|healthcare|government|banking|'
        'fintech|cybersecurity|legal|education|heritage|sports|esg|retail|luxury|startup|'
        'media|travel|food|hr|executive|infrastructure|default>",\n'
        '  "sections": [\n'
        '    {\n'
        '      "heading": "Executive Summary",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<3-4 substantive paragraphs: context, key findings, significance, scope>",\n'
        '      "bullets": [],\n'
        '      "callout": {"label": "Key Highlight", "text": "<single most important finding>"},\n'
        '      "table": null\n'
        '    },\n'
        '    {\n'
        '      "heading": "Background & Context",\n'
        '      "subheading": "<optional sub-label>",\n'
        '      "level": 1,\n'
        '      "content": "<2-3 paragraphs establishing problem space, history, relevance>",\n'
        '      "bullets": ["<context point 1>","<context point 2>","<context point 3>"],\n'
        '      "callout": {"label": "Market Scale", "text": "<key statistic establishing scale or importance>"},\n'
        '      "table": {"headers": ["<Col 1>","<Col 2>","<Col 3>"], "rows": [["<v>","<v>","<v>"],["<v>","<v>","<v>"],["<v>","<v>","<v>"]]}\n'
        '    },\n'
        '    {\n'
        '      "heading": "<Core Analysis Section>",\n'
        '      "subheading": "",\n'
        '      "level": 2,\n'
        '      "content": "<3-4 paragraphs of deep analysis with data points and insights>",\n'
        '      "bullets": ["<finding 1>","<finding 2>","<finding 3>","<finding 4>"],\n'
        '      "callout": {"label": "Data Insight", "text": "<analytical insight from this section>"},\n'
        '      "table": {"headers": ["<Col 1>","<Col 2>","<Col 3>","<Col 4>"], "rows": [["<v>","<v>","<v>","<v>"],["<v>","<v>","<v>","<v>"],["<v>","<v>","<v>","<v>"]]}\n'
        '    },\n'
        '    {\n'
        '      "heading": "<Secondary Analysis Section>",\n'
        '      "subheading": "",\n'
        '      "level": 2,\n'
        '      "content": "<2-3 paragraphs>",\n'
        '      "bullets": ["<point 1>","<point 2>","<point 3>"],\n'
        '      "callout": {"label": "Growth Driver", "text": "<notable observation>"},\n'
        '      "table": null\n'
        '    },\n'
        '    {\n'
        '      "heading": "Challenges & Risk Factors",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<2-3 paragraphs identifying key risks and challenges>",\n'
        '      "bullets": ["<risk 1>","<risk 2>","<risk 3>"],\n'
        '      "callout": {"label": "Risk Alert", "text": "<primary risk or threat>"},\n'
        '      "table": null\n'
        '    },\n'
        '    {\n'
        '      "heading": "Strategic Recommendations",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<2-3 paragraphs with actionable strategic guidance>",\n'
        '      "bullets": ["<recommendation 1>","<recommendation 2>","<recommendation 3>","<recommendation 4>"],\n'
        '      "callout": {"label": "Strategic Vision", "text": "<primary strategic objective>"},\n'
        '      "table": null\n'
        '    },\n'
        '    {\n'
        '      "heading": "Conclusion",\n'
        '      "subheading": "",\n'
        '      "level": 1,\n'
        '      "content": "<strong closing paragraph with forward-looking statement and call to action>",\n'
        '      "bullets": [],\n'
        '      "callout": {"label": "Closing Insight", "text": "<closing insight or call to action>"},\n'
        '      "table": null\n'
        '    }\n'
        '  ]\n'
        '}\n\n'

        + (
            f"REPORT QUALITY RULES:\n"
            f"1. title: 3-7 words, professional noun phrase, NO verbs like generate/create/write/make/report/document.\n"
            f"   GOOD: 'AI Trends in Healthcare 2025'  BAD: 'Generate a Report on AI'\n"
            f"2. {'Produce EXACTLY ' + str(requested_sections) + ' section(s) — this is a hard requirement.' if requested_sections else 'Minimum 6-8 sections — this is a formal analytical report, not a summary.'}\n"
            "3. Each section's 'content' must be 2-4 full analytical paragraphs with real data.\n"
            "4. All data must be realistic and specific: real percentages, years, named entities.\n"
            "5. Bullets must be complete, insightful sentences — not fragments.\n"
            "6. Every section must have a callout box with a DESCRIPTIVE label (e.g. 'Key Highlight',\n"
            "   'Market Scale', 'Data Insight', 'Growth Driver', 'Risk Alert', 'Strategic Vision',\n"
            "   'Market Inflection', 'Critical Finding', 'Action Required', 'Closing Insight').\n"
            "   Do NOT use generic labels like KEY, STAT, INSIGHT, GOAL, NOTE.\n"
            "7. At least 2 sections must include a data table with 3+ columns and 3+ rows.\n"
            "8. domain must be a single lowercase keyword from the allowed list.\n"
            "9. Output tone: Executive briefing — authoritative, data-driven, factual. No filler.\n"
            "10. Output RAW JSON ONLY. Any non-JSON text is a failure.\n"
        )
    )


# ── Continuation system prompt for two-pass large-document generation ─────────
_CONTINUATION_SYSTEM_PROMPT = (
    "You are a document structuring assistant continuing a multi-part document generation.\n\n"
    "You will be given:\n"
    "  - The document title and domain (already decided).\n"
    "  - A list of section headings that have ALREADY been written.\n"
    "  - The remaining section headings that still need to be written.\n\n"
    "Your ONLY job is to produce the JSON for the REMAINING sections.\n"
    "Follow the user's requested content, structure, and tone exactly.\n\n"
    "OUTPUT FORMAT\n"
    "Respond with ONLY valid JSON — no markdown fences, no explanation, no preamble.\n\n"
    'JSON SCHEMA (follow exactly):\n'
    '{\n'
    '  "sections": [\n'
    '    {\n'
    '      "heading": "<section heading>",\n'
    '      "subheading": "<sub-label if applicable, otherwise empty string>",\n'
    '      "level": 1,\n'
    '      "content": "<full section body text, paragraphs separated by \\n\\n>",\n'
    '      "bullets": ["<bullet point>"],\n'
    '      "callout": {"label": "<short label>", "text": "<callout text if applicable, otherwise empty>"},\n'
    '      "table": null\n'
    '    }\n'
    '  ]\n'
    '}\n\n'
    "RULES\n"
    "1. Produce ONLY the remaining sections listed — do not repeat already-written sections.\n"
    "2. If a section requests a table, populate \"table\" as {\"headers\": [...], \"rows\": [[...], ...]}.\n"
    "3. Output RAW JSON ONLY. Any non-JSON text causes a parse failure.\n"
    "4. All text must be grammatically complete and correctly spelled.\n"
)

# Max sections per LLM pass — keeps each call well within API gateway token limits.
# Bumped from 4 → 6 to reduce total passes (20 sections = 4 passes instead of 5)
# which, combined with parallel execution of passes 2-N, cuts total wall-clock
# time roughly in half.
_MAX_SECTIONS_PER_PASS = 6

# Threshold: if the user's question contains more than this many numbered items (e.g. "1)", "2)"),
# use multi-pass chunked generation to avoid API gateway timeouts and token limits.
#
# IMPORTANT: This threshold must be HIGH enough that rich instruction prompts
# (e.g. "1) Executive summary ... 12) Glossary" where each item is a detailed
# instruction, not a bare heading) do NOT trigger multi-pass.  Multi-pass is
# designed for simple numbered-heading lists, not for complex training-document
# prompts.  A prompt with 12 detailed section instructions sent through 4 LLM
# passes takes ~12 minutes and hits the work-horse timeout.
#
# The _build_docx_prompt / _build_pdf_prompt single-pass builders already
# produce McKinsey-quality output for rich prompts — use them instead.
#
# Rule: only trigger multi-pass when the user explicitly lists > 15 bare
# section headings (e.g. a 20-section outline).  Rich instruction prompts
# with ≤ 15 numbered items go through the single-pass builder.
_SPLIT_THRESHOLD = 15


def _parse_llm_json(raw: str, job_id: str = "") -> dict:
    """
    Strip markdown fences and parse JSON robustly.

    When the LLM response is truncated mid-JSON (token limit hit), attempts to
    recover any complete sections that were already serialised before the cut-off.
    This prevents a hard failure when only the last section is incomplete.
    """
    original = raw
    raw = raw.strip()
    if not raw:
        logger.error(
            f"[docgen] worker _parse_llm_json received empty response — "
            f"LLM returned nothing (timeout, context overflow, or model error) | job={job_id}"
        )
        raise ValueError("LLM returned an empty response — possible context overflow or timeout")
    raw = re.sub(r"^```[a-zA-Z]*\s*", "", raw)
    raw = re.sub(r"\s*```\s*$", "", raw.strip())
    raw = raw.strip()
    if not raw.startswith("{"):
        m = re.search(r"\{", raw)
        if m:
            raw = raw[m.start():]
    if raw and not raw.endswith("}"):
        last_brace = raw.rfind("}")
        if last_brace != -1:
            raw = raw[:last_brace + 1]
    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        logger.warning(f"[docgen] worker JSON parse failed — attempting partial recovery | "
                       f"job={job_id} error={exc} raw_preview={original[:300]!r}")
        # ── Partial recovery: extract complete section objects from truncated JSON ──
        # Find all complete {...} objects inside the "sections" array.
        # A complete section ends with a closing brace followed by optional whitespace
        # and then either a comma+newline or the end of the sections array.
        recovered_sections = []
        # Match complete JSON objects in the sections array
        for m in re.finditer(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)?\}', raw):
            try:
                obj = json.loads(m.group())
                # Only keep objects that look like sections (have a "heading" key)
                if "heading" in obj:
                    recovered_sections.append(obj)
            except (json.JSONDecodeError, ValueError):
                continue

        if recovered_sections:
            logger.info(f"[docgen] worker partial recovery succeeded: {len(recovered_sections)} sections | job={job_id}")
            # Try to extract title and domain from the raw text
            title_m  = re.search(r'"title"\s*:\s*"([^"]*)"', raw)
            domain_m = re.search(r'"domain"\s*:\s*"([^"]*)"', raw)
            return {
                "title":    title_m.group(1)  if title_m  else "",
                "domain":   domain_m.group(1) if domain_m else "default",
                "sections": recovered_sections,
            }

        logger.error(f"[docgen] worker JSON parse failed, no recovery possible | "
                     f"job={job_id} error={exc} raw_preview={original[:300]!r}")
        raise


def _build_csv_prompt(question: str) -> str:
    """
    Build the LLM prompt for FLAT CSV generation (data files / test data).

    Unlike the xlsx builder, this returns a column-spec + row_count JSON, not
    a "sections with tables" structure. The downstream code-writing LLM then
    materialises N rows using `random` + curated lists (no Faker) and writes
    them via `pandas.DataFrame(...).to_csv(OUTPUT_PATH, index=False)`.

    Handles three common intents:
      1. Test/synthetic data ("generate a csv of 1000 fake customers")
      2. Data manipulation ("csv of UPI transactions for testing")
      3. Schema-following expansion (an uploaded template is handled by
         _skill_generate's parsed_attachment injection; this prompt just
         instructs the LLM not to invent columns when one is present).
    """
    requested_rows = _extract_row_count(question)  # matches "<N> rows/records/entries"
    if requested_rows:
        row_rule = (
            f"CRITICAL: The user explicitly requested exactly {requested_rows} row(s). "
            f"If csv_mode is \"test_data\", set row_count to exactly {requested_rows} "
            f"(do NOT cap or round it down).\n\n"
        )
    else:
        row_rule = ""

    return (
        "You are a CSV document assistant. The user asked for a `.csv` file. A CSV "
        "file can serve TWO very different purposes, and you must decide which one "
        "this request is BEFORE producing the schema:\n\n"

        "  A) TEST DATA  — the user wants a TABULAR dataset of generated/synthetic "
        "rows for testing, demos, or analysis. Signals: 'test data', 'sample data', "
        "'synthetic', 'dummy', 'fake', 'mock', 'generate N records/rows', 'dataset "
        "of transactions/customers/orders', column-like requests. → csv_mode = "
        "\"test_data\".\n"
        "  B) PLAIN TEXT — the user is asking an INFORMATIONAL / explanatory question "
        "(about a place, a concept, how something works, a summary, an overview) and "
        "merely wants the answer saved with a .csv extension. Signals: 'tell me "
        "about X', 'explain X', 'what is X', 'overview of X', 'about UPI', 'about "
        "Madhya Pradesh' — i.e. there is NO tabular dataset to fabricate. "
        "→ csv_mode = \"plain_text\".\n\n"

        "If you are not clearly being asked to fabricate tabular/test data, choose "
        "\"plain_text\". NEVER invent fake transaction/customer/record tables for an "
        "informational question.\n\n"

        + row_rule +

        "The user's request:\n"
        f"{question}\n\n"

        "Respond with ONLY valid JSON — no markdown fences, no explanation.\n\n"

        "JSON SCHEMA (follow exactly):\n"
        "{\n"
        '  "csv_mode": "<test_data | plain_text>",\n'
        '  "title": "<concise 3-7 word noun phrase — NO request verbs>",\n'
        '  "domain": "<single keyword: payments|banking|fintech|users|orders|'
        'transactions|healthcare|retail|hr|default>",\n'
        '  "row_count": <integer between 50 and 100000 — ONLY for test_data; '
        "use 0 for plain_text>,\n"
        '  "columns": [\n'
        '    {"name":"<snake_case column name>", '
        '"type":"<int|float|string|email|date|datetime|enum|currency|phone|uuid>", '
        '"example":"<one realistic value>", '
        '"notes":"<optional generation hint, e.g. enum values or range>"}\n'
        "  ],   // test_data ONLY — use [] for plain_text\n"
        '  "sections": [\n'
        '    {"heading":"<short heading or empty>", '
        '"content":"<verbatim informational prose, paragraphs split by \\n\\n>", '
        '"bullets":["<bullet>", "..."], "level":2}\n'
        "  ]    // plain_text ONLY — use [] for test_data\n"
        "}\n\n"

        "RULES FOR test_data (when csv_mode == \"test_data\"):\n"
        "1. title: 3-7 words, professional noun phrase. NO verbs like generate/create.\n"
        "2. row_count: if the user gave an explicit number, mirror it EXACTLY (it may "
        "be large, e.g. 10000 — honour it; do not clamp to 2000). Otherwise pick "
        "between 100 and 2000 for test/synthetic data; 50–200 for demo/preview.\n"
        "3. columns: 4–12 columns. First column should be a stable identifier "
        "(id, uuid, or sequence).\n"
        "4. Use snake_case column names (customer_id, full_name, signup_date).\n"
        "5. Pick column types matching the domain. e.g. payments → "
        "[txn_id, timestamp, amount, currency, status, payer_vpa, payee_vpa]; "
        "users → [user_id, full_name, email, phone, city, signup_date, plan].\n"
        "6. Provide one realistic `example` value per column — the downstream "
        "code-writer uses this to infer format.\n"
        "7. `sections` MUST be an empty array for test_data.\n\n"

        "RULES FOR plain_text (when csv_mode == \"plain_text\"):\n"
        "1. Answer the user's actual question fully and accurately in `sections`.\n"
        "2. Put the informational content in `sections` (heading + content + "
        "bullets). `columns` MUST be [] and `row_count` MUST be 0.\n"
        "3. Do NOT fabricate tabular rows, IDs, transactions, or records.\n\n"

        "Output RAW JSON ONLY. Any non-JSON text is a failure.\n"
    )


def _build_xlsx_prompt(question: str) -> str:
    """
    Build the LLM prompt for Excel spreadsheet generation.

    Instructs the LLM to return JSON with sections that include structured
    table dicts (headers + rows) so generate_xlsx() can produce rich,
    multi-sheet workbooks with proper data tables, number formatting, and charts.
    """
    requested_sections = _extract_slide_count(question)
    if requested_sections:
        section_rule = (
            f"CRITICAL: The user explicitly requested exactly {requested_sections} section(s). "
            f"You MUST produce exactly {requested_sections} section(s) — no more, no less.\n\n"
        )
    else:
        section_rule = ""

    return (
        "You are a professional data analyst and Excel report designer. "
        "You produce McKinsey/BCG-quality Excel workbooks with rich structured data.\n\n"

        + section_rule +

        "Create a comprehensive Excel spreadsheet report on:\n"
        f"{question}\n\n"

        "Respond with ONLY valid JSON — no markdown fences, no explanation.\n\n"

        "JSON SCHEMA (follow exactly):\n"
        "{\n"
        '  "title": "<concise 3-7 word professional title — NO request verbs>",\n'
        '  "domain": "<single keyword: payments|banking|ai|fintech|healthcare|government|'
        'education|retail|hr|executive|default>",\n'
        '  "sections": [\n'
        "    {\n"
        '      "heading": "<Section Name — will become a sheet tab>",\n'
        '      "level": 1,\n'
        '      "content": "<1-2 paragraph executive summary of this section>",\n'
        '      "bullets": ["<key insight 1>", "<key insight 2>", "<key insight 3>"],\n'
        '      "table": {\n'
        '        "headers": ["<Label Column>", "<Metric 1>", "<Metric 2>", "<Metric 3>"],\n'
        '        "rows": [\n'
        '          ["<Row Label 1>", "<numeric value>", "<numeric value>", "<numeric value>"],\n'
        '          ["<Row Label 2>", "<numeric value>", "<numeric value>", "<numeric value>"],\n'
        '          ["<Row Label 3>", "<numeric value>", "<numeric value>", "<numeric value>"],\n'
        '          ["<Row Label 4>", "<numeric value>", "<numeric value>", "<numeric value>"],\n'
        '          ["<Row Label 5>", "<numeric value>", "<numeric value>", "<numeric value>"],\n'
        '          ["<Row Label 6>", "<numeric value>", "<numeric value>", "<numeric value>"]\n'
        "        ]\n"
        "      }\n"
        "    }\n"
        "  ]\n"
        "}\n\n"

        "EXCEL DATA QUALITY RULES (follow strictly):\n"
        "1. title: 3-7 words, professional noun phrase, NO verbs like generate/create/write/make.\n"
        "   GOOD: 'UPI Payments Analysis 2025'  BAD: 'Generate a Report on UPI'\n"
        "2. Every section MUST have a 'table' dict with 'headers' and 'rows'.\n"
        "   Tables without data are useless — every row must have real numeric values.\n"
        "3. Numeric values: use realistic numbers with commas for thousands (e.g. '1,245.6').\n"
        "   Percentages: include % symbol (e.g. '23.4%'). Currency: include ₹ (e.g. '₹8,450').\n"
        "4. Each table must have at least 6 data rows (for chart generation).\n"
        "5. First column of each table is the row label (text). Remaining columns are numeric.\n"
        "6. Column headers must be concise (2-4 words) — they will wrap in Excel cells.\n"
        f"7. {'Produce EXACTLY ' + str(requested_sections) + ' section(s).' if requested_sections else 'Produce 3-5 sections, each with a distinct data table.'}\n"
        "8. Use time-series labels where appropriate (e.g. 'Apr-24', 'Q1 2025', 'Jan', 'FY24').\n"
        "9. domain: single lowercase keyword from the allowed list.\n"
        "10. Output RAW JSON ONLY. Any non-JSON text is a failure.\n"
    )


def _llm_call(
        prompt: str,
        job_id: str = "",
        system_prompt: str | None = None,
        model_hint: str = _DEFAULT_DOC_MODEL_HINT,
) -> tuple:
    """
    Call model_router.generate and return (raw_text, llm_meta).
    Raises on failure.
    """
    from models.model_router import model_router
    full_prompt = f"{system_prompt}\n\n{prompt}" if system_prompt else prompt
    # Guard: warn when prompt is very large. Claude Sonnet = 200K tokens ≈ 150K chars.
    # Model router auto-promotes to Gemini (1M tokens) above 80% of tier window,
    # but log a warning so we can diagnose failures.
    _prompt_len = len(full_prompt)
    if _prompt_len > 120000:
        logger.warning(
            f"[docgen] worker large prompt detected | job={job_id} "
            f"prompt_len={_prompt_len} model_hint={model_hint!r} — "
            f"model_router will auto-promote tier if needed"
        )
    logger.info(
        f"[docgen] worker calling model_router.generate | job={job_id} "
        f"prompt_len={_prompt_len} model_hint={model_hint!r}"
    )
    result   = model_router.generate(full_prompt, model_hint=model_hint, return_meta=True)
    raw      = (result.get("text") or "").strip()
    llm_meta = result.get("meta") or {}
    if not raw:
        raise RuntimeError(
            f"LLM returned empty response (prompt_len={_prompt_len}, "
            f"model={llm_meta.get('model','unknown')}) — "
            f"possible context overflow, timeout, or model error"
        )
    if raw.startswith("Error"):
        raise RuntimeError(f"LLM call failed: {raw}")
    logger.info(f"[docgen] worker LLM response | job={job_id} raw_len={len(raw)} model={llm_meta.get('model')}")
    return raw, llm_meta


def _refine_sections(job_id: str, fmt: str, question: str, title: str,
                     sections: list, model_hint: str) -> list:
    """Self-critique + refine pass over the first-draft sections.

    Takes the first-draft sections and asks the authoring model to critique and
    improve them: deepen thin content, tighten prose, fix flow/transitions, and
    fill gaps the request implies — WITHOUT inventing facts or bolting on
    unrequested boilerplate (no forced Executive Summary/Conclusion). Preserves
    the section order and the JSON schema so downstream rendering is unchanged.

    Returns the refined sections, or the original draft if refinement fails or
    returns something implausible (fewer/no sections)."""
    import json as _json
    try:
        draft_json = _json.dumps({"sections": sections}, ensure_ascii=False)[:60000]
    except Exception:  # noqa: BLE001
        return sections

    fmt_label = {"pptx": "presentation", "docx": "Word document", "pdf": "report"}.get(fmt, "document")
    critique_prompt = (
        "You are a world-class editor improving a DRAFT document to publication "
        f"quality. The document is a {fmt_label} titled \"{title}\".\n\n"
        "Improve the draft below by:\n"
        "- Deepening thin or superficial sections with substantive, specific content.\n"
        "- Improving clarity, flow and transitions between ideas.\n"
        "- Making bullets parallel and concrete; removing filler and repetition.\n"
        "- Ensuring the document fully answers the user's request.\n"
        "STRICT CONSTRAINTS:\n"
        "- Do NOT invent facts, statistics, names or figures that aren't supported.\n"
        "- Do NOT add unrequested boilerplate sections (no generic Executive "
        "Summary/Conclusion unless the draft already has one or the user asked).\n"
        "- Keep the SAME section order and the SAME JSON schema.\n"
        "- Output ONLY the JSON object {\"sections\":[...]} — no fences, no prose.\n\n"
        f"USER'S ORIGINAL REQUEST:\n{(question or '').strip()[:1500]}\n\n"
        f"DRAFT:\n{draft_json}\n\nIMPROVED JSON:"
    )
    logger.info(f"[docgen] worker refine pass START | job={job_id} sections={len(sections)}")
    raw, _meta = _llm_call(critique_prompt, job_id=job_id, model_hint=model_hint)
    refined = _parse_llm_json(raw, job_id=job_id)
    new_sections = refined.get("sections") if isinstance(refined, dict) else None
    if isinstance(new_sections, list) and len(new_sections) >= max(1, len(sections) - 1):
        logger.info(f"[docgen] worker refine pass DONE | job={job_id} sections={len(new_sections)}")
        return new_sections
    logger.warning(f"[docgen] worker refine pass produced implausible output — keeping draft | job={job_id}")
    return sections


def _llm_structure(job_id: str, fmt: str, question: str, override_prompt: str | None = None,
                   on_section=None, on_title=None, user_model_hint: str | None = None,
                   xlsx_allow_testdata: bool = False):
    """
    Call LLM to produce a sections list from a free-text question.

    For large free-form prompts (≥ _SPLIT_THRESHOLD numbered items), uses a
    two-pass strategy to avoid API gateway timeouts:
      Pass 1 → title + domain + first half of sections
      Pass 2 → remaining sections via a continuation call

    override_prompt: if provided, skips all prompt builders and uses this prompt
    directly. Used by the preservation path (file conversion) to pass
    _build_preservation_prompt output without hallucination.

    xlsx_allow_testdata: when True, an xlsx request is unconditionally run
    through the CSV test-data classifier first (a probe call) so a bulk
    tabular dataset ("500 employee records, tabular only") can be generated
    programmatically. Reserved for callers that already know the request is
    dataset-shaped (e.g. editing an uploaded/prior .csv/.xlsx data file — see
    the _src_is_data_file call site). Defaults to False: the probe is instead
    triggered on-demand by _wants_flat_dataset(question), which recognises
    explicit dataset language (row/record counts, "sample/synthetic data",
    "tabular only", etc.).

    PERF NOTE: previously this defaulted to True, so every fresh/from-scratch
    xlsx request — including plain narrative reports/plans, which are the
    common case — paid for the CSV probe. Because the probe's `csv_mode` tag
    was then discarded in favour of re-deriving the mode from column presence
    (see the fallback branch below), a narrative xlsx request always threw the
    probe's response away and re-prompted with _build_xlsx_prompt — two full
    LLM round-trips for one document. Defaulting to False plus a good
    _wants_flat_dataset() signal collapses this back to a single call for the
    common narrative case, while still detecting explicit dataset asks.

    When False — i.e. the request edits/expands an EXISTING xlsx (uploaded
    file or prior generated report as context) — we KEEP the narrative
    McKinsey report format and never flatten it into a raw dataset, UNLESS the
    user explicitly asks for a plain/flat/tabular data sheet. This prevents
    "add 20 rows and a column to my report" from converting a styled
    analytical workbook into a bare data dump.

    Returns (sections_or_slides, llm_meta, llm_title) on success, or None on failure.
    llm_meta is augmented with a 'domain' key extracted from the LLM JSON
    response (for DOCX/PDF palette selection).
    llm_title is the LLM-generated document title (str), or "" if not present.
    """
    logger.info(f"[docgen] worker _llm_structure START | job={job_id} fmt={fmt} override={bool(override_prompt)}")
    is_pptx = fmt in ("pptx", "ppt", "powerpoint", "presentation", "slides")

    _csv_prompt_used = False

    # ── If override_prompt provided, skip all prompt builders ────────────────
    if override_prompt:
        struct_prompt = override_prompt
        logger.info(f"[docgen] worker using override_prompt ({len(struct_prompt)} chars) | job={job_id}")
    else:
        # ── Detect large free-form prompts that need multi-pass chunked splitting ──
        requested_count = len(re.findall(r"(?<!\d)\d+\)", question))
        use_split = (not is_pptx) and (requested_count > _SPLIT_THRESHOLD)
        logger.info(f"[docgen] worker section detection: requested_count={requested_count}, "
                    f"threshold={_SPLIT_THRESHOLD}, use_split={use_split} | job={job_id}")

        if use_split:
            logger.info(f"[docgen] worker large free-form prompt detected ({requested_count} sections) "
                        f"— using multi-pass chunked split ({_MAX_SECTIONS_PER_PASS} sections/pass) | job={job_id}")
            return _llm_structure_split(job_id, fmt, question, requested_count,
                                        user_model_hint=user_model_hint)

    # ── Standard single-pass path ─────────────────────────────────────────────
    if not override_prompt:
        if is_pptx:
            logger.info(f"[docgen] worker using PPTX prompt builder | job={job_id}")
            struct_prompt = _build_pptx_prompt(question)
        elif fmt in ("docx", "word", "doc", "pdf"):
            # ── OLD _build_docx_prompt / _build_pdf_prompt DISABLED ──────────
            # The legacy builders hardcoded an "Executive Summary" first section
            # and a "Closing Insight" callout in the JSON schema. That forced
            # every server-generated DOCX/PDF into the old template — which did
            # NOT match the platform skill output the local test
            # runner (test_premium_design_docs.py) produced.
            #
            # We now use the same free-form, structure-preserving prompt as the
            # local test runner so server output matches local output exactly:
            # natural topic-driven headings, no forced sections, no forced
            # callout labels. The skill itself (_skill_generate) handles all
            # premium styling — this prompt only structures the content.
            logger.info(
                f"[docgen] worker using FREE-FORM prompt (skill-aligned) | "
                f"job={job_id} fmt={fmt}"
            )
            struct_prompt = _build_freeform_prompt(question, fmt)
        elif fmt == "csv":
            # CSV gets a flat column-spec prompt — NOT the multi-sheet xlsx
            # prompt — so the downstream code-writer emits df.to_csv() instead
            # of an Excel workbook with sections + tables.
            logger.info(f"[docgen] worker using CSV prompt builder | job={job_id}")
            struct_prompt = _build_csv_prompt(question)
            _csv_prompt_used = True
        elif fmt in ("xlsx", "excel", "spreadsheet"):
            # XLSX has TWO shapes:
            #   • bulk tabular DATASET ("500 employee records, tabular only") →
            #     generate programmatically via the CSV test-data classifier
            #     (exact row count, all columns).
            #   • narrative/analytical REPORT (McKinsey-style sections + tables)
            #     → _build_xlsx_prompt.
            #
            # We only probe the test-data classifier when the question
            # EXPLICITLY signals a bulk dataset (_wants_flat_dataset — strong
            # phrases or an explicit row/record count), OR the caller already
            # knows this is dataset-shaped (xlsx_allow_testdata=True, set only
            # when editing/expanding an uploaded/prior .csv/.xlsx DATA file).
            # Everything else — including plain from-scratch requests like
            # "prepare a plan/report as an excel sheet" — skips the probe
            # entirely and goes straight to the narrative _build_xlsx_prompt
            # below, avoiding a wasted extra LLM round-trip (see the
            # xlsx_allow_testdata docstring on _llm_structure for why this
            # used to run on every fresh xlsx request).
            _explicit_dataset = _wants_flat_dataset(question)
            _use_testdata_probe = xlsx_allow_testdata or _explicit_dataset
            if _use_testdata_probe:
                logger.info(
                    f"[docgen] worker using CSV test-data classifier for XLSX | "
                    f"job={job_id} allow_testdata={xlsx_allow_testdata} "
                    f"explicit_dataset={_explicit_dataset}"
                )
                struct_prompt = _build_csv_prompt(question)
                _csv_prompt_used = True
            else:
                logger.info(
                    f"[docgen] worker XLSX edit/expand of existing report → keeping "
                    f"narrative xlsx format (no test-data probe) | job={job_id}"
                )
                struct_prompt = _build_xlsx_prompt(question)
        else:
            fmt_label = {
                "pptx":         "PowerPoint presentation (slides)",
                "ppt":          "PowerPoint presentation (slides)",
                "powerpoint":   "PowerPoint presentation (slides)",
                "presentation": "PowerPoint presentation (slides)",
                "slides":       "PowerPoint presentation (slides)",
                "docx":         "Word document",
                "word":         "Word document",
                "doc":          "Word document",
                "pdf":          "PDF report",
                "xlsx":         "Excel spreadsheet",
                "excel":        "Excel spreadsheet",
                "spreadsheet":  "Excel spreadsheet",
                "txt":          "plain text document",
                "text":         "plain text document",
                "md":           "Markdown document",
                "markdown":     "Markdown document",
                "csv":          "CSV data file",
            }.get(fmt, "document")
            struct_prompt = (
                f"You are a professional document author. "
                f"The user wants a {fmt_label} about: {question}\n\n"
                f"Respond with ONLY valid JSON — no markdown fences, no explanation:\n"
                f'{{"title":"<title>","domain":"default","sections":['
                f'{{"heading":"<heading>","content":"<2-3 paragraph content>","bullets":["<item1>","<item2>","<item3>"],"level":2}}'
                f"]}}\n\n"
                f"Guidelines:\n"
                f"- 4-6 sections with substantive paragraph content\n"
                f"- Title should be concise and professional\n"
                f"- Do NOT wrap the JSON in ```json``` or any other fence"
            )

    model_hint = _resolve_doc_model_hint(user_model_hint)
    logger.info(
        f"[docgen] worker _llm_structure model resolution | job={job_id} "
        f"user_choice={user_model_hint!r} → effective={model_hint!r}"
    )

    try:
        # Stream tokens for live preview when a caller subscribes; otherwise
        # use the non-streaming path (cheaper end-to-end for batch/test callers).
        if on_section is not None or on_title is not None:
            from agents.doc_generator_agent import _llm_call_stream as _stream
            raw, llm_meta = _stream(
                struct_prompt, context=f"doc:{job_id}",
                on_section=on_section, on_title=on_title,
                model_hint=model_hint,
            )
            if raw.startswith("Error"):
                raise RuntimeError(f"LLM call failed: {raw}")
        else:
            raw, llm_meta = _llm_call(struct_prompt, job_id=job_id, model_hint=model_hint)

        struct = _parse_llm_json(raw, job_id=job_id)

        # Extract domain from LLM response and attach to llm_meta so callers
        # can pass it to generate() for palette selection.
        domain = (struct.get("domain") or "").strip().lower() or None
        llm_meta["domain"] = domain
        logger.info(f"[docgen] worker domain resolved from LLM response: {domain!r} | job={job_id}")

        # Extract LLM-generated title — sanitize to reject raw-prompt leaks
        raw_llm_title = (struct.get("title") or "").strip()
        llm_title = _sanitize_llm_title(raw_llm_title, question)
        logger.info(f"[docgen] worker LLM-generated title: {_safe_log(llm_title)} (raw={_safe_log(raw_llm_title)}) | job={job_id}")

        if is_pptx:
            slides = struct.get("slides") or []
            logger.info(f"[docgen] worker _llm_structure DONE (pptx) | job={job_id} slides={len(slides)}")
            return slides, llm_meta, llm_title
        else:
            sections = struct.get("sections") or []
            # CSV special-case: the structuring LLM classifies the request via
            # `csv_mode` and returns EITHER {columns, row_count} (test_data) OR
            # real informational `sections` (plain_text).
            #
            #  - test_data → wrap the column spec as a synthetic `csv_schema`
            #    section the code-writer materialises into N fabricated rows.
            #  - plain_text → keep the real sections; the caller writes them
            #    verbatim to the .csv as text (one cell per line), so the
            #    downloaded file matches the live preview and contains NO
            #    fabricated/tabular data.
            #
            # When an override_prompt was used (file- or chat-preservation),
            # the structuring LLM never sees the csv schema, so `csv_mode` is
            # absent and `columns` is empty — we default to plain_text, which
            # is exactly what a "convert this reply/file to CSV" request wants.
            # CSV *and* XLSX run through the same test-data classifier
            # (_build_csv_prompt). csv_mode is resolved from whether the model
            # produced a column spec; only a column spec can be materialised into
            # synthetic tabular data. Guarded by _csv_prompt_used so an xlsx
            # request that used the narrative _build_xlsx_prompt (edit/expand of
            # an existing report) skips this and keeps its real `sections`.
            if _csv_prompt_used and fmt in ("csv", "xlsx", "excel", "spreadsheet"):
                _csv_cols  = struct.get("columns") or []
                _csv_rows  = struct.get("row_count")
                _csv_mode  = (struct.get("csv_mode") or "").strip().lower()
                # Resolve mode robustly. The ONLY way to render synthetic tabular
                # data is test_data WITH a column spec, so:
                #   - columns present  → test_data (regardless of a stray tag)
                #   - no columns       → plain_text (nothing to fabricate),
                #     even if the LLM tagged it "test_data" — this prevents the
                #     code-writer from being handed an empty schema.
                _csv_mode = "test_data" if _csv_cols else "plain_text"

                # ── XLSX narrative fallback ──────────────────────────────────
                # An xlsx request with NO column spec is an analytical/report
                # workbook, not a dataset. Re-run the original narrative xlsx
                # prompt and fall through to the standard sections return so
                # report-style multi-sheet workbooks are unchanged.
                if fmt != "csv" and _csv_mode != "test_data":
                    logger.info(
                        f"[docgen] worker XLSX classified plain_text (no columns) → "
                        f"falling back to narrative xlsx prompt | job={job_id}"
                    )
                    raw2, llm_meta2 = _llm_call(
                        _build_xlsx_prompt(question), job_id=job_id, model_hint=model_hint)
                    struct = _parse_llm_json(raw2, job_id=job_id)
                    # Merge cost/meta from the fallback call. The probe call's
                    # tokens/cost are real spend and must be ACCUMULATED, not
                    # overwritten — a plain `.update()` here silently dropped
                    # the discarded probe call's cost from billing/accounting.
                    # `model`/`domain` reflect the call whose output we keep
                    # (the fallback), so those fields are overwritten as before.
                    _probe_meta = llm_meta or {}
                    llm_meta = dict(llm_meta2 or {})
                    for _k in ("tokens", "in_tok", "out_tok"):
                        llm_meta[_k] = int(_probe_meta.get(_k) or 0) + int(llm_meta.get(_k) or 0)
                    llm_meta["cost_usd"] = float(_probe_meta.get("cost_usd") or 0.0) + float(llm_meta.get("cost_usd") or 0.0)
                    domain = (struct.get("domain") or "").strip().lower() or domain
                    llm_meta["domain"] = domain
                    raw_llm_title2 = (struct.get("title") or "").strip()
                    if raw_llm_title2:
                        llm_title = _sanitize_llm_title(raw_llm_title2, question)
                    sections = struct.get("sections") or []
                    logger.info(
                        f"[docgen] worker _llm_structure DONE (xlsx narrative) | "
                        f"job={job_id} sections={len(sections)} domain={domain!r}"
                    )
                    return sections, llm_meta, llm_title

                llm_meta["csv_mode"]      = _csv_mode
                llm_meta["csv_columns"]   = _csv_cols
                llm_meta["csv_row_count"] = _csv_rows
                if _csv_mode == "test_data" and _csv_cols:
                    sections = [{
                        "heading": "csv_schema",
                        "csv_columns":  _csv_cols,
                        "csv_row_count": _csv_rows,
                    }]
                # else: plain_text — leave `sections` as the real content.
                logger.info(
                    f"[docgen] worker _llm_structure DONE ({fmt}) | job={job_id} "
                    f"mode={_csv_mode} columns={len(_csv_cols)} "
                    f"row_count={_csv_rows} sections={len(sections)} domain={domain!r}"
                )
                return sections, llm_meta, llm_title
            # ── Multi-pass authoring: self-critique + refine ─────────────────
            # Elevates quality toward Claude/GPT. Skipped for faithful paths
            # (override_prompt = preservation/convert/summarize) and data formats
            # (csv/xlsx), where invention/rewriting is undesirable.
            _refine_ok = (
                _DOCGEN_REFINE_ENABLED
                and not override_prompt
                and fmt in ("docx", "word", "doc", "pdf", "txt", "text", "md", "markdown")
                and sections
            )
            if _refine_ok:
                try:
                    sections = _refine_sections(job_id, fmt, question, llm_title,
                                                sections, model_hint)
                except Exception as _rerr:  # noqa: BLE001
                    logger.warning(f"[docgen] worker refine pass failed (using draft): {_rerr}")
            logger.info(f"[docgen] worker _llm_structure DONE | job={job_id} sections={len(sections)} domain={domain!r}")
            return sections, llm_meta, llm_title
    except Exception as exc:
        logger.error(f"[docgen] worker _llm_structure FAILED | job={job_id} error={exc}", exc_info=True)
        _fail(job_id, f"Content structuring failed: {exc}")
        return None


def _llm_structure_split(job_id: str, fmt: str, question: str, requested_count: int,
                         user_model_hint: str | None = None):
    """
    Multi-pass chunked LLM document generation for large free-form prompts.

    Splits sections into chunks of _MAX_SECTIONS_PER_PASS to avoid API gateway
    timeouts and token limits.  A 20-section document with chunk-size 6 becomes
    4 passes (6 + 6 + 6 + 2).

    **Performance strategy (parallel execution):**
      Pass 1  → sequential — extracts title + domain (needed by later passes).
      Pass 2…N → **concurrent** via ThreadPoolExecutor — each pass already knows
                 the title/domain from Pass 1 and the headings produced by Pass 1.
                 They do NOT depend on each other, so they run in parallel.

    For a 5-pass document this cuts wall-clock time from ~9 min (sequential)
    to ~4 min (Pass 1 + longest-of-remaining).

    Merges all results in order and returns (sections, llm_meta, llm_title).
    """
    import concurrent.futures as _cf

    # ── Extract numbered items from the question ──────────────────────────────
    items = re.findall(r"(?<!\d)(\d+\))\s*(.*?)(?=\s*(?<!\d)\d+\)|$)", question, re.DOTALL)

    # Preamble = everything before the first numbered item
    preamble_match = re.search(r"(?<!\d)1\)", question)
    preamble = question[:preamble_match.start()].strip() if preamble_match else ""

    # ── Build the free-form system prompt (same as test script) ──────────────
    free_form_system = (
        "You are a document structuring assistant.\n\n"
        "Your ONLY job is to convert the user's document request into a JSON payload that will be\n"
        "rendered into a Word / PDF document. Follow the user's requested content, structure,\n"
        "headings, and tone exactly — do NOT add, remove, or rewrite sections beyond what the user asks.\n\n"
        "OUTPUT FORMAT\n"
        "Respond with ONLY valid JSON — no markdown fences, no explanation, no preamble.\n\n"
        "JSON SCHEMA (follow exactly):\n"
        "{\n"
        '  "title": "<document title derived from the user\'s request>",\n'
        '  "domain": "<single lowercase keyword: payments|banking|ai|technology|cybersecurity|legal|education|heritage|sports|esg|government|fintech|healthcare|executive|default>",\n'
        '  "sections": [\n'
        "    {\n"
        '      "heading": "<section heading as requested by the user>",\n'
        '      "subheading": "<sub-label if the user specified one, otherwise empty string>",\n'
        '      "level": 1,\n'
        '      "content": "<full section body text, paragraphs separated by \\n\\n, written exactly as the user requested>",\n'
        '      "bullets": ["<bullet point as requested>"],\n'
        '      "callout": {"label": "<short label>", "text": "<callout text if applicable, otherwise empty>"},\n'
        '      "table": null\n'
        "    }\n"
        "  ]\n"
        "}\n\n"
        "RULES\n"
        "1. Follow the user's requested structure exactly — produce every section they listed, in the same order.\n"
        "2. Write the content the user asked for — do not invent extra sections or change the scope.\n"
        "3. If the user requests a table, populate \"table\" as {\"headers\": [...], \"rows\": [[...], ...]}.\n"
        "4. Output RAW JSON ONLY. Any non-JSON text causes a parse failure.\n"
        "5. All text must be grammatically complete and correctly spelled.\n"
    )

    if not items:
        # Fallback: no numbered items found — treat as single pass
        logger.warning(f"[docgen] worker no numbered items found in question — falling back to single pass | job={job_id}")
        items_to_chunk = []
    else:
        items_to_chunk = items

    # Split items into chunks of _MAX_SECTIONS_PER_PASS
    chunks = [items_to_chunk[i:i + _MAX_SECTIONS_PER_PASS]
              for i in range(0, len(items_to_chunk), _MAX_SECTIONS_PER_PASS)]
    if not chunks:
        chunks = [items_to_chunk]  # single empty chunk triggers fallback below

    num_passes = len(chunks)
    logger.info(f"[docgen] worker multi-pass chunked split | {requested_count} sections → "
                f"{num_passes} passes of ≤{_MAX_SECTIONS_PER_PASS} sections each | job={job_id}")

    all_sections: list = []
    title    = _derive_title_from_question(question)
    domain   = None
    llm_meta: dict = {}

    # Resolve the model hint once for every pass — same precedence rules
    # as the single-pass path (_llm_structure → _resolve_doc_model_hint).
    _split_model_hint = _resolve_doc_model_hint(user_model_hint)

    # ══════════════════════════════════════════════════════════════════════════
    # PASS 1 — sequential (extracts title + domain needed by later passes)
    # ══════════════════════════════════════════════════════════════════════════
    chunk_1      = chunks[0]
    chunk_1_text = "\n".join(f"{num} {txt.strip()}" for num, txt in chunk_1)
    chunk_1_size = len(chunk_1)

    pass_1_prompt = (
        f"{preamble}\n\n"
        f"IMPORTANT: Produce ONLY the following {chunk_1_size} sections "
        f"(do NOT produce any other sections yet):\n"
        f"{chunk_1_text}"
    ).strip()
    full_prompt_1 = f"{free_form_system}\n\n{pass_1_prompt}"

    logger.info(f"[docgen] worker multi-pass | Pass 1/{num_passes}: "
                f"{chunk_1_size} sections | job={job_id}")
    _publish_progress(
        job_id, 2, 6, "Structuring Content",
        f"LLM pass 1/{num_passes} — generating sections 1–{chunk_1_size}…",
    )
    try:
        raw, pass_meta = _llm_call(full_prompt_1, job_id=job_id, model_hint=_split_model_hint)
        struct = _parse_llm_json(raw, job_id=job_id)
        title  = (struct.get("title") or "").strip() or title
        domain = (struct.get("domain") or "").strip().lower() or None
        llm_meta = pass_meta
        llm_meta["domain"] = domain
        new_sections = struct.get("sections") or []
        all_sections.extend(new_sections)
        logger.info(f"[docgen] worker Pass 1/{num_passes} done | "
                    f"job={job_id} new_sections={len(new_sections)} "
                    f"total={len(all_sections)} title={title!r} domain={domain!r}")
    except Exception as exc:
        logger.error(f"[docgen] worker Pass 1 FAILED | job={job_id} error={exc}", exc_info=True)
        _fail(job_id, f"Content structuring (pass 1) failed: {exc}")
        return None

    # ══════════════════════════════════════════════════════════════════════════
    # PASSES 2…N — PARALLEL (each only needs title + domain + Pass-1 headings)
    # ══════════════════════════════════════════════════════════════════════════
    remaining_chunks = chunks[1:]
    if not remaining_chunks:
        # Only one pass was needed — skip parallelism
        llm_title = _sanitize_llm_title(title, question)
        logger.info(f"[docgen] worker multi-pass chunked split DONE (single pass) | job={job_id} "
                    f"total_sections={len(all_sections)} title={llm_title!r}")
        return all_sections, llm_meta, llm_title

    # Build all continuation prompts upfront.
    # Each pass knows the headings from Pass 1 (for context) but does NOT
    # depend on headings from other parallel passes.
    pass_1_headings = [s.get("heading", "") for s in all_sections]
    continuation_prompts: list[tuple[int, str, int]] = []  # (pass_idx, prompt, chunk_size)

    running_section_count = len(all_sections)
    for rel_idx, chunk in enumerate(remaining_chunks):
        pass_idx   = rel_idx + 2
        chunk_text = "\n".join(f"{num} {txt.strip()}" for num, txt in chunk)
        chunk_size = len(chunk)

        pass_prompt = (
            f"Document title: {title}\n"
            f"Domain: {domain}\n\n"
            f"Sections already written ({len(pass_1_headings)}):\n"
            + "\n".join(f"  - {h}" for h in pass_1_headings)
            + f"\n\nOriginal document context:\n{preamble}\n\n"
            f"Now produce ONLY the following {chunk_size} sections:\n"
            f"{chunk_text}"
        )
        full_prompt = f"{_CONTINUATION_SYSTEM_PROMPT}\n\n{pass_prompt}"
        continuation_prompts.append((pass_idx, full_prompt, chunk_size))
        running_section_count += chunk_size

    _publish_progress(
        job_id, 2, 6, "Structuring Content",
        f"LLM passes 2–{num_passes} running in parallel ({len(remaining_chunks)} concurrent calls)…",
    )
    logger.info(f"[docgen] worker launching passes 2–{num_passes} in parallel | job={job_id}")

    def _run_pass(args: tuple) -> tuple:
        """Execute a single continuation pass. Returns (pass_idx, sections, meta, error)."""
        p_idx, prompt, c_size = args
        try:
            raw_text, p_meta = _llm_call(prompt, job_id=job_id, model_hint=_split_model_hint)
            p_struct = _parse_llm_json(raw_text, job_id=job_id)
            sections = p_struct.get("sections") or []
            logger.info(f"[docgen] worker Pass {p_idx}/{num_passes} done (parallel) | "
                        f"job={job_id} new_sections={len(sections)}")
            _publish_progress(
                job_id, 2, 6, "Structuring Content",
                f"LLM pass {p_idx}/{num_passes} complete — {len(sections)} sections received",
            )
            return (p_idx, sections, p_meta, None)
        except Exception as e:
            logger.warning(f"[docgen] worker Pass {p_idx}/{num_passes} failed (parallel) | "
                           f"job={job_id} error={e}")
            return (p_idx, [], {}, e)

    # Cap workers at number of remaining passes (typically 2-4), never more
    # than 4 to avoid overwhelming the LLM proxy.
    max_parallel = min(len(continuation_prompts), 4)

    with _cf.ThreadPoolExecutor(max_workers=max_parallel) as pool:
        futures = pool.map(_run_pass, continuation_prompts)
        results = list(futures)  # preserves input order

    # Merge results in pass order
    for p_idx, sections, p_meta, err in results:
        if err:
            logger.warning(f"[docgen] worker Pass {p_idx}/{num_passes} had error — "
                           f"skipping its sections | job={job_id}")
            continue
        all_sections.extend(sections)
        # Accumulate token counts from all passes
        for key in ("tokens", "in_tok", "out_tok"):
            llm_meta[key] = int(llm_meta.get(key) or 0) + int(p_meta.get(key) or 0)

    llm_title = _sanitize_llm_title(title, question)
    logger.info(f"[docgen] worker multi-pass chunked split DONE | job={job_id} "
                f"total_sections={len(all_sections)} title={llm_title!r}")
    return all_sections, llm_meta, llm_title

# ── Image generation for PPTX slides ─────────────────────────────────────────

def _enrich_with_images(slides: list) -> list:
    """
    For each PPTX slide that has an image_prompt (title, content, closing),
    generate an image using Gemini Imagen or DALL-E and attach as _image_bytes.
    Fails silently — missing image means the slide uses geometric-only design.

    Per-image timeout: each fetch is capped at PPT_IMAGE_TIMEOUT_SEC (default 45s).
    This prevents a slow proxy from hanging the work-horse long enough for the OS
    to kill it with "Work-horse terminated unexpectedly; waitpid returned None".
    """
    provider = _resolve_image_provider()
    if provider == "disabled":
        return slides

    # Per-image wall-clock cap — keeps total image enrichment bounded.
    # 3 image slides × 45s = 135s max, well within the 1800s job timeout.
    _img_timeout = float(os.getenv("PPT_IMAGE_TIMEOUT_SEC", "45"))

    IMAGE_SLIDE_TYPES = _image_eligible_slide_types()
    logger.info(
        f"[docgen] worker image enrichment | all_slides={_PPT_IMAGES_ALL_SLIDES} "
        f"eligible={sorted(IMAGE_SLIDE_TYPES)}"
    )

    for slide in slides:
        stype = (slide.get("slide_type") or "content").lower()
        if stype not in IMAGE_SLIDE_TYPES:
            continue

        image_prompt = (slide.get("image_prompt") or "").strip()
        if not image_prompt:
            continue

        img_bytes = None
        try:
            import concurrent.futures as _cf
            with _cf.ThreadPoolExecutor(max_workers=1) as _pool:
                if provider == "gemini":
                    fut = _pool.submit(_fetch_gemini_image, image_prompt)
                else:
                    fut = _pool.submit(_fetch_dalle_image, image_prompt)
                try:
                    img_bytes = fut.result(timeout=_img_timeout)
                except _cf.TimeoutError:
                    logger.warning(
                        f"doc_worker: image fetch timed out after {_img_timeout}s "
                        f"for '{stype}' slide — using geometric design"
                    )
                    img_bytes = None
        except Exception as _ie:
            logger.warning(f"doc_worker: image fetch error for '{stype}' slide: {_ie}")
            img_bytes = None

        # DALL-E fallback when Gemini returned nothing (and not already tried)
        if img_bytes is None and provider == "gemini" and os.getenv("OPENAI_API_KEY"):
            try:
                import concurrent.futures as _cf
                with _cf.ThreadPoolExecutor(max_workers=1) as _pool:
                    fut = _pool.submit(_fetch_dalle_image, image_prompt)
                    try:
                        img_bytes = fut.result(timeout=_img_timeout)
                    except _cf.TimeoutError:
                        logger.warning(
                            f"doc_worker: DALL-E fallback timed out after {_img_timeout}s "
                            f"for '{stype}' slide"
                        )
                        img_bytes = None
            except Exception as _ie2:
                logger.warning(f"doc_worker: DALL-E fallback error for '{stype}' slide: {_ie2}")
                img_bytes = None

        if img_bytes:
            slide["_image_bytes"] = img_bytes
            logger.info(f"doc_worker: image generated for '{stype}' slide ({len(img_bytes)} bytes)")
        else:
            logger.info(f"doc_worker: no image for '{stype}' slide — using geometric design")

    return slides


def _fetch_ppt_image_via_proxy(prompt: str, provider: str = "auto") -> bytes | None:
    """
    Generate a PPT slide image by calling the LLM proxy's /llm/generate-ppt-image endpoint.
    The proxy (the LLM proxy server) holds GEMINI_API_KEY / OPENAI_API_KEY and handles compliance + circuit breaking.
    Returns raw image bytes or None on failure.
    """
    import base64 as _b64
    proxy_url = os.getenv("LLM_PROXY_URL", "").rstrip("/")
    if not proxy_url:
        logger.warning("doc_worker: LLM_PROXY_URL not set — cannot fetch PPT image via proxy")
        return None
    try:
        import httpx
        # Use a dedicated image timeout (default 40s), NOT LLM_TIMEOUT_SEC.
        # LLM_TIMEOUT_SEC is for text generation (can be 300s+); using it here
        # causes the work-horse to hang per-image and get killed by the OS.
        _img_req_timeout = float(os.getenv("PPT_IMAGE_TIMEOUT_SEC", "40"))
        resp = httpx.post(
            f"{proxy_url}/llm/generate-ppt-image",
            json={"provider": provider, "prompt": prompt},
            timeout=_img_req_timeout,
        )
        resp.raise_for_status()
        data = resp.json()
        b64  = data.get("image_b64")
        if b64:
            return _b64.b64decode(b64)
    except Exception as exc:
        logger.warning(f"doc_worker: PPT image via proxy failed: {exc}")
    return None


def _fetch_gemini_image(prompt: str) -> bytes | None:
    """
    Generate image via Gemini Imagen 3.
    Routes through LLM proxy (the LLM proxy server) when LLM_PROXY_URL is set (production).
    Falls back to direct gateway call in local dev (no proxy configured).
    """
    # Production path: proxy holds the API key — always use it
    proxy_url = os.getenv("LLM_PROXY_URL", "").rstrip("/")
    if proxy_url:
        return _fetch_ppt_image_via_proxy(prompt, provider="gemini")
    # Local dev fallback: call gateway directly
    try:
        from gateway_gemini import generate_image_gemini
        return generate_image_gemini(prompt)
    except Exception as exc:
        logger.warning(f"doc_worker: Gemini image direct call failed: {exc}")
    return None


def _fetch_dalle_image(prompt: str) -> bytes | None:
    """
    Generate image via DALL-E 3.
    Routes through LLM proxy (the LLM proxy server) when LLM_PROXY_URL is set (production).
    Falls back to direct gateway call in local dev (no proxy configured).
    """
    proxy_url = os.getenv("LLM_PROXY_URL", "").rstrip("/")
    if proxy_url:
        return _fetch_ppt_image_via_proxy(prompt, provider="dalle")
    # Local dev fallback: call gateway directly
    try:
        from gateway_openai import generate_image_dalle
        return generate_image_dalle(prompt)
    except Exception as exc:
        logger.warning(f"doc_worker: DALL-E image direct call failed: {exc}")
    return None


def _publish_progress(job_id: str, step: int, total_steps: int,
                      label: str, detail: str = "") -> None:
    """Publish doc-generation progress to Redis so the frontend can show a
    real-time stepper instead of a blind spinner.  The key is short-lived
    (10 min) — once ``doc:result:{job_id}`` is written the frontend
    switches to the final state and ignores progress entirely."""
    _R.setex(
        f"doc:progress:{job_id}", 600,
        json.dumps({
            "step":        step,
            "total_steps": total_steps,
            "label":       label,
            "detail":      detail,
        }),
    )
    logger.info(
        f"[docgen] worker progress | job={job_id} step={step}/{total_steps} "
        f"label={label!r}" + (f" detail={detail!r}" if detail else "")
    )


def _fail(job_id: str, error: str) -> None:
    _R.setex(
        f"doc:result:{job_id}", 3600,
        json.dumps({"status": "error", "error": error})
    )
    logger.error(f"[docgen] worker FAIL | job={job_id} error={error!r}")


def _clarify(job_id: str, question: str, options: list, *,
             original_question: str = "", fmt: str = "",
             attachment_ids: list | None = None,
             chat_id: str | None = None, user_model_hint: str = "auto",
             doc_intent: str = "") -> None:
    """Publish a 'need clarification' result so the frontend can ask the user a
    quick question (rendered as quick-reply buttons) instead of the worker
    guessing which document a fuzzy reference means. The chosen option resumes
    generation via POST /docs/clarify-resume. We echo back the full request
    context so the resume call rebuilds an equivalent job — crucially including
    chat_id (needed for artifact recall + version chaining). Mirrors _fail."""
    _R.setex(
        f"doc:result:{job_id}", 3600,
        json.dumps({
            "status": "clarify",
            "question": question or "Which document did you mean?",
            "options": options or [],
            "resume": {
                "question":        original_question,
                "format":          fmt,
                "attachment_ids":  attachment_ids or [],
                "chat_id":         chat_id or "",
                "user_model_hint": user_model_hint or "auto",
                "doc_intent":      doc_intent or "",
            },
        })
    )


def _save_audit(
        *,
        file_id: str,
        job_id: str,
        user_id: str,
        chat_id,
        fmt: str,
        title: str,
        filename: str,
        file_path: str,
        content_md: str,
        artifact_id: str | None = None,
        version: int | None = None,
) -> None:
    """Insert the GeneratedDocument audit row that backs /docs/download/{file_id}.

    `artifact_id` + `version` are optional for chat-doc callers (they default to
    NULL / 1 in the DB), but REQUIRED for the cowork skill worker path which
    passes them so successive build_document calls form a version chain
    (Canvas/Pages parity). Previously this function did NOT accept those
    kwargs — doc_skill_worker.py passed them, raised TypeError, the outer
    try/except swallowed it, no row was written, and /docs/download/{file_id}
    returned 404 ({"detail":"Document not found"} → the 1KB "corrupt PDF"
    symptom in CoworkDesktop).
    """
    try:
        from db.database import SessionLocal
        from db.models import GeneratedDocument
        db = SessionLocal()
        try:
            row = GeneratedDocument(
                id=file_id,
                job_id=job_id,
                user_id=user_id,
                chat_id=chat_id or None,
                format=fmt,
                title=title,
                filename=filename,
                file_path=file_path,
                content_md=content_md,
            )
            # Only set artifact versioning fields when provided so we don't
            # override the column defaults (artifact_id NULL, version 1) for
            # legacy chat-doc rows.
            if artifact_id is not None:
                row.artifact_id = artifact_id
            if version is not None:
                row.version = version
            db.add(row)
            db.commit()
        finally:
            db.close()
        # Mirror into chat_attachments so this generated doc can be RE-FED as
        # input later ("combine last week's report with this new PDF").
        try:
            from services.doc_context import mirror_generated_doc_as_attachment
            mirror_generated_doc_as_attachment(
                doc_id=file_id, chat_id=chat_id, user_id=user_id,
                title=title, fmt=fmt, file_path=file_path, content_md=content_md,
            )
        except Exception as _merr:  # noqa: BLE001
            logger.warning(f"doc_worker: attachment mirror failed for {file_id}: {_merr}")
    except Exception as exc:
        logger.warning(f"doc_worker: audit save failed for {file_id}: {exc}")


def _save_md_session_for_chat(
        *,
        chat_id: str,
        job_id: str,
        file_id: str,
        title: str,
        domain,
        sections: list,
        content_md: str,
        filename: str,
        file_path: str,
        original_format: str,
        question: str,
        llm_meta: dict,
) -> None:
    """
    Save an MD session to Redis after any doc format generation.

    This enables @edit_doc follow-up edits regardless of the original format.
    When the user later calls @edit_doc, generate_md_job loads this session,
    applies the edit, and regenerates the file in original_format.

    Session key: md:session:{chat_id}  (TTL 24 h — matches RESULT_TTL)
    """
    if not chat_id:
        return
    try:
        session = {
            "schema_version": "1.0",
            "chat_id":        chat_id,
            "document": {
                "title":             title,
                "domain":            domain,
                "file_id":           file_id,
                "filename":          filename,
                "file_path":         file_path,
                "word_count":        len(content_md.split()),
                "section_count":     len(sections),
                "content_snapshot":  content_md,
                # Preserved so edit flow can regenerate in the same format
                "original_format":   original_format,
                "original_question": question,
            },
            "sections":     sections,
            "conversation": [
                {
                    "turn":    1,
                    "role":    "user",
                    "type":    "generate",
                    "content": question,
                },
                {
                    "turn":    1,
                    "role":    "assistant",
                    "type":    "generate",
                    "content": (
                        f"Generated document: {filename} "
                        f"({len(sections)} sections)"
                    ),
                    "meta": {
                        "model":    llm_meta.get("model"),
                        "tokens":   llm_meta.get("tokens"),
                        "cost_usd": llm_meta.get("cost_usd"),
                        "latency":  llm_meta.get("latency"),
                    },
                },
            ],
            "llm_meta": {
                "model":          llm_meta.get("model"),
                "total_tokens":   int(llm_meta.get("tokens") or 0),
                "total_cost_usd": float(llm_meta.get("cost_usd") or 0.0),
                "total_calls":    1,
            },
        }
        _R.setex(
            f"md:session:{chat_id}",
            RESULT_TTL,
            json.dumps(session, ensure_ascii=False),
        )
        logger.info(
            f"[docgen] worker MD session saved | chat_id={chat_id!r} "
            f"fmt={original_format!r} job={job_id}"
        )
    except Exception as exc:
        logger.warning(
            f"[docgen] worker MD session save failed for chat_id={chat_id!r}: {exc}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Re-export: workers/doc_skill_worker.py imports purge_expired_docs from this
# module. The actual implementation now lives in workers/purge_worker.py —
# doc_purge.py, image_purge.py, and upload_purge.py were consolidated into it
# (see purge_worker.py's module docstring). This alias was left pointing at
# the old, now-nonexistent workers.doc_purge module, so every single job in
# this worker logged a spurious "doc_purge re-export failed: No module named
# 'workers.doc_purge'" warning (see the [docgen] log). Pointed at the correct
# module so doc_skill_worker's existing import contract stays valid without
# the per-job warning noise.
# ─────────────────────────────────────────────────────────────────────────────
try:
    from workers.purge_worker import purge_expired_docs  # noqa: E402,F401
except Exception as _purge_exc:  # pragma: no cover — keep import-safe
    logger.warning(f"[docgen] worker purge_worker re-export failed: {_purge_exc}")

    def purge_expired_docs():  # type: ignore[override]
        """Fallback no-op when workers.purge_worker cannot be imported."""
        return None
